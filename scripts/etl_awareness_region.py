#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
지역별 인지도(기상도) ETL
========================
입력 : data/dialect_gisangdo/*.xlsx  (2024년 차 어휘 조사 원자료, 제보자 1명 = 파일 1개)
출력 : data/processed/awareness_by_region.json

파일명 규약 : {지역2}{연차2}{세대2}{성별1}VE.xlsx  예) CB2420FVE = 충북·24년차·20대·여성

원자료 구조
-----------
한 행 = (제보자, 항목, 어형) 하나. J열 '사용도/인지도'의 1~4가 그 **어형**에 대한 등급이다.
  1 사용(현재 일상에서 씀) / 2 이해(써 봤으나 지금은 안 씀) / 3 인지(들어는 봄) / 4 무지(처음 들음)
'방언형(어절)'이 '*'이거나 비어 있으면 조사자가 제시만 하고 제보자가 발화하지 않은 형태다.

상태 판정 (지역 단위)
--------------------
  w1~w4 : 지역어형에 등급이 있는 경우. 사용률(등급1 비율)을 기상도 임계값에 태운다.
  std   : 조사는 됐으나 응답이 전부 표준어형 → '표준어권'(제5상태). 결측이 아니다.
  w0    : 조사 자체가 없거나, 지역어형 행은 있는데 등급이 비어 판단 불가 → '관측 없음'.
"""
import collections, glob, io, json, os, re, sys

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC  = os.path.join(BASE, 'data', 'dialect_gisangdo')
OUT  = os.path.join(BASE, 'data', 'processed', 'awareness_by_region.json')

REGION_ORDER = ['GG', 'GW', 'CB', 'CN', 'JB', 'JN', 'GB', 'GN', 'JJ']
REGION_NAMES = {'GG': '경기', 'GW': '강원', 'CB': '충북', 'CN': '충남', 'JB': '전북',
                'JN': '전남', 'GB': '경북', 'GN': '경남', 'JJ': '제주'}
VALID = {'1', '2', '3', '4'}

# 원자료에 전각 숫자('１')로 찍힌 칸이 있다. 같은 숫자의 다른 코드포인트일 뿐이라
# 반각으로 맞추는 것은 값의 재해석이 아니다. 그 밖의 값(빈칸·'*'·'1,4' 같은 복수기입)은
# 손대지 않고 '등급 없음'으로 남긴다 — 어느 등급인지는 사람이 원본을 고쳐야 한다.
_FULLWIDTH = str.maketrans('１２３４', '1234')


def grade_of(v):
    """엑셀 칸 값 → 등급 문자열('1'~'4') 또는 None. 정규화 규칙의 유일한 출처."""
    if v in (None, ''):
        return None
    g = str(v).strip().translate(_FULLWIDTH)
    return g if g in VALID else None


# 표제어 → dialect_gisangdo.html WORDLIST 표기 (동음이의 구분 괄호)
WORD_ALIAS = {'가': '가(邊)', '새끼': '새끼(繩)', '아우 타다': '아우타다', '키': '키(箕)'}

norm = lambda s: re.sub(r'[:\s]', '', re.sub(r'\([^)]*\)', '', s or ''))


def head_forms(hw):
    """표제어의 표준형 집합. '할머니(호칭)'→{할머니}, '서 되/세 되'→{서되, 세되}."""
    return {norm(p) for p in re.split(r'[/·]', hw or '') if norm(p)}


QC = {'files': 0, 'rowsTotal': 0, 'rowsBadCode': 0, 'gradeFilled': 0, 'skipped': []}


def _norm(h):
    return re.sub(r'\s+', '', str(h or ''))


def _hdr(h):
    """열 찾기용 머리글 — 2022·2023 자료는 '(수정)' 같은 꼬리가 붙는다.

    _norm 은 서식 종류를 세는 데도 쓰므로 그대로 두고, 열을 찾을 때만 꼬리를 뗀다.
    """
    return re.sub(r'\((수정|보충|최종|검수)\)$', '', _norm(h))


def resolve_columns(ws):
    """원자료 서식이 13가지로 제각각이라 열 위치를 헤더에서 찾는다.
       (충남 20대 여·70대 여 파일은 타임코드 열이 없어 항목번호가 1번째에 온다.)"""
    try:
        hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return None
    idx = {_hdr(h): i for i, h in enumerate(hdr) if h}

    def find(*names):
        for n in names:
            if n in idx:
                return idx[n]
        return None

    col = {
        'code':  find('항목번호'),
        'head':  find('표제어', '표준어형', '표제어형', '대응표준어'),
        'base':  find('방언형(기저형)'),
        'grade': find('인지도/사용도', '사용도/인지도'),
    }
    return None if col['code'] is None or col['grade'] is None else col


def cell(row, i):
    return row[i] if i is not None and i < len(row) else None


def load_records():
    import openpyxl
    recs = []
    # ~$ 로 시작하는 파일은 엑셀이 열려 있을 때 생기는 잠금 파일이라 제외한다
    files = sorted(f for f in glob.glob(os.path.join(SRC, '*.xlsx'))
                   if not os.path.basename(f).startswith('~$'))
    if not files:
        sys.exit('원자료를 찾지 못했습니다: %s' % SRC)
    layouts = collections.defaultdict(list)   # 서식(열 구성)이 몇 가지인지
    for path in files:
        f = os.path.basename(path)
        m = re.match(r'([A-Z]{2})(\d{2})(\d{2})([MF])VE', f)
        if not m:
            print('  건너뜀(파일명 규약 불일치):', f)
            continue
        rg, yr, age, sx = m.groups()
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        col = resolve_columns(ws)
        try:
            hdr0 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            layouts['|'.join(_norm(h) for h in hdr0)].append(f)
        except StopIteration:
            pass
        if col is None:
            print('  건너뜀(열 구성을 알 수 없음):', f)
            QC['skipped'].append(f)
            wb.close()
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = cell(row, col['code'])
            if code is None:
                continue
            QC['rowsTotal'] += 1
            mm = re.match(r'^(\d{5})', str(code).strip())
            if not mm:          # 항목번호 칸에 표제어가 들어간 오류 행
                QC['rowsBadCode'] += 1
                continue
            pres = str(cell(row, col['head'])).strip() if cell(row, col['head']) else ''
            base = str(cell(row, col['base'])).strip() if cell(row, col['base']) else ''
            form = base if base and base != '*' else pres
            g = grade_of(cell(row, col['grade']))
            if g:
                QC['gradeFilled'] += 1
            recs.append({'rg': rg, 'year': yr, 'age': int(age), 'sx': sx,
                         'it': mm.group(1), 'pres': pres, 'form': form, 'g': g,
                         # 화면이 '그 지역에서 조사된 표준어형·방언형' 을 그대로 보여줄 수 있게
                         # 두 칸을 원문대로 싣는다. form 은 둘을 합친 파생값이라 대신 못 쓴다.
                         'base': base})
        wb.close()
    QC['files'] = len(files)
    # 원자료 서식이 제각각이라 열 위치를 헤더로 찾는다. 몇 가지였는지 검수 화면에 알린다.
    QC['layouts'] = len(layouts)
    common = max(layouts.values(), key=len) if layouts else []
    QC['layoutOdd'] = sorted(f for v in layouts.values() if v is not common for f in v)
    return recs, len(files)


# 사용도/인지도 등급 → 점수. 조사표의 4등급이 그대로 척도가 된다.
#   ① 사용 100 · ② 이해 75 · ③ 인지 50 · ④ 무지 25
GRADE_SCORE = {1: 100, 2: 75, 3: 50, 4: 25}

# 상태 경계 = 등급 사이의 중간값. 임의로 정한 값이 아니라 등급 척도에서 나온다.
#   w1 등급1~1.5 · w2 1.5~2.5 · w3 2.5~3.5 · w4 3.5~4
SCORE_BANDS = [(87.5, 'w1'), (62.5, 'w2'), (37.5, 'w3')]


def region_score(grades):
    """제보자별 등급 → 지역 점수(0~100). 등급이 없으면 None."""
    g = [int(x) for x in grades if int(x) in GRADE_SCORE]
    if not g:
        return None
    return round(sum(GRADE_SCORE[x] for x in g) / len(g), 2)


def weather_of(score):
    """점수 → 상태. dialect_gisangdo.html weatherOf() 와 같아야 한다."""
    if score is None:
        return 'w0'
    for lo, st in SCORE_BANDS:
        if score >= lo:
            return st
    return 'w4'


WEATHER_DB = os.path.join(BASE, 'data', 'gisangdo.db')


def load_records_from_db(db_path=None, year=None):
    """전용 테이블(wb_weather_response) → build_output 이 받는 레코드 형태.

    엑셀을 다시 읽지 않고 DB 를 원천으로 삼는 경로. 관리자에서 원자료를 올린 뒤
    정적 JSON 을 다시 뽑을 때(--from-db) 와 프론트 API(server.py) 가 함께 쓴다.

    year: 조사 연차(파일명 2자리, 예 '24'). 주면 그 연차만 읽는다.
      연차를 섞으면 '제보자 38명' 같은 수치가 서로 다른 조사의 합이 되어 뜻을 잃는다.
      그래서 연차는 화면 필터가 아니라 원천을 가르는 조건으로 다룬다.
    """
    import sqlite3
    path = db_path or WEATHER_DB
    if not os.path.exists(path):
        raise FileNotFoundError('기상도 DB가 없습니다: %s' % path)
    yr = re.sub(r'\D', '', str(year or ''))[-2:]        # '2024' · '24' 모두 받는다
    con = sqlite3.connect(path)
    con.row_factory = sqlite3.Row
    try:
        where = "r.item_base IS NOT NULL AND f.use_yn = 'Y' AND r.use_yn = 'Y'"
        params = []
        if yr:
            where += ' AND f.research_degree = ?'
            params.append(yr)
        rows = con.execute(
            """SELECT r.response_id rid,
                      f.region_cd rg, f.research_degree yr, f.generation age, f.sex sx,
                      r.item_base it, r.headword pres, r.dialect_form base, r.grade g,
                      r.upt_dt upt
               FROM wb_weather_response r
               JOIN wb_weather_file f ON f.weather_file_id = r.weather_file_id
               WHERE """ + where, params
        ).fetchall()
        q = "SELECT COUNT(*) FROM wb_weather_file WHERE use_yn='Y'"
        if yr:
            q += ' AND research_degree=?'
            nfiles = con.execute(q, (yr,)).fetchone()[0]
        else:
            nfiles = con.execute(q).fetchone()[0]
    finally:
        con.close()
    recs = []
    for r in rows:
        pres = (r['pres'] or '').strip()
        base = (r['base'] or '').strip()
        g = grade_of(r['g'])
        recs.append({'rg': r['rg'], 'year': r['yr'], 'age': int(r['age']), 'sx': r['sx'],
                     'it': r['it'], 'pres': pres, 'base': base,
                     'form': base if base and base != '*' else pres,
                     'g': g,
                     # 관리자 편집이 '어느 행을 고칠지' 알 수 있게 행 id 를 함께 싣는다.
                     # 제보자 한 명·한 항목에 행이 여러 개라(표준어형 응답 + 지역어형 제시)
                     # 이게 없으면 어느 행을 고쳐야 하는지 알 수 없다.
                     'rid': r['rid'],
                     # 관리자가 고친 행. 목록의 '관리자가 고침' 검색이 이걸 센다.
                     'upt': r['upt'] if 'upt' in r.keys() else None})
    return recs, nfiles


def fill_db_qc(out, db_path=None, year=None):
    """QC 는 엑셀을 읽을 때만 채워지는 값이라, DB 원천일 때 다시 채운다.
       year 를 주면 그 연차 기준으로 센다. 화면의 연차 선택지도 함께 실어 준다."""
    import sqlite3
    path = db_path or WEATHER_DB
    yr = re.sub(r'\D', '', str(year or ''))[-2:]
    out['meta']['source'] = 'wb_weather_response (전용 테이블)'
    out['meta']['origin'] = 'db'
    out['meta']['year'] = yr or ''
    if not os.path.exists(path):
        return out
    con = sqlite3.connect(path)
    try:
        fw = "use_yn='Y'" + (' AND research_degree=?' if yr else '')
        fp = (yr,) if yr else ()
        rw = ("WHERE weather_file_id IN (SELECT weather_file_id FROM wb_weather_file WHERE %s)" % fw) if yr else ''
        qc = out['meta']['qc']
        qc['files'] = con.execute('SELECT COUNT(*) FROM wb_weather_file WHERE ' + fw, fp).fetchone()[0]
        qc['rowsTotal'] = con.execute(
            'SELECT COUNT(*) FROM wb_weather_response ' + rw, fp).fetchone()[0]
        qc['gradeFilled'] = con.execute(
            "SELECT COUNT(*) FROM wb_weather_response " + (rw + " AND " if rw else "WHERE ")
            + "grade_valid_yn='Y'", fp).fetchone()[0]
        qc['gradeBad'] = con.execute(
            "SELECT COUNT(*) FROM wb_weather_response " + (rw + " AND " if rw else "WHERE ")
            + "grade IS NOT NULL AND grade<>'*' AND grade_valid_yn='N'", fp).fetchone()[0]
        qc['layouts'] = con.execute(
            'SELECT COUNT(DISTINCT src_layout) FROM wb_weather_file WHERE ' + fw, fp).fetchone()[0]
        qc['calcDt'] = con.execute('SELECT MAX(reg_dt) FROM wb_weather_file').fetchone()[0]
        qc.pop('layoutOdd', None)
        # 화면 셀렉트를 하드코딩하지 않도록 실제 있는 연차를 알려 준다
        out['meta']['years'] = [
            {'degree': r[0], 'year': r[1], 'files': r[2]}
            for r in con.execute(
                """SELECT research_degree, MAX(research_year), COUNT(*)
                   FROM wb_weather_file WHERE use_yn='Y'
                   GROUP BY research_degree ORDER BY research_degree""")]
    finally:
        con.close()
    return out


def build_output(recs, nfiles):
    """레코드 목록 → awareness_by_region.json 과 같은 구조.

    원자료 엑셀(load_records)에서도, DB(wb_weather_response)에서도 같은 함수를 쓴다.
    구조와 판정이 두 곳으로 갈라지면 화면이 달라지므로 조립은 반드시 여기 한 곳에서만 한다."""
    print('제보자 파일 %d개 / 레코드 %d건' % (nfiles, len(recs)))

    # 9개 지역 전부에서 등급이 관측된 항목만 서비스 대상으로 삼는다 (=101개)
    seen = collections.defaultdict(set)
    for r in recs:
        if r['g']:
            seen[r['it']].add(r['rg'])
    core = sorted(it for it, s in seen.items() if len(s) == len(REGION_ORDER))
    print('전 지역 관측 항목: %d개' % len(core))

    headword = {}
    for it in core:
        c = collections.Counter(r['pres'] for r in recs if r['it'] == it and r['pres'])
        headword[it] = c.most_common(1)[0][0]

    by_item = collections.defaultdict(list)
    for r in recs:
        if r['it'] in set(core):
            by_item[r['it']].append(r)

    # 제보자 명부 — 한 세대·성별에 여러 명이 올 수 있으므로 (연차,세대,성별)로 식별한다.
    # 식별자는 원자료 파일명 어간과 같다: GG2420F = 경기·24년차·20대·여
    _seen = collections.defaultdict(set)
    for r in recs:
        _seen[r['rg']].add((r['year'], r['age'], r['sx']))
    roster = {rg: [{'id': '%s%s%02d%s' % (rg, y, a, sx), 'year': y, 'age': a, 'sex': sx}
                   for (y, a, sx) in sorted(v)]
              for rg, v in _seen.items()}

    items, tally = [], collections.Counter()
    def is_dialect(it, form):
        """그 어형을 지역어형으로 볼지 — 엑셀의 표제어와 다르면 지역어형이다.

        판정 근거를 원자료 안에서만 찾는다. 엑셀 밖의 판정 목록은 쓰지 않는다.
        """
        f = norm(form)
        return bool(f) and f not in head_forms(headword[it])

    for it in core:
        H = norm(headword[it])
        entry = {'code': it,
                 'word': WORD_ALIAS.get(headword[it], headword[it]),
                 'headword': headword[it],
                 'regions': {}}
        for rg in REGION_ORDER:
            rows = [r for r in by_item[it] if r['rg'] == rg]  # H 대신 is_dialect() 사용
            informants = sorted({(r['year'], r['age'], r['sx']) for r in rows})

            # 제보자별 지역어형 최선 등급(작을수록 살아 있음)
            best, bestform, bestrid, forms = {}, {}, {}, collections.Counter()
            for key in informants:
                mine = [(int(r['g']), r['form'], r.get('rid')) for r in rows
                        if r['g'] and is_dialect(it, r['form'])
                        and (r['year'], r['age'], r['sx']) == key]
                if mine:
                    lo = min(mine, key=lambda x: (x[0], x[1] or ''))
                    best[key] = lo[0]
                    bestform[key] = lo[1]
                    bestrid[key] = lo[2]
                    forms[lo[1]] += 1

            # 명부 전원을 한 줄씩 — 관리자 그리드가 '칸 하나 = 사람 한 명'으로
            # 편집할 수 있게, 집계 이전의 원자료 상태를 그대로 남긴다.
            #   d 지역어형 응답(등급 있음) · s 표준어형만 응답 · x 등급 미기입/무응답
            panel = []
            for inf in roster.get(rg, []):
                key = (inf['year'], inf['age'], inf['sex'])
                base = {'id': inf['id'], 'age': inf['age'], 'sex': inf['sex']}
                mine = [r for r in rows if (r['year'], r['age'], r['sx']) == key]
                if key in best:
                    base.update({'st': 'd', 'grade': best[key], 'form': bestform[key]})
                    if bestrid.get(key) is not None:
                        base['rid'] = bestrid[key]      # 편집 대상 행 (DB 원천일 때만)
                elif mine and not [r for r in mine if is_dialect(it, r['form'])]:
                    base['st'] = 's'
                else:
                    d = [r for r in mine if is_dialect(it, r['form'])]
                    base.update({'st': 'x', 'form': d[0]['form'] if d else None})
                panel.append(base)

            # 그 지역에서 실제로 적힌 표준어형·방언형(기저형). 등급 유무와 무관하게 원문 그대로.
            heads, bases = collections.Counter(), collections.Counter()
            for r in rows:
                pv = (r.get('pres') or '').strip()
                bv = (r.get('base') or '').strip()
                if pv and pv != '*':
                    heads[pv] += 1
                if bv and bv != '*':
                    bases[bv] += 1

            dial_rows = [r for r in rows if is_dialect(it, r['form'])]
            n = len(best)
            if n:
                score = region_score(best.values())
                state = weather_of(score)
                cell = {'state': state, 'n': n, 'score': score,
                        'dist': {str(k): sum(1 for v in best.values() if v == k) for k in (1, 2, 3, 4)},
                        'forms': [{'form': f, 'n': c} for f, c in forms.most_common()],
                        'cases': [{'id': '%s%s%02d%s' % (rg, y, a, s), 'age': a, 'sex': s,
                                   'grade': best[(y, a, s)], 'form': bestform[(y, a, s)]}
                                  for (y, a, s) in sorted(best)]}
                # 세대 칸, 그리고 세대×성별 칸.
                # 담당자가 '70대 남 / 70대 여' 를 나눠 보길 원해서 성별 축을 함께 낸다.
                # 판정 규칙은 둘이 같아야 한다 — 여기 한 곳에서만 만든다.
                def gen_cell(sel_rows, sel_best):
                    if sel_best:
                        # 지역 칸·전체·범례와 같은 식(등급 점수 평균)을 쓴다.
                        # 예전에는 세대 칸만 '최선 등급'(min)이라, 등급 1·4 두 명이면
                        # 맑음으로 칠해져 범례의 점수 구간과 어긋났다.
                        sc = region_score(sel_best.values())
                        return {'state': weather_of(sc), 'score': sc,
                                'n': len(sel_best),
                                'cases': [{'sex': sx, 'grade': sel_best[(y, a, sx)]}
                                          for (y, a, sx) in sorted(sel_best)]}
                    # 등급 있는 지역어형이 없다 → 조사는 됐는데 표준어형만이면 std, 아니면 관측 없음
                    return {'state': 'std' if sel_rows and not [r for r in sel_rows
                                                               if is_dialect(it, r['form'])] else 'w0',
                            'n': 0}

                cell['gens'] = {}
                cell['gensex'] = {}
                for g in (20, 50, 70):
                    cell['gens'][str(g)] = gen_cell(
                        [r for r in rows if r['age'] == g],
                        {k: v for k, v in best.items() if k[1] == g})
                    for sx in ('M', 'F'):
                        cell['gensex']['%d%s' % (g, sx)] = gen_cell(
                            [r for r in rows if r['age'] == g and r['sx'] == sx],
                            {k: v for k, v in best.items() if k[1] == g and k[2] == sx})
            elif rows and not dial_rows:
                # 조사됐고 응답이 전부 표준어형 → 표준어권(확정)
                state = 'std'
                cell = {'state': state, 'n': 0, 'score': None,
                        'respondents': len(informants),
                        'note': '조사된 %d명 모두 표준어형만 응답' % len(informants)}
            else:
                state = 'w0'
                cell = {'state': state, 'n': 0,
                        'note': ('지역어형은 나왔으나 사용도/인지도 미기입'
                                 if dial_rows else '해당 항목 응답 없음')}
            # 관리 목록이 '어디에 얼마나 쌓였나'를 보여주려면 원자료 행수가 필요하다.
            # 세는 곳이 갈라지면 화면끼리 숫자가 달라지므로 여기서만 센다.
            # 집계가 아니라 '그 지역에서 조사된 것' 을 원문대로 모은 목록.
            # forms 는 제보자별 최선 등급 어형만 담아 조사된 전부를 보여주지 못한다.
            # 지역어형이 안 나온 칸(std)에도 표준어형은 적혀 있으므로 모든 칸에 붙인다.
            cell['heads'] = [{'form': f, 'n': c} for f, c in heads.most_common()]
            cell['bases'] = [{'form': f, 'n': c} for f, c in bases.most_common()]
            cell['rows'] = len(rows)                                   # 그 지역의 응답 행
            cell['graded'] = sum(1 for r in rows if r['g'])            # 그중 등급이 적힌 행
            cell['people'] = len(informants)                           # 조사된 제보자
            cell['edited'] = sum(1 for r in rows if r.get('upt'))      # 관리자가 고친 행
            cell['panel'] = panel
            entry['regions'][rg] = cell
            tally[state] += 1
        items.append(entry)

    # ── 전 항목 통합 기상도 ──
    # 규칙은 항목별 칸과 똑같다: '그 지역 제보자가 지역어형에 준 등급의 점수 평균'.
    # 항목 하나가 아니라 101개 전부의 등급을 모아 같은 식에 넣을 뿐이라 새 해석이 아니다.
    # 제보자마다 항목 수가 다르므로 (제보자, 항목)마다 최선 등급 하나씩을 모은다.
    overall = {'code': '__ALL__', 'word': '전체', 'headword': '전체', 'regions': {}}
    for rg in REGION_ORDER:
        grades, per_key = [], collections.defaultdict(list)
        for it in core:
            for r in by_item[it]:
                if r['rg'] != rg or not r['g'] or not is_dialect(it, r['form']):
                    continue
                per_key[(it, r['year'], r['age'], r['sx'])].append(int(r['g']))
        for (it, y, a, sx), gs in per_key.items():
            grades.append(min(gs))                       # 제보자 한 명의 그 항목 최선 등급
        people = sorted({(k[1], k[2], k[3]) for k in per_key})
        score = region_score(grades) if grades else None
        cell = {'state': weather_of(score), 'n': len(people), 'score': score,
                'answers': len(grades),                  # 평균에 들어간 (제보자,항목) 수
                'items': len({k[0] for k in per_key}),   # 지역어형이 나온 항목 수
                'dist': {str(k): sum(1 for g in grades if g == k) for k in (1, 2, 3, 4)}}
        for axis, keys in (('gens', [(g, None) for g in (20, 50, 70)]),
                           ('gensex', [(g, sx) for g in (20, 50, 70) for sx in ('M', 'F')])):
            cell[axis] = {}
            for g, sx in keys:
                sel = [min(v) for k, v in per_key.items()
                       if k[2] == g and (sx is None or k[3] == sx)]
                key = str(g) if sx is None else '%d%s' % (g, sx)
                sc = region_score(sel) if sel else None
                # cases 는 싣지 않는다 — 통합은 (제보자,항목) 조합이 수백 개라
                # 사례 나열이 뜻을 갖지 않고, 화면이 그 길이를 '명 수'로 잘못 읽는다.
                cell[axis][key] = {
                    'state': weather_of(sc), 'score': sc,
                    'n': len({(k[1], k[3]) for k in per_key
                              if k[2] == g and (sx is None or k[3] == sx)}),   # 제보자 수
                    'answers': len(sel),                                       # 평균에 들어간 응답 수
                    'used': sum(1 for x in sel if x == 1)}                     # 그중 '지금도 씀'
        overall['regions'][rg] = cell

    # 지역별로 실제 조사된 제보자 구성 — '조사 안 함'과 '지역어형 안 나옴'을 구분하기 위해

    out = {
        'meta': {
            'title': '지역별 인지도(기상도) — 어휘 조사 실측',
            'source': 'data/dialect_gisangdo/*.xlsx (2024년 차 어휘 조사 원자료)',
            'informants': nfiles,
            'items': len(items),
            'scale': '① 사용 100 · ② 이해 75 · ③ 인지 50 · ④ 무지 25 (어형 단위 부여)',
            'metric': '지역 점수 = 그 지역 제보자가 지역어형에 준 등급의 점수 평균 (0~100)',
            'thresholds': ('w1 ≥87.5 · w2 ≥62.5 · w3 ≥37.5 · w4 <37.5 '
                           '— 경계는 등급 사이의 중간값이며 별도로 정한 임계값이 아니다'),
            'qc': dict(QC, coreItems=len(core),
                       cells=len(REGION_ORDER) * len(core),
                       states=dict(tally)),
            'caveat': ('셀당 제보자 1~6명. 제보자가 1~2명인 칸이 전체의 약 1/4이므로 점수를 '
                       '단독으로 읽지 말고 n을 함께 볼 것. 지역×세대 셀은 1~2명이라 사례(cases)로만 제공. '
                       '등급 기입률이 조사자별 63~100%로 달라 지역 간 절대 비교는 피할 것. '
                       '서울은 조사 지역에 포함되지 않음.'),
        },
        'regionOrder': REGION_ORDER,
        'regionNames': REGION_NAMES,
        'regionSites': {'GG': '부천', 'GW': '홍천', 'CB': '진천', 'CN': '천안', 'GB': '구미',
                        'GN': '김해', 'JB': '정읍', 'JN': '광양', 'JJ': '서귀포 대정읍'},
        'regionRoster': roster,
        'states': {
            'w1':  {'label': '맑음',      'icon': 'ti-sun',       'emoji': '☀️', 'desc': '지역어를 지금도 씀'},
            'w2':  {'label': '구름조금',  'icon': 'ti-cloud',     'emoji': '🌤️', 'desc': '써 봤으나 지금은 안 씀'},
            'w3':  {'label': '흐리고 비', 'icon': 'ti-cloud-rain','emoji': '🌧️', 'desc': '들어는 봄'},
            'w4':  {'label': '천둥번개',  'icon': 'ti-cloud-bolt','emoji': '⛈️', 'desc': '처음 들음'},
            'std': {'label': '표준어권',  'icon': 'ti-circle-check', 'emoji': '🔵',
                    'desc': '조사됐으나 지역어형이 나오지 않음 · 표준어로 통일'},
            'w0':  {'label': '관측 없음', 'icon': 'ti-cloud-off', 'emoji': '🌫️', 'desc': '자료 없음'},
        },
        'items': items,
        # 단어를 고르지 않고 전 항목을 한 장으로 볼 때 쓴다 (items 와 같은 모양)
        'overall': overall,
    }
    return out


def main(argv=None):
    import argparse
    ap = argparse.ArgumentParser(description='지역별 인지도(기상도) ETL')
    ap.add_argument('--from-db', action='store_true',
                    help='엑셀 대신 전용 테이블(wb_weather_response)에서 읽는다')
    ap.add_argument('--db', default=None, help='기상도 DB 경로 (기본 data/gisangdo.db)')
    ap.add_argument('--year', default=None, help="조사 연차만 추림 (예: 24 또는 2024)")
    ap.add_argument('--out', default=OUT, help='산출 JSON 경로')
    a = ap.parse_args(argv)

    if a.from_db:
        recs, nfiles = load_records_from_db(a.db, a.year)
        print('DB 원천%s — 제보자 %d명 / 레코드 %d건'
              % ((' (연차 %s)' % a.year) if a.year else '', nfiles, len(recs)))
        out = build_output(recs, nfiles)
        fill_db_qc(out, a.db, a.year)
    else:
        recs, nfiles = load_records()
        out = build_output(recs, nfiles)

    with io.open(a.out, 'w', encoding='utf-8') as fp:
        json.dump(out, fp, ensure_ascii=False, indent=1)
    OUT_USED = a.out

    tally = out['meta']['qc']['states']
    print('상태 분포 (항목×지역 %d셀):' % sum(tally.values()))
    for k in ['w1', 'w2', 'w3', 'w4', 'std', 'w0']:
        print('  %-4s %4d' % (k, tally.get(k, 0)))
    print('저장:', OUT_USED)


if __name__ == '__main__':
    main()
