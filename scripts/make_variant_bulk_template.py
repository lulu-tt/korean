#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
지역별 이형태(음운) 일괄등록 서식 — 도별 9파일 + zip 생성

모양을 국립국어원 통합자료와 **똑같이** 맞춘다.
  · 파일 1개 = 도 1개, 시트는 「음운」 한 장.
  · 1행 제목 / 2행 헤더 / 3행부터 데이터.
  · 도는 파일명에서 판별하므로 파일명에 도 이름이 들어가야 한다.
  · 2행 C열부터의 조사지점은 site_map.json 의 실제 raw_header 로 미리 채운다.
    헤더 문자열이 곧 지점의 열쇠라 손으로 적으면 오타가 곧 새 지점이 된다.
  · 경기 파일에만 작성 예시 9행을 넣는다(실제 경기 원문 발췌라 값이 정합).

사용
  python3 scripts/make_variant_bulk_template.py
"""
from __future__ import annotations

import argparse
import json
import sys
import tempfile
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
_OUT_DIR = ROOT / "neibis-cms" / "mariadb" / "neibis" / "survey" / "data"
ZIP_OUT = _OUT_DIR / "지역별이형태_일괄등록_양식.zip"
SITE_MAP = ROOT / "data" / "processed" / "site_map.json"
SRC_DIR = ROOT / "data" / "dialect_phonology_compare"

PROV_ORDER = ["GG", "GW", "CB", "CN", "JB", "JN", "GB", "GN", "JJ"]
PROV_NAME = {
    "GG": "경기", "GW": "강원", "CB": "충북", "CN": "충남", "JB": "전북",
    "JN": "전남", "GB": "경북", "GN": "경남", "JJ": "제주",
}

# ── 서식 스타일 ────────────────────────────────────────────────────
C_HEAD = "FF1F3864"      # 헤더 배경(감청)
C_HEAD_REQ = "FF7B2D26"  # 필수열 헤더 배경(적벽돌)
C_SAMPLE = "FFFFF7E6"    # 예시행 배경(연노랑)
C_NOTE = "FF64748B"      # 설명 텍스트

F_TITLE = Font(name="맑은 고딕", size=14, bold=True, color="FF1F3864")
F_H2 = Font(name="맑은 고딕", size=11, bold=True, color="FF1F3864")
F_HEAD = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFFFF")
F_BODY = Font(name="맑은 고딕", size=10)
F_NOTE = Font(name="맑은 고딕", size=9, color=C_NOTE)
F_SAMPLE = Font(name="맑은 고딕", size=10, color="FF7C4A03", italic=True)

THIN = Side(style="thin", color="FFBFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def put(ws, row, col, value, font=F_BODY, align=WRAP, fill=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.alignment = align
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if border:
        c.border = BOX
    return c


def write_header(ws, row, cols):
    """cols = [(제목, 폭, 필수여부)] → 헤더행 작성 + 열 폭 지정."""
    for i, (title, width, req) in enumerate(cols, start=1):
        put(
            ws, row, i,
            title + ("*" if req else ""),
            font=F_HEAD, align=CENTER,
            fill=C_HEAD_REQ if req else C_HEAD,
        )
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 30


def write_sample(ws, start_row, rows):
    for r, vals in enumerate(rows, start=start_row):
        for c, v in enumerate(vals, start=1):
            put(ws, r, c, v, font=F_SAMPLE, fill=C_SAMPLE)
    return start_row + len(rows)


# ── 열 정의 ───────────────────────────────────────────────────────
# 서식에는 더 이상 쓰이지 않는다. scripts/export_variant_bulk_data.py 가
# 내보내기 시트를 그릴 때 가져다 쓰므로 남겨 둔다.
ITEM_COLS = [
    ("항목코드", 16, True),
    ("표준어", 26, True),
    ("구분", 10, True),
    ("상위항목코드", 16, False),
    ("대역", 8, False),
    ("정렬순서", 10, False),
    ("사용여부", 10, False),
    ("비고", 40, False),
]

SITE_COLS = [
    ("지점열 헤더(원문)", 26, True),
    ("도코드", 10, True),
    ("도명", 10, False),
    ("지점명", 14, True),
    ("조사연도", 10, False),
    ("지점ID(수정 시)", 30, False),
    ("사용여부", 10, False),
    ("비고", 34, False),
]

LONG_COLS = [
    ("항목코드", 16, True),
    ("지점열 헤더(원문)", 26, True),
    ("도코드", 10, False),
    ("지점명", 14, False),
    ("조사연도", 10, False),
    ("응답 원문", 40, True),
    ("결측", 8, False),
    ("비고", 30, False),
]


# ── 시트: 음운 (유일한 데이터 시트) ────────────────────────────────
# 국립국어원 통합자료와 동일 레이아웃: 1행 제목 / 2행 헤더 / 3행부터 데이터.
FIXED_COLS = ["항목번호", "표준어"]

# 경기 통합자료 실제 원문 발췌 (data/dialect_phonology_compare/…경기…xlsx).
# 열 순서는 site_map 의 경기 지점 순서와 같다.
GG_SAMPLE = [
    ["31001", "테(輪)", "테", "", "", "", "", ""],
    ["31001-0-1", "-이/가", "테가 이:쁘다", "태가", "테가", "테가", "테가", "테가"],
    ["31001-0-2", "-보다", "테보덤", "태보다", "", "", "테를", ""],
    ["31002", "태(胎)", "태", "", "", "", "", ""],
    ["31002-0-1", "-이/가", "태가", "태가", "태가", "태가", "태가", "태가"],
    ["31002-0-2", "-보다", "*", "태보다", "", "", "태를", ""],
    ["32001", "막-(防)[ㄱ]", "막는다", "", "", "", "", "망는다"],
    ["32001-0-1", "-지", "막찌", "막찌", "막찌 마", "막찌", "망는다 | 흐르지", "막찌"],
    ["32001-0-2", "-고", "막꼬", "마꼬 이따 | 막꼬", "막꾸", "막꾸", "막꾸", "마꾸"],
]


def _source_header_order() -> dict:
    """
    도별 지점 열 순서를 원본 엑셀에서 복원한다. site_map.json 은 가나다순이라
    그대로 쓰면 열 순서가 원본과 달라진다(예시 데이터가 엉뚱한 열에 붙는다).
    """
    seq = {}
    if not SRC_DIR.is_dir():
        return seq
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import etl_phonology as E  # noqa: PLC0415
    for f in sorted(SRC_DIR.iterdir()):
        if not f.is_file() or f.suffix.lower() not in (".xls", ".xlsx") or f.name.startswith("~$"):
            continue
        pc, _ = E.detect_province(f.name)
        if not pc:
            continue
        r = E.parse_phonology_sheet(f)
        if r.get("ok"):
            seq[pc] = list(r["site_headers"])
    return seq


def load_site_headers() -> dict:
    """
    도코드 → 조사지점 raw_header 목록. 손으로 적으면 오타 하나가 새 지점이 되므로
    적재된 실제 값(site_map.json)을 그대로 쓰되, 순서는 원본 엑셀에 맞춘다.
    """
    if not SITE_MAP.is_file():
        raise SystemExit("site_map.json 이 없습니다: %s" % SITE_MAP)
    rows = json.loads(SITE_MAP.read_text(encoding="utf-8"))
    by_prov = {}
    for r in rows:
        by_prov.setdefault(r["province_code"], []).append(r)

    seq = _source_header_order()
    out = {}
    for pc, lst in by_prov.items():
        idx = {h: i for i, h in enumerate(seq.get(pc) or [])}
        # 원본을 못 읽으면 (조사연도, 헤더) 순으로 대체
        lst = sorted(lst, key=lambda r: (idx.get(r["raw_header"], 999),
                                         r.get("survey_year") or 9999,
                                         r["raw_header"]))
        out[pc] = [r["raw_header"] for r in lst]

    missing = [c for c in PROV_ORDER if not out.get(c)]
    if missing:
        raise SystemExit("지점이 없는 도: %s" % ", ".join(missing))
    return out


def build_province(prov_code: str, headers: list, out: Path, sample=None) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    # 시트명은 국립국어원 통합자료와 똑같이 「음운」. 업로더가 이 이름을 찾는다.
    ws = wb.create_sheet("음운")
    cols = FIXED_COLS + list(headers)

    put(ws, 1, 1,
        "%s 지역별 이형태 — 행 = 항목, 열 = 조사지점 (3행부터 입력)" % PROV_NAME[prov_code],
        font=F_H2, border=False)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))

    for i, title in enumerate(cols, start=1):
        put(ws, 2, i, title, font=F_HEAD, align=CENTER,
            fill=C_HEAD_REQ if i <= 2 else C_HEAD)
        ws.column_dimensions[get_column_letter(i)].width = 16 if i <= 2 else 22
    ws.column_dimensions["B"].width = 20
    ws.row_dimensions[2].height = 30

    if sample:
        write_sample(ws, 3, sample)
    ws.freeze_panes = "C3"
    # 데이터 영역 아래에는 아무것도 쓰지 않는다 — 파서가 설명문을 데이터 행으로 읽는다.
    wb.properties.title = "%s 지역별 이형태 일괄등록 서식" % PROV_NAME[prov_code]
    wb.properties.creator = "NEIBIS"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def build_zip(zip_out: Path = ZIP_OUT) -> Path:
    """도별 9파일을 만들어 zip 하나로 묶는다. 낱개 xlsx 는 남기지 않는다."""
    headers = load_site_headers()
    zip_out.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory() as tmp:
        made = []
        for code in PROV_ORDER:
            # 파일명에 도 이름이 있어야 업로더가 도를 판별한다(detect_province).
            name = "지역별이형태_%s.xlsx" % PROV_NAME[code]
            made.append(build_province(
                code, headers[code], Path(tmp) / name,
                sample=GG_SAMPLE if code == "GG" else None))
        with zipfile.ZipFile(zip_out, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in made:
                zf.write(f, arcname=f.name)
    return zip_out


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=str(ZIP_OUT), help="zip 경로")
    args = ap.parse_args()
    print("생성:", build_zip(Path(args.out)))
