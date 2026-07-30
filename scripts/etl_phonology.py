#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
음운(phonology) 시트 ETL — 실엑셀 원문만 적재

원칙
  · 셀 값은 엑셀에 있는 문자열만 저장 (추정·시연 데이터 생성 금지)
  · 허용 가공: trim, 코드 float 정규화(20101.0→20101), HTML 엔티티 복원
  · 결측: 빈칸 / '*' → is_missing=1, raw_text는 원문 유지('', '*')
  · 지점: 엑셀 헤더 문자열을 raw_header로 보존, site_map으로만 정규화

사용
  python3 scripts/etl_phonology.py
  python3 scripts/etl_phonology.py --data-dir ./data --out-dir ./data/processed
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import re
import sqlite3
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    import xlrd
except ImportError:
    xlrd = None

# ── 상수 ──────────────────────────────────────────────────────────
HEADER_KEYS = {"code", "코드", "항목번호"}
SHEET_NAME = "음운"
DOMAIN = "phonology"

PROVINCE_FROM_FILENAME = [
    ("경기", "GG", "경기"),
    ("강원", "GW", "강원"),
    ("충북", "CB", "충북"),
    ("충남", "CN", "충남"),
    ("전북", "JB", "전북"),
    ("전남", "JN", "전남"),
    ("경북", "GB", "경북"),
    ("경남", "GN", "경남"),
    ("제주", "JJ", "제주"),
]


def detect_province(filename: str) -> tuple[str | None, str | None]:
    # macOS 파일명이 NFD일 수 있어 NFC로 통일 후 매칭
    name_n = unicodedata.normalize("NFC", filename)
    for key, code, name in PROVINCE_FROM_FILENAME:
        if key in name_n:
            return code, name
    return None, None


def norm_code(val) -> str | None:
    """항목번호 정규화. 엑셀 float(20101.0) → '20101'. 내용 창작 없음."""
    if val is None:
        return None
    if isinstance(val, float):
        if val != val:  # NaN
            return None
        if val == int(val):
            return str(int(val))
        s = str(val).strip()
    else:
        s = str(val).strip()
    if not s:
        return None
    # 20101.0 형태
    m = re.fullmatch(r"(\d+)\.0+", s)
    if m:
        return m.group(1)
    return s


def clean_cell(val) -> str:
    """표시/저장용 최소 정리. 의미 변경 없이 공백·엔티티만."""
    if val is None:
        return ""
    if isinstance(val, float):
        if val != val:
            return ""
        if val == int(val):
            s = str(int(val))
        else:
            s = str(val)
    else:
        s = str(val)
    s = html.unescape(s)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.strip()
    return s


def is_missing_text(raw: str) -> bool:
    return raw == "" or raw == "*"


def is_header_cell(val) -> bool:
    if val is None:
        return False
    return str(val).strip().lower() in HEADER_KEYS


def site_id_for(province_code: str | None, raw_header: str) -> str:
    """안정적 site_id: 도코드 + 헤더 해시 (헤더 원문 변경 시에만 바뀜)."""
    base = f"{province_code or 'XX'}|{raw_header}"
    h = hashlib.sha1(base.encode("utf-8")).hexdigest()[:10]
    # 사람이 읽기 쉬운 prefix
    slug = re.sub(r"[^\w가-힣]+", "_", raw_header)[:24].strip("_")
    pc = province_code or "XX"
    return f"{pc}_{slug}_{h}" if slug else f"{pc}_{h}"


def parse_site_header(raw_header: str) -> dict:
    """
    헤더 문자열에서 지명·연도 추정 (실패 시 null).
    추정치이며 raw_header가 정본. 서비스는 raw 우선.
    """
    place = None
    year = None
    # 연도
    ym = re.search(r"(19|20)\d{2}", raw_header)
    if ym:
        year = int(ym.group(0))
    # SJB_IS(임실) / SCB_CJ 충북 제천(2005)
    m = re.search(r"\(([^)]+)\)\s*$", raw_header)
    if m and not re.fullmatch(r"(19|20)\d{2}", m.group(1).strip()):
        inner = m.group(1).strip()
        # 임실, 완주 등
        if not re.search(r"\d{4}", inner):
            place = re.sub(r"^(강원|충북|충남|전북|전남|경기|경북|경남|제주)\s*", "", inner).strip()
    # 용인(2004), 곡성(2005), 고성(2004)
    m2 = re.match(r"^([가-힣]+)\s*\(\s*((?:19|20)\d{2})\s*\)\s*$", raw_header)
    if m2:
        place = m2.group(1)
        year = int(m2.group(2))
    # 강원 삼척, 충북 제천(2005)
    m3 = re.match(
        r"^(?:강원|충북|충남|전북|전남|경기|경북|경남|제주)\s+([가-힣]+)(?:\s*\(\s*((?:19|20)\d{2})\s*\))?$",
        raw_header,
    )
    if m3:
        place = m3.group(1)
        if m3.group(2):
            year = int(m3.group(2))
    # 코드접두 제거: SJB_IS(임실)
    m4 = re.search(r"[A-Z]{2,3}_[A-Z]{1,3}\s*\(\s*([가-힣]+)", raw_header)
    if m4:
        place = m4.group(1)
    # 의성, 청도, 가시 등 단독 지명
    if place is None and re.fullmatch(r"[가-힣]{2,6}", raw_header):
        place = raw_header
    # SCB_CJ 충북 제천(2005)
    m5 = re.search(r"([가-힣]{2,6})\s*\(\s*((?:19|20)\d{2})\s*\)", raw_header)
    if m5 and place is None:
        place = m5.group(1)
        year = int(m5.group(2))
    return {"place_name": place, "survey_year": year}


def parent_code_of(item_code: str) -> str | None:
    if "-" in item_code:
        return item_code.split("-", 1)[0]
    return None


def is_parent_code(item_code: str) -> bool:
    return "-" not in item_code and bool(re.fullmatch(r"\d+", item_code))


# ── 엑셀 읽기 ─────────────────────────────────────────────────────
def iter_sheet_rows_xlsx(path: Path, sheet_name: str):
    if openpyxl is None:
        raise RuntimeError("openpyxl 필요: pip install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return None
    ws = wb[sheet_name]
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        yield i, list(row) if row else []
    wb.close()


def iter_sheet_rows_xls(path: Path, sheet_name: str):
    if xlrd is None:
        raise RuntimeError("xlrd 필요: pip install xlrd")
    wb = xlrd.open_workbook(str(path))
    if sheet_name not in wb.sheet_names():
        return None
    ws = wb.sheet_by_name(sheet_name)
    for i in range(ws.nrows):
        yield i + 1, [ws.cell_value(i, j) for j in range(ws.ncols)]


def iter_sheet_rows(path: Path, sheet_name: str = SHEET_NAME):
    suf = path.suffix.lower()
    if suf == ".xlsx":
        gen = iter_sheet_rows_xlsx(path, sheet_name)
    elif suf == ".xls":
        gen = iter_sheet_rows_xls(path, sheet_name)
    else:
        return None
    if gen is None:
        return None
    return gen


def parse_phonology_sheet(path: Path) -> dict:
    """
    한 파일의 음운 시트 → 구조화 결과.
    returns dict with keys: ok, error, header_row, site_headers, rows, stats
    """
    prov_code, prov_name = detect_province(path.name)
    gen = iter_sheet_rows(path, SHEET_NAME)
    if gen is None:
        return {
            "ok": False,
            "error": f"시트 '{SHEET_NAME}' 없음 또는 형식 미지원",
            "source_file": path.name,
            "province_code": prov_code,
            "province_name": prov_name,
        }

    header_row_idx = None
    site_headers: list[str] = []
    col_code = 0
    col_std = 1
    site_cols: list[tuple[int, str]] = []  # (col_index, raw_header)
    items_order: list[str] = []
    items: dict[str, str] = {}  # code -> standard_form (마지막 non-empty 우선)
    responses: list[dict] = []
    empty_code_rows = 0
    data_rows = 0

    buffered = list(gen)
    for row_idx, cells in buffered:
        if not cells:
            continue
        # 헤더 탐지
        if header_row_idx is None and cells and is_header_cell(cells[0]):
            header_row_idx = row_idx
            # 지점 열: 2번째 이후 non-empty
            for ci, val in enumerate(cells):
                if ci < 2:
                    continue
                h = clean_cell(val)
                if h:
                    site_cols.append((ci, h))
                    site_headers.append(h)
            continue

        if header_row_idx is None:
            continue

        code = norm_code(cells[0] if len(cells) > 0 else None)
        if not code:
            # 데이터 영역에서 코드 없는 행 스킵
            if any(clean_cell(c) for c in cells[1:]):
                empty_code_rows += 1
            continue

        std = clean_cell(cells[1] if len(cells) > 1 else "")
        data_rows += 1
        if code not in items:
            items_order.append(code)
        # 표준어: 비어 있지 않으면 갱신 (부모 행에 표제어 있는 경우)
        if std or code not in items:
            items[code] = std if std else items.get(code, "")

        for ci, raw_header in site_cols:
            raw = clean_cell(cells[ci] if ci < len(cells) else "")
            responses.append(
                {
                    "item_code": code,
                    "raw_header": raw_header,
                    "raw_text": raw,
                    "is_missing": 1 if is_missing_text(raw) else 0,
                    "source_row": row_idx,
                }
            )

    if header_row_idx is None:
        return {
            "ok": False,
            "error": "헤더 행(Code/항목번호)을 찾지 못함",
            "source_file": path.name,
            "province_code": prov_code,
            "province_name": prov_name,
        }

    return {
        "ok": True,
        "error": None,
        "source_file": path.name,
        "source_path": str(path),
        "province_code": prov_code,
        "province_name": prov_name,
        "header_row": header_row_idx,
        "site_headers": site_headers,
        "items_order": items_order,
        "items": items,
        "responses": responses,
        "stats": {
            "data_rows": data_rows,
            "n_sites": len(site_cols),
            "n_responses": len(responses),
            "n_items": len(items),
            "empty_code_rows": empty_code_rows,
            "n_missing": sum(1 for r in responses if r["is_missing"]),
            "n_filled": sum(1 for r in responses if not r["is_missing"]),
        },
    }


# ── DB ────────────────────────────────────────────────────────────
SCHEMA = """
PRAGMA journal_mode=WAL;
PRAGMA foreign_keys=ON;

CREATE TABLE IF NOT EXISTS meta (
  key TEXT PRIMARY KEY,
  value TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS import_batch (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  source_file TEXT NOT NULL,
  province_code TEXT,
  province_name TEXT,
  sheet_name TEXT NOT NULL,
  header_row INTEGER,
  n_sites INTEGER,
  n_items INTEGER,
  n_responses INTEGER,
  n_filled INTEGER,
  n_missing INTEGER,
  imported_at TEXT NOT NULL,
  status TEXT NOT NULL,
  error TEXT
);

CREATE TABLE IF NOT EXISTS survey_site (
  site_id TEXT PRIMARY KEY,
  province_code TEXT,
  province_name TEXT,
  raw_header TEXT NOT NULL,
  place_name TEXT,
  survey_year INTEGER,
  source_file TEXT,
  UNIQUE(province_code, raw_header, source_file)
);

CREATE TABLE IF NOT EXISTS item (
  item_code TEXT PRIMARY KEY,
  domain TEXT NOT NULL,
  standard_form TEXT,
  parent_code TEXT,
  is_parent INTEGER NOT NULL DEFAULT 0
);

CREATE TABLE IF NOT EXISTS response (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  item_code TEXT NOT NULL,
  site_id TEXT NOT NULL,
  raw_text TEXT NOT NULL,
  is_missing INTEGER NOT NULL DEFAULT 0,
  source_file TEXT NOT NULL,
  source_row INTEGER,
  import_batch_id INTEGER,
  FOREIGN KEY (item_code) REFERENCES item(item_code),
  FOREIGN KEY (site_id) REFERENCES survey_site(site_id),
  FOREIGN KEY (import_batch_id) REFERENCES import_batch(id),
  UNIQUE(item_code, site_id, source_file)
);

CREATE INDEX IF NOT EXISTS idx_response_item ON response(item_code);
CREATE INDEX IF NOT EXISTS idx_response_site ON response(site_id);
CREATE INDEX IF NOT EXISTS idx_item_parent ON item(parent_code);
CREATE INDEX IF NOT EXISTS idx_site_prov ON survey_site(province_code);
"""


def init_db(db_path: Path) -> sqlite3.Connection:
    if db_path.exists():
        db_path.unlink()
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    conn.executescript(SCHEMA)
    return conn


def load_into_db(conn: sqlite3.Connection, parsed_list: list[dict]) -> dict:
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    all_items: dict[str, str] = {}
    site_map_rows: list[dict] = []
    totals = Counter()

    for parsed in parsed_list:
        if not parsed.get("ok"):
            conn.execute(
                """INSERT INTO import_batch
                   (source_file, province_code, province_name, sheet_name, imported_at, status, error,
                    n_sites, n_items, n_responses, n_filled, n_missing)
                   VALUES (?,?,?,?,?,?,?,0,0,0,0,0)""",
                (
                    parsed["source_file"],
                    parsed.get("province_code"),
                    parsed.get("province_name"),
                    SHEET_NAME,
                    now,
                    "error",
                    parsed.get("error"),
                ),
            )
            totals["files_error"] += 1
            continue

        st = parsed["stats"]
        cur = conn.execute(
            """INSERT INTO import_batch
               (source_file, province_code, province_name, sheet_name, header_row,
                n_sites, n_items, n_responses, n_filled, n_missing, imported_at, status, error)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,NULL)""",
            (
                parsed["source_file"],
                parsed["province_code"],
                parsed["province_name"],
                SHEET_NAME,
                parsed["header_row"],
                st["n_sites"],
                st["n_items"],
                st["n_responses"],
                st["n_filled"],
                st["n_missing"],
                now,
                "ok",
            ),
        )
        batch_id = cur.lastrowid
        totals["files_ok"] += 1

        # sites
        header_to_sid: dict[str, str] = {}
        for h in parsed["site_headers"]:
            sid = site_id_for(parsed["province_code"], h)
            # 동일 헤더가 파일마다 있으면 source_file 포함해 구분
            sid = site_id_for(parsed["province_code"], f"{h}|{parsed['source_file']}")
            header_to_sid[h] = sid
            parsed_h = parse_site_header(h)
            conn.execute(
                """INSERT INTO survey_site
                   (site_id, province_code, province_name, raw_header, place_name, survey_year, source_file)
                   VALUES (?,?,?,?,?,?,?)
                   ON CONFLICT(site_id) DO UPDATE SET
                     province_code=excluded.province_code,
                     province_name=excluded.province_name,
                     raw_header=excluded.raw_header,
                     place_name=excluded.place_name,
                     survey_year=excluded.survey_year,
                     source_file=excluded.source_file""",
                (
                    sid,
                    parsed["province_code"],
                    parsed["province_name"],
                    h,
                    parsed_h["place_name"],
                    parsed_h["survey_year"],
                    parsed["source_file"],
                ),
            )
            site_map_rows.append(
                {
                    "site_id": sid,
                    "province_code": parsed["province_code"],
                    "province_name": parsed["province_name"],
                    "raw_header": h,
                    "place_name": parsed_h["place_name"],
                    "survey_year": parsed_h["survey_year"],
                    "source_file": parsed["source_file"],
                }
            )

        for code, std in parsed["items"].items():
            if code not in all_items or (std and not all_items.get(code)):
                all_items[code] = std or all_items.get(code, "")
            # FK: response 이전에 item 존재 필요
            parent = parent_code_of(code)
            prev = conn.execute(
                "SELECT standard_form FROM item WHERE item_code=?", (code,)
            ).fetchone()
            form = std or (prev["standard_form"] if prev else "") or all_items.get(code, "")
            conn.execute(
                """INSERT INTO item (item_code, domain, standard_form, parent_code, is_parent)
                   VALUES (?,?,?,?,?)
                   ON CONFLICT(item_code) DO UPDATE SET
                     standard_form=CASE
                       WHEN excluded.standard_form != '' THEN excluded.standard_form
                       ELSE item.standard_form END,
                     parent_code=excluded.parent_code,
                     is_parent=excluded.is_parent""",
                (code, DOMAIN, form, parent, 1 if is_parent_code(code) else 0),
            )

        # responses
        for r in parsed["responses"]:
            sid = header_to_sid[r["raw_header"]]
            # 안전: 응답 코드가 items에 누락된 경우 최소 item 생성
            if not conn.execute(
                "SELECT 1 FROM item WHERE item_code=?", (r["item_code"],)
            ).fetchone():
                pc = parent_code_of(r["item_code"])
                conn.execute(
                    """INSERT INTO item (item_code, domain, standard_form, parent_code, is_parent)
                       VALUES (?,?,?,?,?)
                       ON CONFLICT(item_code) DO NOTHING""",
                    (r["item_code"], DOMAIN, "", pc, 1 if is_parent_code(r["item_code"]) else 0),
                )
            conn.execute(
                """INSERT INTO response
                   (item_code, site_id, raw_text, is_missing, source_file, source_row, import_batch_id)
                   VALUES (?,?,?,?,?,?,?)
                   ON CONFLICT(item_code, site_id, source_file) DO UPDATE SET
                     raw_text=excluded.raw_text,
                     is_missing=excluded.is_missing,
                     source_row=excluded.source_row,
                     import_batch_id=excluded.import_batch_id""",
                (
                    r["item_code"],
                    sid,
                    r["raw_text"],
                    r["is_missing"],
                    parsed["source_file"],
                    r["source_row"],
                    batch_id,
                ),
            )
            totals["responses"] += 1
            if r["is_missing"]:
                totals["missing"] += 1
            else:
                totals["filled"] += 1

    totals["items"] = conn.execute("SELECT COUNT(*) FROM item").fetchone()[0]

    conn.execute(
        "INSERT OR REPLACE INTO meta(key,value) VALUES (?,?)",
        ("domain", DOMAIN),
    )
    conn.execute(
        "INSERT OR REPLACE INTO meta(key,value) VALUES (?,?)",
        ("imported_at", now),
    )
    conn.execute(
        "INSERT OR REPLACE INTO meta(key,value) VALUES (?,?)",
        ("policy", "raw_excel_only_no_synthetic"),
    )
    conn.commit()
    return {"totals": dict(totals), "site_map_rows": site_map_rows, "imported_at": now, "all_items": all_items}


# ── export ────────────────────────────────────────────────────────
def export_json(conn: sqlite3.Connection, out_dir: Path, site_map_rows: list, imported_at: str):
    out_dir.mkdir(parents=True, exist_ok=True)

    sites = [dict(r) for r in conn.execute("SELECT * FROM survey_site ORDER BY province_code, raw_header")]
    items = [dict(r) for r in conn.execute("SELECT * FROM item ORDER BY item_code")]
    batches = [dict(r) for r in conn.execute("SELECT * FROM import_batch ORDER BY id")]

    # 트리: parent -> children
    parents = [it for it in items if it["is_parent"]]
    children_by_parent: dict[str, list] = defaultdict(list)
    for it in items:
        if it["parent_code"]:
            children_by_parent[it["parent_code"]].append(
                {"item_code": it["item_code"], "standard_form": it["standard_form"]}
            )

    meta = {
        "domain": DOMAIN,
        "policy": "raw_excel_only_no_synthetic",
        "imported_at": imported_at,
        "description": "국립국어원 지역어 조사 통합자료 음운 시트. 셀 값은 엑셀 원문만 포함.",
        "counts": {
            "sites": len(sites),
            "items": len(items),
            "parent_items": len(parents),
            "responses": conn.execute("SELECT COUNT(*) FROM response").fetchone()[0],
            "responses_filled": conn.execute("SELECT COUNT(*) FROM response WHERE is_missing=0").fetchone()[0],
            "responses_missing": conn.execute("SELECT COUNT(*) FROM response WHERE is_missing=1").fetchone()[0],
            "import_batches_ok": conn.execute("SELECT COUNT(*) FROM import_batch WHERE status='ok'").fetchone()[0],
            "import_batches_error": conn.execute("SELECT COUNT(*) FROM import_batch WHERE status='error'").fetchone()[0],
        },
        "batches": batches,
        "sites": sites,
    }
    (out_dir / "phonology_meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # items index (no response payload)
    items_index = {
        "domain": DOMAIN,
        "imported_at": imported_at,
        "items": [
            {
                "item_code": it["item_code"],
                "standard_form": it["standard_form"],
                "parent_code": it["parent_code"],
                "is_parent": bool(it["is_parent"]),
                "children": children_by_parent.get(it["item_code"], []) if it["is_parent"] else [],
            }
            for it in items
        ],
    }
    (out_dir / "phonology_items.json").write_text(
        json.dumps(items_index, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # full responses as JSONL (streaming-friendly)
    jsonl_path = out_dir / "phonology_responses.jsonl"
    with jsonl_path.open("w", encoding="utf-8") as f:
        q = conn.execute(
            """
            SELECT r.item_code, r.site_id, r.raw_text, r.is_missing, r.source_file, r.source_row,
                   s.province_code, s.province_name, s.raw_header, s.place_name, s.survey_year,
                   i.standard_form, i.parent_code
            FROM response r
            JOIN survey_site s ON s.site_id = r.site_id
            JOIN item i ON i.item_code = r.item_code
            ORDER BY r.item_code, s.province_code, s.raw_header
            """
        )
        for row in q:
            f.write(json.dumps(dict(row), ensure_ascii=False) + "\n")

    # 뷰어용: 항목별 묶음 (파일이 커질 수 있어 parent 단위 compact)
    # 상위 사용: item_code -> [{site fields + raw}]
    by_item: dict[str, list] = defaultdict(list)
    q2 = conn.execute(
        """
        SELECT r.item_code, r.raw_text, r.is_missing,
               s.site_id, s.province_code, s.province_name, s.raw_header, s.place_name, s.survey_year, s.source_file
        FROM response r
        JOIN survey_site s ON s.site_id = r.site_id
        ORDER BY r.item_code, s.province_code, s.raw_header
        """
    )
    for row in q2:
        d = dict(row)
        code = d.pop("item_code")
        by_item[code].append(d)

    # 너무 큰 단일 파일 방지: items 디렉터리에 코드별 저장 + 인덱스
    items_dir = out_dir / "items"
    items_dir.mkdir(exist_ok=True)
    for code, rows in by_item.items():
        # 파일명에 안전한 코드
        safe = code.replace("/", "_")
        (items_dir / f"{safe}.json").write_text(
            json.dumps(
                {
                    "item_code": code,
                    "standard_form": conn.execute(
                        "SELECT standard_form FROM item WHERE item_code=?", (code,)
                    ).fetchone()[0],
                    "parent_code": parent_code_of(code),
                    "domain": DOMAIN,
                    "policy": "raw_excel_only_no_synthetic",
                    "n": len(rows),
                    "responses": rows,
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

    # site_map CSV + YAML-like JSON
    (out_dir / "site_map.json").write_text(
        json.dumps(site_map_rows, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    with (out_dir / "site_map.csv").open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "site_id",
                "province_code",
                "province_name",
                "raw_header",
                "place_name",
                "survey_year",
                "source_file",
            ],
        )
        w.writeheader()
        for row in site_map_rows:
            w.writerow(row)

    return {
        "meta": out_dir / "phonology_meta.json",
        "items": out_dir / "phonology_items.json",
        "jsonl": jsonl_path,
        "items_dir": items_dir,
        "site_map": out_dir / "site_map.json",
    }


def write_report(conn: sqlite3.Connection, out_dir: Path, parsed_list: list, paths: dict, imported_at: str):
    lines = []
    lines.append("# 음운 시트 ETL 검증 리포트")
    lines.append("")
    lines.append(f"- 생성 시각(UTC): `{imported_at}`")
    lines.append(f"- 정책: **엑셀 원문만 적재** (합성·추정 응답 없음)")
    lines.append(f"- 시트: `{SHEET_NAME}`")
    lines.append("")
    lines.append("## 요약")
    lines.append("")
    n_ok = conn.execute("SELECT COUNT(*) FROM import_batch WHERE status='ok'").fetchone()[0]
    n_err = conn.execute("SELECT COUNT(*) FROM import_batch WHERE status='error'").fetchone()[0]
    n_sites = conn.execute("SELECT COUNT(*) FROM survey_site").fetchone()[0]
    n_items = conn.execute("SELECT COUNT(*) FROM item").fetchone()[0]
    n_parents = conn.execute("SELECT COUNT(*) FROM item WHERE is_parent=1").fetchone()[0]
    n_resp = conn.execute("SELECT COUNT(*) FROM response").fetchone()[0]
    n_filled = conn.execute("SELECT COUNT(*) FROM response WHERE is_missing=0").fetchone()[0]
    n_miss = conn.execute("SELECT COUNT(*) FROM response WHERE is_missing=1").fetchone()[0]
    lines.append("| 항목 | 값 |")
    lines.append("|------|-----|")
    lines.append(f"| 성공 파일 | {n_ok} |")
    lines.append(f"| 실패 파일 | {n_err} |")
    lines.append(f"| 지점(site) | {n_sites} |")
    lines.append(f"| 항목(item) | {n_items} |")
    lines.append(f"| 부모 항목 | {n_parents} |")
    lines.append(f"| 응답 셀 | {n_resp} |")
    lines.append(f"| 응답(내용 있음) | {n_filled} |")
    lines.append(f"| 응답(결측 빈칸/*) | {n_miss} |")
    lines.append("")
    lines.append("## 파일별")
    lines.append("")
    lines.append("| 파일 | 도 | 상태 | 지점 | 항목 | 응답 | 채움 | 결측 | 비고 |")
    lines.append("|------|----|------|------|------|------|------|------|------|")
    for r in conn.execute("SELECT * FROM import_batch ORDER BY id"):
        note = r["error"] or ""
        lines.append(
            f"| {r['source_file'][:40]} | {r['province_name'] or '—'} | {r['status']} | "
            f"{r['n_sites'] or 0} | {r['n_items'] or 0} | {r['n_responses'] or 0} | "
            f"{r['n_filled'] or 0} | {r['n_missing'] or 0} | {note} |"
        )
    lines.append("")
    lines.append("## 지점 헤더 (raw_header = 정본)")
    lines.append("")
    for r in conn.execute(
        "SELECT province_name, raw_header, place_name, survey_year, source_file FROM survey_site ORDER BY province_code, raw_header"
    ):
        place = r["place_name"] or "∅"
        year = r["survey_year"] or "∅"
        lines.append(
            f"- **{r['province_name']}** `{r['raw_header']}` → place={place}, year={year}"
        )
    lines.append("")
    lines.append("## 샘플 원문 (임의 추출 아님 — DB 항목 코드 순 채움 셀 12건)")
    lines.append("")
    samples = conn.execute(
        """
        SELECT r.item_code, i.standard_form, s.province_name, s.raw_header, r.raw_text
        FROM response r
        JOIN item i ON i.item_code=r.item_code
        JOIN survey_site s ON s.site_id=r.site_id
        WHERE r.is_missing=0
        ORDER BY r.item_code, s.province_code
        LIMIT 12
        """
    ).fetchall()
    for s in samples:
        raw = (s["raw_text"] or "")[:120].replace("\n", " ")
        lines.append(
            f"- `{s['item_code']}` {s['standard_form'] or ''} / {s['province_name']} · {s['raw_header']}: {raw}"
        )
    lines.append("")
    lines.append("## 산출물")
    lines.append("")
    for k, p in paths.items():
        lines.append(f"- `{k}`: `{p}`")
    lines.append("")
    lines.append("## 주의")
    lines.append("")
    lines.append("- `place_name` / `survey_year` 는 헤더 문자열 파싱 **보조 필드**이며, 서비스 표시 정본은 `raw_header` 와 `raw_text` 입니다.")
    lines.append("- 유지/합류 등 언어학적 판정은 본 ETL에 **포함하지 않습니다**.")
    lines.append("")

    report_path = out_dir / "etl_report.md"
    report_path.write_text("\n".join(lines), encoding="utf-8")
    return report_path


def main():
    ap = argparse.ArgumentParser(description="음운 시트 ETL (실엑셀 원문 only)")
    root = Path(__file__).resolve().parent.parent
    ap.add_argument("--data-dir", type=Path, default=root / "data")
    ap.add_argument("--out-dir", type=Path, default=root / "data" / "processed")
    args = ap.parse_args()

    data_dir: Path = args.data_dir
    out_dir: Path = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    files = sorted(
        [
            p
            for p in data_dir.iterdir()
            if p.suffix.lower() in (".xlsx", ".xls") and not p.name.startswith("~$")
        ],
        key=lambda p: p.name,
    )
    if not files:
        print(f"[error] 엑셀 없음: {data_dir}", file=sys.stderr)
        sys.exit(1)

    print(f"[etl] data_dir={data_dir}")
    print(f"[etl] files={len(files)}")

    parsed_list = []
    for p in files:
        print(f"  reading {p.name} ...", end=" ", flush=True)
        try:
            result = parse_phonology_sheet(p)
        except Exception as e:
            result = {
                "ok": False,
                "error": f"{type(e).__name__}: {e}",
                "source_file": p.name,
                "province_code": detect_province(p.name)[0],
                "province_name": detect_province(p.name)[1],
            }
        parsed_list.append(result)
        if result.get("ok"):
            st = result["stats"]
            print(f"ok sites={st['n_sites']} items={st['n_items']} resp={st['n_responses']}")
        else:
            print(f"FAIL {result.get('error')}")

    db_path = out_dir / "dialect_phonology.db"
    conn = init_db(db_path)
    load_info = load_into_db(conn, parsed_list)
    paths = export_json(conn, out_dir, load_info["site_map_rows"], load_info["imported_at"])
    paths["db"] = db_path
    report = write_report(conn, out_dir, parsed_list, {k: str(v) for k, v in paths.items()}, load_info["imported_at"])
    paths["report"] = report
    conn.close()

    print(f"[etl] db → {db_path}")
    print(f"[etl] report → {report}")
    print(f"[etl] meta → {paths['meta']}")
    print(f"[etl] done totals={load_info['totals']}")


if __name__ == "__main__":
    main()
