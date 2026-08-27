#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
기상도 원자료(엑셀) 양식 정리
=============================
입력 : data/dialect_gisangdo/*.xlsx   (원본 46개 — 건드리지 않는다)
출력 : data/dialect_gisangdo_정리/
         · 기상도_변이조사_양식.xlsx        빈 양식 + 작성 안내
         · 기상도_점검필요.xlsx             값이 규약을 벗어난 행 목록
         · 파일별/{원본명}.xlsx             파일 단위 정리본 (업로드 대상)

제보자 1명 = 파일 1개 관행을 유지한다. 지역·세대·성별은 파일명에서 읽으므로
(etl_awareness_region.py 와 동일) 시트에 출처 열을 두지 않는다. 통합은 업로드 후 DB에서 한다.

원본 서식이 15가지로 흔들려(헤더 표기·열 순서·열 개수) 위치가 아니라 **헤더 이름으로** 열을 찾는다.
출력 폴더를 원본과 분리하는 이유: etl_awareness_region.py 가 data/dialect_gisangdo/*.xlsx 를
통째로 훑기 때문에 같은 폴더에 두면 정리본까지 집계에 섞인다.
"""
import glob, os, re, sys, collections

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC  = os.path.join(BASE, 'data', 'dialect_gisangdo')
OUT  = os.path.join(BASE, 'data', 'dialect_gisangdo_정리')

# 지시받은 삭제 대상 (9열) — ETL(etl_awareness_region.py)도 이 열들을 읽지 않는다
DROP = ['시작시간', '종료시간', '지속시간', '방언형(어절)', '발음정보', '부가정보',
        '개인정보유무', '음성상태', '비고']

# 정리 후 표준 열 (5)
KEEP = ['일련번호', '항목번호', '표제어형', '방언형(기저형)', '사용도/인지도']

REGION_NAMES = {'GG': '경기', 'GW': '강원', 'CB': '충북', 'CN': '충남', 'JB': '전북',
                'JN': '전남', 'GB': '경북', 'GN': '경남', 'JJ': '제주'}


def norm_header(h):
    """표기 흔들림을 표준 이름으로 모은다."""
    s = re.sub(r'\s+', '', str(h or ''))
    if s.startswith('시작시간'):
        return '시작시간'
    if s.startswith('종료시간'):
        return '종료시간'
    if s.startswith('지속시간'):
        return '지속시간'
    alias = {'표제어': '표제어형', '표제어형': '표제어형', '표준어형': '표제어형',
             '일련번호': '일련번호', '항목번호': '항목번호',
             '인지도/사용도': '사용도/인지도', '사용도/인지도': '사용도/인지도',
             '개인정보유무': '개인정보유무', '음성상태': '음성상태', '비고': '비고',
             '방언형(어절)': '방언형(어절)', '방언형(기저형)': '방언형(기저형)',
             '발음정보': '발음정보', '부가정보': '부가정보'}
    return alias.get(s, s)


def parse_name(fname):
    """파일명 {지역2}{연차2}{세대2}{성별1}VE → 출처 4항목 (점검 목록 표시용)."""
    m = re.match(r'^([A-Z]{2})(\d{2})(\d{2})([MF])VE', os.path.basename(fname))
    if not m:
        return ('', '', '', '')
    r, yy, gen, sx = m.groups()
    return (REGION_NAMES.get(r, r), '20' + yy, gen + '대', '여' if sx == 'F' else '남')


def read_rows(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # 첫 시트가 정본이다. 12개 파일이 항목범위별 분할 시트를 함께 갖고 있고
    # 그중 GB2450MVE 는 active 가 마지막 분할 시트라 wb.active 를 쓰면 517행이 유실된다.
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    try:
        hdr = [norm_header(x) for x in next(it)]
    except StopIteration:
        wb.close()
        return [], []
    idx = {}
    for i, name in enumerate(hdr):
        if name and name not in idx:      # 같은 이름이 두 번이면 첫 열을 쓴다
            idx[name] = i
    out = []
    for row in it:
        if row is None or all(v is None or str(v).strip() == '' for v in row):
            continue
        rec = []
        for k in KEEP:
            i = idx.get(k)
            v = row[i] if i is not None and i < len(row) else None
            rec.append('' if v is None else str(v).strip())
        out.append(rec)
    wb.close()
    return hdr, out


def style_header(ws, ncol):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    fill = PatternFill('solid', fgColor='1F3864')
    font = Font(color='FFFFFF', bold=True, size=10)
    thin = Side(style='thin', color='BFBFBF')
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=thin)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


WIDTHS = {'일련번호': 18, '항목번호': 11, '표제어형': 14,
          '방언형(기저형)': 18, '사용도/인지도': 13}


def set_widths(ws, cols):
    from openpyxl.utils import get_column_letter
    for i, name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(name, 12)


def guide_sheet(wb, stats):
    from openpyxl.styles import Font
    ws = wb.create_sheet('작성 안내')
    lines = [
        ('기상도 변이 조사 원자료 — 작성 안내', True),
        ('', False),
        ('◆ 열 구성 (5열)', True),
        ('  일련번호        {지역2}{연차2}{세대2}{성별1}VE + 항목번호. 예) CB2420FVE20101', False),
        ('                  같은 항목의 추가 어형 행은 비워 둡니다.', False),
        ('  항목번호        조사표 항목번호 (필수). 예) 20101', False),
        ('  표제어형        조사 항목의 표준어. 예) 흰자위', False),
        ('  방언형(기저형)  제보자가 응답한 어형의 기저형. 미발화는 * 로 적습니다.', False),
        ('  사용도/인지도   1 사용 · 2 이해 · 3 인지 · 4 무지 (숫자만, 반각)', False),
        ('', False),
        ('◆ 삭제된 열 (9열)', True),
        ('  시작시간 · 종료시간 · 지속시간 · 방언형(어절) · 발음정보 · 부가정보', False),
        ('  개인정보유무 · 음성상태 · 비고', False),
        ('  집계(etl_awareness_region.py)가 읽지 않아 제외했습니다.', False),
        ('  ETL 이 실제로 쓰는 열: 항목번호 · 표준어형 · 방언형(기저형) · 사용도/인지도', False),
        ('', False),
        ('◆ 지켜 주세요', True),
        ('  · 헤더 문구를 바꾸지 마세요. 집계가 헤더 이름으로 열을 찾습니다.', False),
        ('  · 사용도/인지도는 전각(１)이 아니라 반각(1)으로 적습니다.', False),
        ('  · 열을 새로 끼우거나 순서를 바꾸지 마세요.', False),
        ('  · 파일 1개 = 제보자 1명. 파일명 규약을 지켜야 지역·세대·성별이 읽힙니다.', False),
        ('    {지역2}{연차2}{세대2}{성별1}VE.xlsx   예) CB2420FVE = 충북·2024·20대·여', False),
        ('', False),
        ('◆ 원본 현황 (%s 기준)' % stats['date'], True),
        ('  파일 %d개 · 데이터 %s행' % (stats['files'], format(stats['rows'], ',')), False),
        ('  헤더 표기가 %d가지로 갈려 있었습니다 (표제어/표제어형/표준어형 등).' % stats['hdr_variants'], False),
        ('  일련번호가 빈 행 %s행(%.1f%%) — 추가 어형 행입니다. 지역·세대·성별은 파일명에서 읽습니다.'
         % (format(stats['blank_sn'], ','), stats['blank_pct']), False),
    ]
    for r, (txt, bold) in enumerate(lines, start=1):
        c = ws.cell(row=r, column=1, value=txt)
        if bold:
            c.font = Font(bold=True, size=11)
    ws.column_dimensions['A'].width = 100
    return ws


def main():
    import openpyxl
    os.makedirs(OUT, exist_ok=True)
    per_dir = os.path.join(OUT, '파일별')
    os.makedirs(per_dir, exist_ok=True)

    files = sorted(f for f in glob.glob(os.path.join(SRC, '*.xlsx'))
                   if not os.path.basename(f).startswith('~$'))
    if not files:
        sys.exit('원자료를 찾지 못했습니다: %s' % SRC)

    hdr_variants = set()
    total = blank_sn = 0
    issues = []

    for path in files:
        hdr, rows = read_rows(path)
        hdr_variants.add(tuple(hdr))
        name = os.path.basename(path)
        prov = parse_name(path)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '조사자료'          # 시트명도 통일 (원본은 Sheet1/파일명 등 제각각)
        ws.append(KEEP)
        for r in rows:
            ws.append(r)
        style_header(ws, len(KEEP))
        set_widths(ws, KEEP)
        wb.save(os.path.join(per_dir, name))
        wb.close()

        total += len(rows)
        for r in rows:
            if not r[0]:
                blank_sn += 1
            g = r[4]
            if g and g not in ('1', '2', '3', '4', '*'):
                issues.append([name, ' '.join(x for x in prov if x), r[0], r[1], g])

    stats = {'files': len(files), 'rows': total, 'hdr_variants': len(hdr_variants),
             'blank_sn': blank_sn, 'blank_pct': blank_sn / max(1, total) * 100,
             'date': '원본 46개'}

    # 빈 양식
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '조사자료'
    ws.append(KEEP)
    style_header(ws, len(KEEP))
    set_widths(ws, KEEP)
    guide_sheet(wb, stats)
    form_path = os.path.join(OUT, '기상도_변이조사_양식.xlsx')
    wb.save(form_path)
    wb.close()

    # 점검 필요 목록
    issue_path = None
    if issues:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '점검 필요'
        ws.append(['파일', '출처', '일련번호', '항목번호', '사용도/인지도 입력값'])
        for r in issues:
            ws.append(r)
        style_header(ws, 5)
        for L, w in zip('ABCDE', (26, 22, 20, 12, 44)):
            ws.column_dimensions[L].width = w
        issue_path = os.path.join(OUT, '기상도_점검필요.xlsx')
        wb.save(issue_path)
        wb.close()

    print('파일 %d개 / 데이터 %s행 / 원본 헤더 표기 %d가지'
          % (len(files), format(total, ','), len(hdr_variants)))
    print('일련번호 빈 행 %s (%.1f%%) — 출처는 파일명에서 읽으므로 문제 없음'
          % (format(blank_sn, ','), stats['blank_pct']))
    print('점검 필요 %d건' % len(issues))
    print('  %s' % os.path.relpath(per_dir, BASE))
    for p in (form_path, issue_path):
        if p:
            print('  %s  (%.1f KB)' % (os.path.relpath(p, BASE), os.path.getsize(p) / 1024))


if __name__ == '__main__':
    main()
