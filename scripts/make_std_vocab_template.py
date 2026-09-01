#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
검색 표준어 어휘 — 일괄등록 서식(xlsx) 생성

업로더(serve.py: api_std_vocab_bulk)의 실제 동작에 맞춘다.
  · '표준어' 열이 있는 **첫 시트**를 대상으로 삼는다 → 데이터 시트를 맨 앞에 둔다.
    (「안내」를 앞에 두면 그쪽이 대상으로 잡힐 수 있다)
  · 읽는 열은 «표준어»(필수)와 «항목번호»(선택) 둘뿐. 나머지 열은 무시된다.
  · **업로드하면 기존 목록을 통째로 교체한다.** 그래서 예시 행을 넣지 않는다 —
    예시가 든 서식을 그대로 올리면 전체 목록이 그 몇 줄로 바뀐다.
    (빈 채로 올리면 "등록할 어휘가 없습니다"로 거부되므로 빈 서식은 안전하다)

사용
  python3 scripts/make_std_vocab_template.py
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(Path(__file__).resolve().parent))

from make_variant_bulk_template import (  # noqa: E402
    C_HEAD, C_HEAD_REQ, CENTER, F_H2, F_NOTE, F_TITLE, WRAP,
    put, write_header,
)

DEFAULT_OUT = (
    ROOT / "neibis-cms" / "mariadb" / "neibis" / "survey" / "data"
    / "검색표준어어휘_일괄등록_양식.xlsx"
)

COLS = [
    ("항목번호", 18, False),
    ("표준어", 52, True),
]

RULES = [
    "«표준어»만 필수입니다. 한 행에 한 낱말씩, 3행부터 적으세요.",
    "«항목번호»는 조사 항목과 이어 붙일 때 쓰는 참고값입니다. 없으면 비워 두세요.",
    "업로드하면 **현재 등록된 검색 어휘 목록 전체가 이 파일 내용으로 바뀝니다.** "
    "일부만 고치려는 경우에도 남길 낱말을 모두 포함해야 합니다.",
    "같은 낱말이 두 번 나오면 뒤엣것은 무시하고 «중복 n건 제외»로 알려 줍니다.",
    "표준어 칸이 빈 행은 건너뜁니다.",
    "사용여부는 모두 «사용(Y)»으로 등록됩니다. 개별 해제는 목록 화면에서 하세요.",
    "«.xlsx» 파일만 올릴 수 있습니다(.xls 는 받지 않습니다).",
    "이 시트에는 낱말 말고 다른 것을 적지 마세요. 표준어 칸에 적은 메모는 "
    "낱말로 등록됩니다.",
]


def sheet_data(wb: Workbook):
    """데이터 시트 — 반드시 첫 시트여야 한다(업로더가 첫 '표준어' 시트를 고른다)."""
    ws = wb.create_sheet("표준어")
    # 제목에 «표준어»를 쓰지 않는다. 업로더의 _find_col 이 부분일치라
    # 1행 제목이 헤더 행으로 잡히고, 그러면 2행("항목번호")이 낱말로 등록된다.
    put(ws, 1, 1, "검색 어휘 일괄등록 — 한 행 = 한 낱말 (3행부터 입력)",
        font=F_H2, border=False)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    write_header(ws, 2, COLS)
    ws.freeze_panes = "A3"
    # 데이터 영역 아래에 아무것도 쓰지 않는다 — 작성 규칙은 「안내」 시트에만 둔다.
    return ws


def sheet_guide(wb: Workbook):
    ws = wb.create_sheet("안내")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 100

    r = 2
    ws.cell(row=r, column=2, value="검색 표준어 어휘 — 일괄등록 서식").font = F_TITLE
    r += 1
    c = ws.cell(row=r, column=2,
                value="NEIBIS 관리자 > 조사자료 > 검색 표준어 어휘 관리 > 일괄 등록")
    c.font = F_NOTE
    r += 2

    ws.cell(row=r, column=2, value="작성 규칙").font = F_H2
    r += 1
    for i, text in enumerate(RULES, start=1):
        put(ws, r, 1, str(i), align=CENTER, fill="FFF1F5F9")
        cell = put(ws, r, 2, text.replace("**", ""))
        cell.alignment = WRAP
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    c = ws.cell(row=r, column=2,
                value="※ 낱말은 「표준어」 시트에 적습니다. 이 시트는 업로드 대상이 아닙니다.")
    c.font = F_NOTE
    return ws


def build(out: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    sheet_data(wb)      # 첫 시트 = 데이터
    sheet_guide(wb)
    wb.properties.title = "검색 표준어 어휘 일괄등록 서식"
    wb.properties.creator = "NEIBIS"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    args = ap.parse_args()
    print("생성:", build(Path(args.out)))
