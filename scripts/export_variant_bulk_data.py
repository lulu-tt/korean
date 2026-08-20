#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
음운 데이터 → 일괄등록 서식(xlsx) 내보내기

data/processed/dialect_phonology.db 에 적재된 음운 전량을
scripts/make_variant_bulk_template.py 가 정의한 서식 형태로 출력한다.

산출물
  ① 전체 1파일  — ③-1응답(개별) long 형태. response 1행 = 엑셀 1행.
                  경북처럼 일부 코드에 응답행이 없는 도가 있어도 유령 결측이 생기지 않는다.
  ② 도별 9파일  — ③응답(음운) wide 형태. 국립국어원 원본과 열 순서까지 동일.
                  해당 도가 실제로 응답을 가진 코드만 행으로 내보낸다.

두 형태 모두 재업로드 시 원본 DB와 동일해지는 무손실 왕복을 목표로 한다.

사용
  python3 scripts/export_variant_bulk_data.py
  python3 scripts/export_variant_bulk_data.py --out-dir data/exports
"""
from __future__ import annotations

import argparse
import sqlite3
import sys
import unicodedata
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

import etl_phonology as E  # noqa: E402  (site 열 순서 복원에 사용)
from make_variant_bulk_template import (  # noqa: E402
    C_HEAD, C_HEAD_REQ, F_H2, F_NOTE, ITEM_COLS, LONG_COLS, SITE_COLS,
    put, write_header,
)

DB_PATH = ROOT / "data" / "processed" / "dialect_phonology.db"
SRC_DIR = ROOT / "data" / "dialect_phonology_compare"
DEFAULT_OUT = ROOT / "data" / "exports"

PROV_ORDER = ["GG", "GW", "CB", "CN", "JB", "JN", "GB", "GN", "JJ"]
PROV_NAME = {
    "GG": "경기", "GW": "강원", "CB": "충북", "CN": "충남", "JB": "전북",
    "JN": "전남", "GB": "경북", "GN": "경남", "JJ": "제주",
}


# ── 정렬 ──────────────────────────────────────────────────────────
def item_sort_key(code: str):
    """부모(31001) → 하위(31001-0-1, -0-2) 순. 원본 시트 행 순서와 같다."""
    base, _, rest = code.partition("-")
    if not rest:
        return (int(base), -1, -1)
    a, _, b = rest.partition("-")
    return (int(base), int(a), int(b))


# ── DB 읽기 ───────────────────────────────────────────────────────
def load_db(con: sqlite3.Connection) -> dict:
    items = OrderedDict()
    for code, std, parent, is_parent in con.execute(
        "select item_code, standard_form, parent_code, is_parent from item"
    ):
        items[code] = {
            "code": code,
            "std": std or "",
            "parent": parent or "",
            "is_parent": bool(is_parent),
        }
    ordered_codes = sorted(items, key=item_sort_key)

    sites = {}
    for sid, pc, pn, raw, place, year, src in con.execute(
        "select site_id, province_code, province_name, raw_header, place_name,"
        " survey_year, source_file from survey_site"
    ):
        sites[sid] = {
            "site_id": sid, "prov_code": pc or "", "prov_name": pn or "",
            "raw_header": raw, "place": place or "", "year": year or "",
            "source_file": unicodedata.normalize("NFC", src or ""),
        }

    responses = []
    for code, sid, raw, miss in con.execute(
        "select item_code, site_id, raw_text, is_missing from response"
    ):
        responses.append((code, sid, raw, int(miss)))

    return {"items": items, "ordered_codes": ordered_codes,
            "sites": sites, "responses": responses}


def site_column_order(sites: dict) -> dict:
    """
    도별 지점 열 순서를 원본 엑셀 헤더 순서로 복원한다.
    원본을 못 읽으면 (연도, 헤더) 순으로 대체한다.
    """
    order = {p: [] for p in PROV_ORDER}
    by_prov = {p: [] for p in PROV_ORDER}
    for s in sites.values():
        if s["prov_code"] in by_prov:
            by_prov[s["prov_code"]].append(s)

    header_seq = {}
    if SRC_DIR.is_dir():
        for p in SRC_DIR.iterdir():
            if not p.is_file() or p.suffix.lower() not in (".xls", ".xlsx"):
                continue
            pc, _ = E.detect_province(p.name)
            if not pc:
                continue
            r = E.parse_phonology_sheet(p)
            if r.get("ok"):
                header_seq[pc] = list(r["site_headers"])

    for pc, lst in by_prov.items():
        seq = header_seq.get(pc)
        if seq:
            idx = {h: i for i, h in enumerate(seq)}
            lst = sorted(lst, key=lambda s: (idx.get(s["raw_header"], 999), s["raw_header"]))
        else:
            lst = sorted(lst, key=lambda s: (s["year"] or 9999, s["raw_header"]))
        order[pc] = lst
    return order


# ── 공통 시트 ─────────────────────────────────────────────────────
def sheet_summary(wb: Workbook, title: str, rows: list[tuple[str, object]], notes: list[str]):
    ws = wb.create_sheet("요약")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 78

    ws.cell(row=2, column=2, value=title).font = Font(
        name="맑은 고딕", size=14, bold=True, color="FF1F3864"
    )
    r = 4
    ws.cell(row=r, column=2, value="내보낸 내용").font = F_H2
    r += 1
    for k, v in rows:
        put(ws, r, 2, k, fill="FFF1F5F9")
        put(ws, r, 3, v)
        r += 1
    r += 1
    ws.cell(row=r, column=2, value="유의사항").font = F_H2
    r += 1
    for n in notes:
        c = ws.cell(row=r, column=2, value="· " + n)
        c.font = F_NOTE
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        r += 1
    return ws


def sheet_items(wb: Workbook, data: dict, codes: list[str]):
    ws = wb.create_sheet("①항목")
    put(ws, 1, 1, "① 항목(비교단위) — 한 행 = 한 항목", font=F_H2, border=False)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ITEM_COLS))
    write_header(ws, 2, ITEM_COLS)

    r = 3
    ordinal = {}
    for code in codes:
        it = data["items"][code]
        band = code[:3]
        if it["is_parent"]:
            ordinal[band] = ordinal.get(band, 0) + 1
            sort_ordr = ordinal[band]
        else:
            sort_ordr = int(code.rsplit("-", 1)[-1])
        ws.append([
            code, it["std"], "부모" if it["is_parent"] else "환경",
            "" if it["is_parent"] else it["parent"],
            band, sort_ordr, "Y", "",
        ])
        r += 1
    ws.freeze_panes = "A3"
    return ws, r - 3


def sheet_sites(wb: Workbook, site_list: list[dict]):
    ws = wb.create_sheet("②조사지점")
    put(ws, 1, 1, "② 조사지점 — 한 행 = 한 지점", font=F_H2, border=False)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(SITE_COLS))
    write_header(ws, 2, SITE_COLS)
    for s in site_list:
        ws.append([
            s["raw_header"], s["prov_code"], s["prov_name"], s["place"],
            s["year"], s["site_id"], "Y", "",
        ])
    ws.freeze_panes = "A3"
    return ws, len(site_list)


# ── 전체 파일: ③-1 long ───────────────────────────────────────────
def build_all(data: dict, order: dict, out: Path) -> tuple[Path, int]:
    wb = Workbook()
    wb.remove(wb.active)

    site_list = [s for pc in PROV_ORDER for s in order[pc]]
    resp_by_key = {(c, sid): (raw, miss) for c, sid, raw, miss in data["responses"]}

    n_resp = len(data["responses"])
    n_missing = sum(1 for *_x, m in data["responses"] if m)
    sheet_summary(
        wb,
        "변이형(음운) 비교 — 전체 데이터 (일괄등록 서식)",
        [
            ("출처", "data/processed/dialect_phonology.db (음운 시트만)"),
            ("항목", f"{len(data['items']):,}건 (부모 {sum(1 for v in data['items'].values() if v['is_parent']):,} / 환경 {sum(1 for v in data['items'].values() if not v['is_parent']):,})"),
            ("조사지점", f"{len(site_list):,}개 (9개 도)"),
            ("응답", f"{n_resp:,}건 (채움 {n_resp - n_missing:,} / 결측 {n_missing:,})"),
            ("응답 시트 형태", "③-1 개별(long) — 한 행 = 한 응답"),
        ],
        [
            "응답을 long 형태로 담았습니다. DB의 response 1행이 엑셀 1행이므로 무손실입니다.",
            "54지점을 wide 한 시트로 펴면 경북에 응답행이 없는 462개 코드가 빈칸이 되고, 재업로드 시 결측 2,772건이 새로 생깁니다. 그래서 전체 파일은 long을 씁니다.",
            "도별 wide 형태(원본과 동일)가 필요하면 같은 폴더의 「도별」 파일을 사용하세요.",
            "지점은 «지점열 헤더(원문)»으로 식별합니다. 「②조사지점」의 지점ID는 기존 지점을 그대로 가리키기 위한 값입니다.",
        ],
    )

    sheet_items(wb, data, data["ordered_codes"])
    sheet_sites(wb, site_list)

    ws = wb.create_sheet("③-1응답(개별)")
    put(ws, 1, 1, "③-1 지점 응답(개별형) — 한 행 = 한 응답", font=F_H2, border=False)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(LONG_COLS))
    write_header(ws, 2, LONG_COLS)

    written = 0
    for code in data["ordered_codes"]:
        for s in site_list:
            hit = resp_by_key.get((code, s["site_id"]))
            if hit is None:
                continue
            raw, miss = hit
            ws.append([
                code, s["raw_header"], s["prov_code"], s["place"], s["year"],
                raw, "Y" if miss else "N", "",
            ])
            written += 1
    ws.freeze_panes = "A3"

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out, written


# ── 도별 파일: ③ wide ─────────────────────────────────────────────
def build_province(data: dict, order: dict, pc: str, out_dir: Path,
                   lean: bool = True) -> tuple[Path, int, int]:
    sites = order[pc]
    sids = [s["site_id"] for s in sites]
    sid_set = set(sids)

    resp = {}
    for code, sid, raw, miss in data["responses"]:
        if sid in sid_set:
            resp[(code, sid)] = raw

    codes = [c for c in data["ordered_codes"]
             if any((c, sid) in resp for sid in sids)]

    wb = Workbook()
    wb.remove(wb.active)

    n_cells = len(codes) * len(sids)
    n_filled = sum(1 for c in codes for sid in sids if (resp.get((c, sid)) or "") not in ("", "*"))

    # 경량(기본): ③응답 시트만. 업로드에 필요한 정보는 모두 ③에서 유도된다.
    #   항목코드·표준어 = A·B열 / 부모·환경 = 코드 형식 / 지점명·연도 = 지점열 헤더 / 도 = 파일명
    # --with-master: 요약·①항목·②조사지점을 함께 담는다(마스터 통제·관리필드 입력용).
    if not lean:
        sheet_summary(
            wb,
            f"변이형(음운) 비교 — {PROV_NAME[pc]} (일괄등록 서식)",
            [
                ("출처", "data/processed/dialect_phonology.db (음운 시트만)"),
                ("도", f"{PROV_NAME[pc]} ({pc})"),
                ("항목", f"{len(codes):,}건"),
                ("조사지점", f"{len(sites)}개 — " + ", ".join(s["raw_header"] for s in sites)),
                ("응답", f"{n_cells:,}셀 (채움 {n_filled:,} / 결측 {n_cells - n_filled:,})"),
                ("응답 시트 형태", "③ 통합(wide) — 국립국어원 원본과 동일"),
            ],
            [
                "「③응답(음운)」 시트는 원본 통합자료와 레이아웃·열 순서가 같습니다. 1행 제목 / 2행 헤더 / 3행부터 데이터.",
                "이 도가 실제로 응답을 가진 항목만 행으로 담았습니다. 없는 행을 임의로 추가하지 마세요.",
                "빈칸과 «*»는 결측입니다. 원문 그대로이므로 값을 채워 넣지 마세요.",
                "일괄등록 팝업에서 «③ 지점 응답(통합자료)» 유형으로 올리세요.",
            ],
        )
        sheet_items(wb, data, codes)
        sheet_sites(wb, sites)

    ws = wb.create_sheet("③응답(음운)")
    put(
        ws, 1, 1,
        f"③ 지점 응답(통합자료형) — {PROV_NAME[pc]} / 행 = 항목, 열 = 조사지점",
        font=F_H2, border=False,
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(sites))

    header = ["항목번호", "표준어"] + [s["raw_header"] for s in sites]
    for i, title in enumerate(header, start=1):
        put(ws, 2, i, title, font=Font(name="맑은 고딕", size=10, bold=True, color="FFFFFFFF"),
            fill=C_HEAD_REQ if i <= 2 else C_HEAD)
        ws.column_dimensions[get_column_letter(i)].width = 16 if i <= 2 else 22
    ws.column_dimensions["B"].width = 20
    ws.row_dimensions[2].height = 30

    for code in codes:
        ws.append([code, data["items"][code]["std"]]
                  + [resp.get((code, sid), "") for sid in sids])
    ws.freeze_panes = "C3"

    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"변이형비교_음운_{PROV_NAME[pc]}.xlsx"
    wb.save(out)
    return out, len(codes), n_cells


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out-dir", default=str(DEFAULT_OUT))
    ap.add_argument(
        "--with-master", action="store_true",
        help="도별 파일에 요약·①항목·②조사지점 시트를 함께 담는다 (기본: ③응답 시트만)",
    )
    args = ap.parse_args()
    out_dir = Path(args.out_dir)
    lean = not args.with_master

    if not DB_PATH.is_file():
        raise SystemExit(f"DB 없음: {DB_PATH}")

    con = sqlite3.connect(f"file:{DB_PATH}?mode=ro", uri=True)
    data = load_db(con)
    order = site_column_order(data["sites"])

    p, n = build_all(data, order, out_dir / "변이형비교_음운_전체.xlsx")
    print(f"전체: {p.name}  응답 {n:,}행")

    total_cells = 0
    print(f"도별 ({'③응답 시트만' if lean else '요약+①+②+③'}):")
    for pc in PROV_ORDER:
        if not order[pc]:
            continue
        fp, n_items, n_cells = build_province(data, order, pc, out_dir / "도별", lean=lean)
        total_cells += n_cells
        print(f"  {PROV_NAME[pc]}: {fp.name}  항목 {n_items:,} / 응답셀 {n_cells:,}"
              f"  ({fp.stat().st_size / 1024:.0f}KB)")
    print(f"도별 응답셀 합계: {total_cells:,}")


if __name__ == "__main__":
    main()
