#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
기상도 원자료 → 전용 테이블 적재
================================
입력 : data/dialect_gisangdo_정리/파일별/*.xlsx  (정리 양식 5열)
       data/dialect_gisangdo/*.xlsx             (옛 14열 원본도 그대로 받는다)
출력 : data/gisangdo.db  (SQLite — CUBRID 이관 전 검증용)

DDL 은 scripts/sql/wb_weather_ddl.sql 과 같은 구조다. 열을 헤더 이름으로 찾으므로
정리 양식(5열)과 옛 원본(14열·열순서 뒤바뀜)을 모두 읽는다.

  python3 scripts/load_weather_xlsx.py            # 정리본 적재
  python3 scripts/load_weather_xlsx.py --raw      # 옛 원본 적재
"""
import argparse, glob, os, re, sqlite3, sys

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TIDY = os.path.join(BASE, 'data', 'dialect_gisangdo_정리', '파일별')
RAW  = os.path.join(BASE, 'data', 'dialect_gisangdo')
DB   = os.path.join(BASE, 'data', 'gisangdo.db')

REGION_NAMES = {'GG': '경기', 'GW': '강원', 'CB': '충북', 'CN': '충남', 'JB': '전북',
                'JN': '전남', 'GB': '경북', 'GN': '경남', 'JJ': '제주'}
VALID_GRADE = {'1', '2', '3', '4'}

DDL = """
CREATE TABLE IF NOT EXISTS wb_weather_file (
  weather_file_id INTEGER PRIMARY KEY, file_nm TEXT NOT NULL UNIQUE,
  region_cd TEXT NOT NULL, region_nm TEXT, research_year INTEGER,
  research_degree TEXT, generation INTEGER, sex TEXT,
  row_cnt INTEGER DEFAULT 0, item_cnt INTEGER DEFAULT 0, src_layout TEXT,
  use_yn TEXT DEFAULT 'Y', reg_id TEXT, reg_dt TEXT, upt_id TEXT, upt_dt TEXT);
CREATE INDEX IF NOT EXISTS ix_wwf_region ON wb_weather_file (region_cd, generation, sex);

CREATE TABLE IF NOT EXISTS wb_weather_response (
  response_id INTEGER PRIMARY KEY, weather_file_id INTEGER NOT NULL,
  line_no INTEGER NOT NULL, serial_no TEXT, item_cd TEXT NOT NULL, item_base TEXT,
  headword TEXT, dialect_form TEXT, grade TEXT, grade_valid_yn TEXT DEFAULT 'N',
  use_yn TEXT DEFAULT 'Y', reg_dt TEXT,
  FOREIGN KEY (weather_file_id) REFERENCES wb_weather_file (weather_file_id));
CREATE INDEX IF NOT EXISTS ix_wwr_file ON wb_weather_response (weather_file_id, line_no);
CREATE INDEX IF NOT EXISTS ix_wwr_item ON wb_weather_response (item_base, grade_valid_yn);
CREATE INDEX IF NOT EXISTS ix_wwr_head ON wb_weather_response (headword);

CREATE TABLE IF NOT EXISTS wb_weather_region_stat (
  region_cd TEXT NOT NULL, item_base TEXT NOT NULL, headword TEXT,
  state TEXT NOT NULL, use_rate REAL, informant_cnt INTEGER DEFAULT 0,
  dialect_cnt INTEGER DEFAULT 0, std_only_yn TEXT DEFAULT 'N', core_yn TEXT DEFAULT 'N',
  note TEXT, calc_dt TEXT,
  PRIMARY KEY (region_cd, item_base));

CREATE TABLE IF NOT EXISTS wb_weather_std_form (
  std_form_id INTEGER PRIMARY KEY, item_base TEXT NOT NULL, std_form TEXT NOT NULL,
  memo TEXT, use_yn TEXT DEFAULT 'Y', reg_id TEXT, reg_dt TEXT,
  UNIQUE (item_base, std_form));
"""


def norm_header(h):
    s = re.sub(r'\s+', '', str(h or ''))
    if s.startswith('시작시간'):
        return '시작시간'
    if s.startswith('종료시간'):
        return '종료시간'
    if s.startswith('지속시간'):
        return '지속시간'
    alias = {'표제어': '표제어형', '표제어형': '표제어형', '표준어형': '표제어형',
             '인지도/사용도': '사용도/인지도', '사용도/인지도': '사용도/인지도'}
    return alias.get(s, s)


def parse_name(fname):
    m = re.match(r'^([A-Z]{2})(\d{2})(\d{2})([MF])VE', os.path.basename(fname))
    if not m:
        return None
    r, yy, gen, sx = m.groups()
    return {'region_cd': r, 'region_nm': REGION_NAMES.get(r, r),
            'research_year': 2000 + int(yy), 'research_degree': yy,
            'generation': int(gen), 'sex': sx}


def item_base(code):
    m = re.match(r'^(\d{5})', str(code or '').strip())
    return m.group(1) if m else None


def read_file(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # 첫 시트가 정본 (etl_awareness_region.py 와 동일 규칙). wb.active 는 쓰면 안 된다.
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    try:
        hdr = [norm_header(x) for x in next(it)]
    except StopIteration:
        wb.close()
        return 'UNKNOWN', []
    idx = {}
    for i, name in enumerate(hdr):
        if name and name not in idx:
            idx[name] = i
    layout = 'V5' if len(hdr) <= 6 else 'RAW'

    def g(row, key):
        i = idx.get(key)
        if i is None or i >= len(row):
            return ''
        v = row[i]
        return '' if v is None else str(v).strip()

    out = []
    n = 0
    for row in it:
        if row is None or all(v is None or str(v).strip() == '' for v in row):
            continue
        n += 1
        out.append({'line_no': n, 'serial_no': g(row, '일련번호') or None,
                    'item_cd': g(row, '항목번호'), 'headword': g(row, '표제어형'),
                    'dialect_form': g(row, '방언형(기저형)'), 'grade': g(row, '사용도/인지도')})
    wb.close()
    return layout, out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--raw', action='store_true', help='정리본 대신 옛 원본을 적재')
    ap.add_argument('--db', default=DB)
    a = ap.parse_args()

    src = RAW if a.raw else TIDY
    files = sorted(f for f in glob.glob(os.path.join(src, '*.xlsx'))
                   if not os.path.basename(f).startswith('~$'))
    if not files:
        sys.exit('원자료를 찾지 못했습니다: %s' % src)

    if os.path.exists(a.db):
        os.remove(a.db)
    con = sqlite3.connect(a.db)
    con.executescript(DDL)

    fid = rid = 0
    skipped = []
    for path in files:
        meta = parse_name(path)
        if not meta:
            skipped.append(os.path.basename(path))
            continue
        layout, rows = read_file(path)
        rows = [r for r in rows if r['item_cd']]        # 항목번호 없는 행은 적재 제외
        fid += 1
        items = {item_base(r['item_cd']) for r in rows}
        con.execute("""INSERT INTO wb_weather_file
            (weather_file_id,file_nm,region_cd,region_nm,research_year,research_degree,
             generation,sex,row_cnt,item_cnt,src_layout,use_yn,reg_id,reg_dt)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,'Y','loader',datetime('now'))""",
            (fid, os.path.basename(path), meta['region_cd'], meta['region_nm'],
             meta['research_year'], meta['research_degree'], meta['generation'],
             meta['sex'], len(rows), len(items - {None}), layout))
        batch = []
        for r in rows:
            rid += 1
            gv = 'Y' if r['grade'] in VALID_GRADE else 'N'
            batch.append((rid, fid, r['line_no'], r['serial_no'], r['item_cd'],
                          item_base(r['item_cd']), r['headword'], r['dialect_form'],
                          r['grade'] or None, gv))
        con.executemany("""INSERT INTO wb_weather_response
            (response_id,weather_file_id,line_no,serial_no,item_cd,item_base,
             headword,dialect_form,grade,grade_valid_yn,use_yn,reg_dt)
            VALUES (?,?,?,?,?,?,?,?,?,?,'Y',datetime('now'))""", batch)
    con.commit()

    q = lambda s: list(con.execute(s))[0][0]
    print('DB: %s' % os.path.relpath(a.db, BASE))
    print('  wb_weather_file     %6d' % q('SELECT COUNT(*) FROM wb_weather_file'))
    print('  wb_weather_response %6d' % q('SELECT COUNT(*) FROM wb_weather_response'))
    print('  등급 유효(1~4)      %6d' % q("SELECT COUNT(*) FROM wb_weather_response WHERE grade_valid_yn='Y'"))
    print('  일련번호 NULL       %6d' % q('SELECT COUNT(*) FROM wb_weather_response WHERE serial_no IS NULL'))
    print('  서로 다른 항목      %6d' % q('SELECT COUNT(DISTINCT item_base) FROM wb_weather_response'))
    print('  서로 다른 표제어    %6d' % q("SELECT COUNT(DISTINCT headword) FROM wb_weather_response WHERE headword<>''"))
    if skipped:
        print('  파일명 규약 불일치로 제외: %s' % ', '.join(skipped))
    con.close()


if __name__ == '__main__':
    main()
