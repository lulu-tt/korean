#!/usr/bin/env python3
"""
NEIBIS 프로토타입 서버.
- neibis-cms/ 정적 파일 서빙
- .do → 같은 이름 .html 매핑
- 로컬 dialect_local.db JSON API
"""
from __future__ import annotations

import functools
import http.server
import io
import json
import os
import re
import socketserver
import sqlite3
import time
import unicodedata
import urllib.parse
import wave
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent
USER_MAP_ROOT = Path(os.environ.get("USER_MAP_ROOT", "/Users/aaa/inseq/korean"))
PORT = int(os.environ.get("PORT", "8877"))

# 로컬 지역어 DB (환경변수로 덮어쓰기 가능)
DB_PATH = Path(
    os.environ.get(
        "DIALECT_DB",
        str(Path.home() / "inseq/korean/dialect_local.db"),
    )
)
# 워크스페이스 상대 경로 fallback
if not DB_PATH.is_file():
    alt = Path("/Users/aaa/inseq/korean/dialect_local.db")
    if alt.is_file():
        DB_PATH = alt

# 구술발화(ELAN .eaf + .wav) 원천 데이터 루트. 하위 폴더(예: JB2320FUT)를 스캔.
ORAL_DATA_ROOT = Path(
    os.environ.get("ORAL_DATA_ROOT", "/Users/aaa/inseq/korean/data")
)


def db_connect():
    if not DB_PATH.is_file():
        raise FileNotFoundError(f"DB not found: {DB_PATH}")
    con = sqlite3.connect(str(DB_PATH))
    con.row_factory = sqlite3.Row
    return con


def _fmt_epoch_ms(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    # pure epoch ms
    if s.isdigit() and len(s) >= 12:
        try:
            return datetime.fromtimestamp(int(s) / 1000).strftime("%Y-%m-%d")
        except Exception:
            return None
    # already date-like
    if len(s) >= 8 and s[0:4].isdigit():
        return s[:10].replace("/", "-").replace(".", "-")
    return None


def fmt_reg_dt(reg_dt, upt_dt=None, file_nm=None) -> str:
    """등록일: reg_dt → upt_dt → 파일명(에폭 ms) 순으로 표시."""
    for cand in (reg_dt, upt_dt):
        formatted = _fmt_epoch_ms(cand)
        if formatted:
            return formatted
    if file_nm:
        base = str(file_nm).rsplit(".", 1)[0]
        formatted = _fmt_epoch_ms(base)
        if formatted:
            return formatted
    return "-"


# ── 구술발화(ELAN) 파서 ────────────────────────────────────────────────
# 국립국어원 지역어 구술발화 표준 대주제(항목 블록 코드 → 주제명)
ORAL_TOPIC_MAP = {
    "10100": "조사 마을의 환경과 배경",
    "10200": "일생 의례와 경험",
    "10300": "생업 활동과 경제 생활",
    "10400": "의생활",
    "10500": "식생활",
    "10600": "주생활",
    "10700": "질병과 건강",
    "10800": "세시 풍속과 여가 문화",
    "10900": "언어·자유 발화",
}

# 시도명 → 프런트 select 코드 (oral.html 검색폼과 동일)
SIDO_NM_TO_CD = {
    "서울": "11", "부산": "26", "대구": "27", "인천": "28", "광주": "29",
    "대전": "30", "울산": "31", "세종": "36", "경기": "41", "강원": "42",
    "충청북도": "43", "충북": "43", "충청남도": "44", "충남": "44",
    "전라북도": "45", "전북": "45", "전라남도": "46", "전남": "46",
    "경상북도": "47", "경북": "47", "경상남도": "48", "경남": "48",
    "제주": "50",
}


def _sido_cd_of(region: str) -> str:
    """조사지역 문자열 앞부분에서 시도 코드 추정."""
    r = (region or "").strip()
    for nm, cd in SIDO_NM_TO_CD.items():
        if r.startswith(nm):
            return cd
    return ""


def _ms_to_hms(ms) -> str:
    try:
        total = int(ms) // 1000
    except (TypeError, ValueError):
        return "00:00:00"
    h, rem = divmod(total, 3600)
    m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def _oral_id_of(eaf_path: Path) -> str:
    """파일 식별자: '_업로드용' 접미사와 확장자를 뗀 베이스명 (예: JB2320FUT10100).

    macOS 파일명은 NFD(자모 분해)로 저장되므로 NFC 정규화 후 접미사 제거.
    """
    stem = unicodedata.normalize("NFC", eaf_path.stem)
    return re.sub(r"_업로드용$", "", stem)


_ORAL_CACHE: dict[str, dict] = {}  # id -> {"mtime": float, "data": {...}}


def _meta_key(k: str) -> str:
    """'조사 지역' · '제보자 이름'처럼 띄어쓰기가 흔들려도 같은 키로 본다."""
    return re.sub(r"\s+", "", k or "")


def _meta_pick(meta: dict, *names: str) -> str:
    """띄어쓰기 흔들림과 별칭(주제보자이름 등)을 모두 받아 첫 값을 돌려준다."""
    for n in names:
        v = meta.get(_meta_key(n))
        if v:
            return str(v).strip()
    return ""


def parse_eaf(eaf_path: Path) -> dict:
    """ELAN .eaf 파싱 → 메타데이터 + 시간정렬 세그먼트.

    반환:
      {
        oralId, region, sidoCd, year, contentCode, topic,
        informant, sex, birth, durationMs, segmentCount,
        segments: [{seq, speaker, startMs, endMs, form, std, item}],
        tiers: {tierId: annotationCount, ...},
      }
    형태음소전사/표준어대역은 같은 화자 내에서 시작 TIME_SLOT 기준으로 짝지음.
    """
    root = ET.parse(str(eaf_path)).getroot()

    # 1) TIME_SLOT_ID → ms
    tsmap: dict[str, int] = {}
    for ts in root.iter("TIME_SLOT"):
        try:
            tsmap[ts.get("TIME_SLOT_ID")] = int(ts.get("TIME_VALUE") or 0)
        except (TypeError, ValueError):
            tsmap[ts.get("TIME_SLOT_ID")] = 0

    def tier_annotations(tier) -> list[dict]:
        out = []
        for ann in tier.findall(".//ALIGNABLE_ANNOTATION"):
            v = ann.find("ANNOTATION_VALUE")
            out.append({
                "start": tsmap.get(ann.get("TIME_SLOT_REF1"), 0),
                "end": tsmap.get(ann.get("TIME_SLOT_REF2"), 0),
                "text": (v.text or "").strip() if v is not None else "",
            })
        return out

    tiers_by_id: dict[str, list] = {}
    tier_counts: dict[str, int] = {}
    meta_raw = ""
    item_anns: list[dict] = []
    voice_anns: list[dict] = []   # 음성상태 (시간정렬)
    remark_anns: list[dict] = []  # 비고 (시간정렬)
    annotators: list[str] = []   # 전사자(ANNOTATOR) 수집
    for tier in root.findall("TIER"):
        tid = tier.get("TIER_ID") or ""
        anns = tier_annotations(tier)
        tiers_by_id[tid] = anns
        tier_counts[tid] = len(anns)
        ann_name = (tier.get("ANNOTATOR") or "").strip()
        if ann_name and ann_name not in annotators:
            annotators.append(ann_name)
        if tid == "메타데이터" and anns:
            meta_raw = anns[0]["text"]
        if tid == "항목번호":
            item_anns = anns
        if tid == "음성상태":
            voice_anns = [a for a in anns if a["text"]]
        if tid == "비고":
            remark_anns = [a for a in anns if a["text"]]

    # 2) 메타데이터 파싱 (key:value / 구분)
    meta: dict[str, str] = {}
    for kv in meta_raw.split("/"):
        if ":" in kv:
            k, val = kv.split(":", 1)
            meta[_meta_key(k)] = val.strip()

    region = _meta_pick(meta, "조사지역")
    content_code = _meta_pick(meta, "조사내용")
    oral_id = _oral_id_of(eaf_path)
    if not content_code:
        m = re.search(r"(\d{5})", oral_id)
        content_code = m.group(1) if m else ""

    # 3) 화자별 형태음소전사 ↔ 표준어대역 짝짓기
    def pair_speaker(speaker: str) -> list[dict]:
        form_anns = tiers_by_id.get(f"{speaker}(형태음소전사)", [])
        std_anns = tiers_by_id.get(f"{speaker}(표준어대역)", [])
        std_by_start = {a["start"]: a["text"] for a in std_anns}
        segs = []
        for i, a in enumerate(form_anns):
            std = std_by_start.get(a["start"])
            if std is None and i < len(std_anns):  # 시작슬롯 불일치 시 순번 폴백
                std = std_anns[i]["text"]
            segs.append({
                "speaker": speaker,
                "startMs": a["start"],
                "endMs": a["end"],
                "form": a["text"],
                "std": std or "",
            })
        return segs

    # 전사자료에 실제로 등장한 화자 — 계층 이름에서 뽑는다.
    #  · 제보자가 2명 이상인지 여기서만 알 수 있다(메타데이터 계층엔 주제보자 이름만 적힌다)
    #  · 조사자/제보자1/제보자2 를 하드코딩하면 '제보자'(숫자 없음)·'제보자3' 이상의 발화가 통째로 유실된다
    speakers = []
    for tid, cnt in tier_counts.items():
        m2 = re.match(r"^(.*?)\s*\(형태음소전사\)$", tid)
        if m2 and cnt:
            speakers.append({"name": m2.group(1).strip(), "lines": cnt})
    speakers.sort(key=lambda x: (0 if x["name"].startswith("조사자") else 1, x["name"]))

    segments = []
    for sp in speakers:
        segments += pair_speaker(sp["name"])
    segments.sort(key=lambda s: (s["startMs"], 0 if s["speaker"] == "조사자" else 1))

    # 항목번호를 시간구간으로 매핑 (희소)
    def item_at(start_ms: int) -> str:
        cur = ""
        for it in item_anns:
            if it["start"] <= start_ms:
                cur = it["text"]
        return cur

    for i, s in enumerate(segments, 1):
        s["seq"] = i
        s["item"] = item_at(s["startMs"])

    duration_ms = max((tsmap.values() or [0]))

    # 전사자(ANNOTATOR) / 조사자 — 있을 때만 채움
    #  · 전사자: 메타데이터 '전사자' 키 우선, 없으면 ANNOTATOR 속성(단일이면 그 값)
    #  · 조사자: 메타데이터 '조사자' 키가 있을 때만 (EAF에 이름 없으면 빈 값)
    transcriber = _meta_pick(meta, "전사자", "전사자명")
    if not transcriber and len(annotators) == 1:
        transcriber = annotators[0]
    investigator = _meta_pick(meta, "조사자", "조사자명")

    return {
        "oralId": oral_id,
        "region": region,
        "sidoCd": _sido_cd_of(region),
        "year": _meta_pick(meta, "조사연도"),
        "contentCode": content_code,
        "topic": ORAL_TOPIC_MAP.get(content_code, content_code or "-"),
        # 파일마다 '제보자 이름' / '주제보자이름' 으로 갈린다 — 둘 다 받는다
        "informant": _meta_pick(meta, "제보자이름", "주제보자이름"),
        "sex": _meta_pick(meta, "성별"),
        "birth": _meta_pick(meta, "출생연도"),
        "speakers": speakers,
        "investigator": investigator,
        "transcriber": transcriber,
        "durationMs": duration_ms,
        "segmentCount": len(segments),
        "segments": segments,
        "tiers": tier_counts,
        # 시간정렬 계층(항목번호/음성상태/비고) — 세그먼트 매핑용 스팬
        "itemSpans": [{"start": a["start"], "end": a["end"], "text": a["text"].strip()} for a in item_anns],
        "voiceSpans": [{"start": a["start"], "end": a["end"], "text": a["text"].strip()} for a in voice_anns],
        "remarkSpans": [{"start": a["start"], "end": a["end"], "text": a["text"].strip()} for a in remark_anns],
    }


def parse_trs(trs_path: Path) -> dict:
    """Transcriber .trs (XML) 파싱 → parse_eaf 와 같은 구조.

    Sync 시각 + `10604@ 방언 {표준어}` / `10604# …` 라인을 세그먼트로 만든다.
    첫 Sync의 헤더에서 제보자·지역·연도 등을 추정한다.
    """
    raw = trs_path.read_bytes()
    text = None
    for enc in ("utf-8", "cp949", "euc-kr"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode("utf-8", "replace")
    try:
        root = ET.fromstring(text)
    except ET.ParseError:
        # DTD / 제어문자 등 — 선언부만 걷어내고 재시도
        cleaned = re.sub(r"<!DOCTYPE[^>]*>", "", text, count=1, flags=re.I)
        root = ET.fromstring(cleaned)

    trans = root if (root.tag or "").endswith("Trans") else root.find(".//Trans")
    if trans is None:
        trans = root
    scribe = (trans.get("scribe") or "").strip()
    if scribe.lower() in ("", "(unknown)", "unknown"):
        scribe = ""
    audio_attr = (trans.get("audio_filename") or "").strip()

    section_end = 0.0
    for sec in root.iter("Section"):
        try:
            section_end = max(section_end, float(sec.get("endTime") or 0))
        except (TypeError, ValueError):
            pass

    syncs: list[tuple[float, str]] = []
    for turn in root.iter("Turn"):
        for ch in list(turn):
            if (ch.tag or "").endswith("Sync"):
                try:
                    t = float(ch.get("time") or 0)
                except (TypeError, ValueError):
                    t = 0.0
                syncs.append((t, (ch.tail or "").strip()))

    header = ""
    line_re = re.compile(r"^(?:(\d{3,6})\s*)?([@#]\d*)\s*(.*)$", re.S)
    segments: list[dict] = []
    items_seen: list[str] = []
    for i, (start, body) in enumerate(syncs):
        if not body:
            continue
        end = syncs[i + 1][0] if i + 1 < len(syncs) else section_end
        m = line_re.match(body)
        if not m:
            if not header:
                header = body
            continue
        item, mark, rest = m.group(1) or "", m.group(2), (m.group(3) or "").strip()
        if mark.startswith("@"):
            speaker = "조사자"
        else:
            n = mark[1:] or "1"
            speaker = "제보자" + n
        std = ""
        bm = re.search(r"\{(.*)\}\s*$", rest, re.S)
        if bm:
            std = bm.group(1).strip()
            rest = rest[: bm.start()].strip()
        start_ms = int(round(start * 1000))
        end_ms = int(round(end * 1000)) if end and end > start else start_ms
        segments.append({
            "speaker": speaker,
            "startMs": start_ms,
            "endMs": end_ms,
            "form": rest,
            "std": std,
            "item": item,
        })
        if item:
            items_seen.append(item)

    for i, s in enumerate(segments, 1):
        s["seq"] = i

    meta: dict[str, str] = {}
    ym = re.search(r"(19\d{2}|20\d{2})\s*년", header)
    if ym:
        meta["year"] = ym.group(1)
    nm = re.search(r"([가-힣]{2,10})\s*\(([^)]*세[^)]*)\)", header)
    if nm:
        meta["informant"] = nm.group(1)
        info = nm.group(2)
        parts = [p.strip() for p in info.split(",") if p.strip()]
        if parts:
            meta["region"] = parts[0]
        am = re.search(r"(\d+)\s*세", info)
        if am:
            meta["age"] = am.group(1)
        if re.search(r"여(성|자)?", info) and not re.search(r"남(성|자)?", info):
            meta["sex"] = "여"
        elif re.search(r"남(성|자)?", info):
            meta["sex"] = "남"
    if not meta.get("region"):
        rm = re.search(r"\(([^)]*(?:특별|광역|자치)?(?:시|도)[^)]*)\)\s*$", header)
        if rm:
            meta["region"] = rm.group(1).strip()
    year = meta.get("year") or ""
    age = meta.get("age") or ""
    birth = ""
    if year.isdigit() and age.isdigit():
        birth = str(int(year) - int(age) + 1)

    oral_id = _oral_id_of(trs_path)
    content_code = ""
    if items_seen:
        content_code = items_seen[0][:3] + "00" if len(items_seen[0]) >= 3 else items_seen[0]
    if not content_code:
        m = re.search(r"(\d{5})", oral_id)
        content_code = m.group(1) if m else ""
    duration_ms = int(round(section_end * 1000)) if section_end else (
        max((s["endMs"] for s in segments), default=0)
    )
    topic = ORAL_TOPIC_MAP.get(content_code, content_code or "-")
    if header:
        rest = header
        if nm:
            rest = header[nm.end():].strip()
        rest = re.sub(r"(?:19|20)\d{2}\s*년.*$", "", rest).strip(" ,.-")
        if rest:
            topic = rest

    # 화자 목록 — .trs 는 계층이 없어 세그먼트 화자를 센다
    spk_counts: dict[str, int] = {}
    for sg in segments:
        spk_counts[sg["speaker"]] = spk_counts.get(sg["speaker"], 0) + 1
    speakers = [{"name": k, "lines": v} for k, v in spk_counts.items()]
    speakers.sort(key=lambda x: (0 if x["name"].startswith("조사자") else 1, x["name"]))

    return {
        "oralId": oral_id or audio_attr,
        "region": meta.get("region", ""),
        "sidoCd": _sido_cd_of(meta.get("region", "")),
        "year": year,
        "contentCode": content_code,
        "topic": topic,
        "informant": meta.get("informant", ""),
        "sex": meta.get("sex", ""),
        "birth": birth,
        "speakers": speakers,
        "investigator": "",
        "transcriber": scribe,
        "durationMs": duration_ms,
        "segmentCount": len(segments),
        "segments": segments,
        "tiers": {},
        "itemSpans": [],
        "voiceSpans": [],
        "remarkSpans": [],
        "format": "trs",
    }


def parse_oral_transcript(path: Path) -> dict:
    """업로드된 전사자료(.eaf / .trs) 파싱."""
    ext = path.suffix.lower()
    if ext == ".trs":
        return parse_trs(path)
    if ext == ".eaf":
        return parse_eaf(path)
    raise ValueError(f"지원하지 않는 전사 형식: {path.suffix}")


def _scan_oral_files() -> list[Path]:
    """데이터 루트 하위의 모든 .eaf 경로 (정렬)."""
    if not ORAL_DATA_ROOT.is_dir():
        return []
    return sorted(ORAL_DATA_ROOT.rglob("*.eaf"))


def _load_oral(eaf_path: Path) -> dict:
    """mtime 캐시가 적용된 parse_eaf."""
    oral_id = _oral_id_of(eaf_path)
    try:
        mtime = eaf_path.stat().st_mtime
    except OSError:
        mtime = 0.0
    cached = _ORAL_CACHE.get(oral_id)
    if cached and cached["mtime"] == mtime and cached.get("path") == str(eaf_path):
        return cached["data"]
    data = parse_eaf(eaf_path)
    # WAV 매칭 (같은 폴더, 같은 베이스, .wav)
    wav = eaf_path.with_name(_oral_id_of(eaf_path) + ".wav")
    data["hasMedia"] = wav.is_file()
    try:
        data["relPath"] = str(eaf_path.relative_to(ORAL_DATA_ROOT))
    except ValueError:
        data["relPath"] = eaf_path.name
    _ORAL_CACHE[oral_id] = {"mtime": mtime, "path": str(eaf_path), "data": data}
    return data


# 시도 코드 → region_nm 접두어 (wb_research_region.region_nm 매칭용)
SIDO_CD_TO_PREFIX = {
    "11": "서울", "26": "부산", "27": "대구", "28": "인천", "29": "광주",
    "30": "대전", "31": "울산", "36": "세종", "41": "경기", "42": "강원",
    "43": "충청북도", "44": "충청남도", "45": "전라북도", "46": "전라남도",
    "47": "경상북도", "48": "경상남도", "50": "제주",
}


def _sec_to_hms(sec) -> str:
    try:
        total = int(float(sec))
    except (TypeError, ValueError):
        return "00:00:00"
    h, rem = divmod(total, 3600)
    m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def _parse_trs_line(txt: str) -> dict:
    """.trs 라인 마커 파싱 (첫 발화 턴 기준).
    @=조사자, #n=제보자n / %1=형태음소전사, %2 {..}=표준어대역.
    한 행에 여러 턴이 들어간 복합 라인(2009 .trs 등)은 singleTurn=False로 표시하고
    첫 턴만 파싱해 표시한다(저장 시 손실 방지를 위해 편집 대상에서 제외).
    """
    raw = txt or ""
    t = raw.strip()
    speaker = "조사자"
    m = re.match(r"^\s*(@|#\d*)", t)
    if m:
        tok = m.group(1)
        speaker = "조사자" if tok == "@" else ("제보자" + (tok[1:] or "1"))
        t = t[m.end():].strip()
    # 표준어대역: 첫 번째 %2 {...} (non-greedy → 첫 턴만)
    m2 = re.search(r"%2\s*\{(.*?)\}", t, re.S)
    std = m2.group(1).strip() if m2 else ""
    # 형태음소전사: 첫 %1 부터 첫 %2 직전까지
    m1 = re.search(r"%1\s*(.*?)(?=\s*%2|$)", t, re.S)
    if m1:
        form = m1.group(1).strip()
    else:
        form = (t[: m2.start()].strip() if m2 else t).strip()
    if not form and not std:
        form = raw.strip()
    single_turn = raw.count("%1") == 1
    return {"speaker": speaker, "form": form, "std": std, "singleTurn": single_turn}


def _oral_media_id(audio_filename, trs_file_nm) -> str:
    """미디어 파일 베이스명 (audio_filename 우선, 없으면 파일명에서 확장자 제거)."""
    if audio_filename:
        return str(audio_filename)
    base = str(trs_file_nm or "")
    return base.rsplit(".", 1)[0] if "." in base else base


def api_oral_list(qs: dict) -> dict:
    """구술발화조사자료 목록 — 기존 DB(wb_trs_file_talk) 기반."""
    def q1(key, default=""):
        return (qs.get(key) or [default])[0].strip()

    try:
        page = max(1, int(q1("page", "1")))
    except ValueError:
        page = 1
    try:
        page_size = int(q1("pageSize", "10"))
    except ValueError:
        page_size = 10
    if page_size not in (10, 50, 100, 200, 300):
        page_size = 10

    sido = q1("sido") or q1("sidoCd")
    year = q1("year") or q1("researchYear")
    sex = q1("sex")           # '' | man | woman | wom | 남 | 여
    q = q1("q") or q1("searchValue")
    topic_upper = q1("topicUpper") or q1("topic")   # 발화 주제 대분류
    topic_sub = q1("topicSub")                      # 발화 주제 소분류

    where: list[str] = []
    params: list = []
    if year:
        where.append("IFNULL(rr.research_year,'') = ?")
        params.append(year)
    if sido:
        pref = SIDO_CD_TO_PREFIX.get(sido)
        if pref:
            where.append("(IFNULL(rr.region_nm,'') LIKE ? OR IFNULL(rr.sigungu_nm,'') LIKE ?)")
            params += [pref + "%", pref + "%"]
    if q:
        like = "%" + q + "%"
        where.append(
            "(IFNULL(f.trs_file_nm,'') LIKE ? OR IFNULL(f.upper_headword,'') LIKE ? "
            "OR IFNULL(f.headword,'') LIKE ? OR IFNULL(rr.region_nm,'') LIKE ?)"
        )
        params += [like, like, like, like]
    if topic_upper:
        where.append("IFNULL(f.upper_headword,'') = ?")
        params.append(topic_upper)
    if topic_sub:
        where.append("IFNULL(f.headword,'') = ?")
        params.append(topic_sub)
    sex_code = "1" if sex in ("man", "남") else ("0" if sex in ("woman", "wom", "여") else "")  # wb_source: 0=여,1=남
    if sex_code:
        where.append(
            "EXISTS (SELECT 1 FROM wb_source s WHERE s.research_region_id = f.research_region_id AND s.sex = ?)"
        )
        params.append(sex_code)
    wh = ("WHERE " + " AND ".join(where)) if where else ""
    offset = (page - 1) * page_size

    with db_connect() as con:
        total = con.execute(
            f"""SELECT COUNT(*) FROM wb_trs_file_talk f
                LEFT JOIN wb_research_region rr ON rr.research_region_id = f.research_region_id {wh}""",
            params,
        ).fetchone()[0]
        rows = con.execute(
            f"""
            SELECT f.trs_id, f.trs_file_nm, f.wave_file_nm, f.audio_filename,
                   f.upper_headword, f.headword, f.trs_time, f.use_yn,
                   f.research_degree, f.remark,
                   rr.region_nm, rr.sigungu_nm, rr.research_year
            FROM wb_trs_file_talk f
            LEFT JOIN wb_research_region rr ON rr.research_region_id = f.research_region_id
            {wh}
            ORDER BY CAST(f.trs_id AS INTEGER) DESC
            LIMIT ? OFFSET ?
            """,
            params + [page_size, offset],
        ).fetchall()

        trs_ids = [str(r["trs_id"]) for r in rows if r["trs_id"] is not None]
        durmap: dict[str, float] = {}
        segmap: dict[str, int] = {}
        if trs_ids:
            ph = ",".join("?" * len(trs_ids))
            for nr in con.execute(
                f"""SELECT trs_id, MAX(CAST(end_time AS REAL)) dur, COUNT(*) cnt
                    FROM wb_trs_line_talk
                    WHERE trs_line_se='text' AND trs_id IN ({ph})
                    GROUP BY trs_id""",
                trs_ids,
            ).fetchall():
                durmap[str(nr["trs_id"])] = nr["dur"] or 0
                segmap[str(nr["trs_id"])] = nr["cnt"] or 0

    items = []
    for idx, r in enumerate(rows, start=offset + 1):
        tid = str(r["trs_id"])
        dur_sec = durmap.get(tid, 0)
        audio_id = _oral_media_id(r["audio_filename"], r["trs_file_nm"])
        items.append({
            "no": idx,
            "oralId": tid,
            "year": r["research_year"] or "",
            "region": r["region_nm"] or r["sigungu_nm"] or "-",
            "topicUpper": r["upper_headword"] or "",
            "topicSub": r["headword"] or "",
            # 하위 호환(구 화면): 대분류 우선
            "topic": r["upper_headword"] or r["headword"] or "-",
            "fileName": r["trs_file_nm"] or (audio_id + ".eaf"),
            "duration": _sec_to_hms(dur_sec),
            "durationMs": int((dur_sec or 0) * 1000),
            "segmentCount": segmap.get(tid, 0),
            "researchDegree": r["research_degree"] or "",
            "hasMedia": _find_oral_media(audio_id, "wav") is not None,
            "openYn": (r["use_yn"] or "Y"),
            "audioState": "이상없음",
        })

    total_pages = max(1, (total + page_size - 1) // page_size)
    return {
        "ok": True,
        "total": total,
        "page": page,
        "pageSize": page_size,
        "totalPages": total_pages,
        "list": items,
        "db": str(DB_PATH),
    }


def api_oral_topics(qs: dict) -> dict:
    """발화 주제 분류(대분류 → 소분류) — 검색폼 2단 select용. DB 실제 값 기준."""
    with db_connect() as con:
        rows = con.execute(
            """SELECT IFNULL(upper_headword,'') AS up, IFNULL(headword,'') AS sub,
                      COUNT(*) AS cnt
                 FROM wb_trs_file_talk
                GROUP BY up, sub"""
        ).fetchall()

    tree: dict[str, dict] = {}
    for r in rows:
        up, sub, cnt = r["up"].strip(), r["sub"].strip(), r["cnt"] or 0
        if not up:
            continue
        node = tree.setdefault(up, {"name": up, "count": 0, "subs": {}})
        node["count"] += cnt
        if sub:
            node["subs"][sub] = node["subs"].get(sub, 0) + cnt

    out = []
    for up in sorted(tree, key=lambda k: (-tree[k]["count"], k)):
        node = tree[up]
        out.append({
            "name": node["name"],
            "count": node["count"],
            "subs": [{"name": nm, "count": node["subs"][nm]}
                     for nm in sorted(node["subs"], key=lambda k: (-node["subs"][k], k))],
        })
    return {"ok": True, "list": out}


def _find_source_eaf(trs_file_nm: str) -> Path | None:
    """trs_file_nm(예: 'CB2370MUT10200.eaf') → 디스크의 원천 EAF ('_업로드용' 포함/미포함)."""
    base = re.sub(r"\.eaf$", "", str(trs_file_nm or ""), flags=re.I).strip()
    if not base:
        return None
    for d in (ORAL_DATA_ROOT, ORAL_UPLOAD_DIR):
        if not d or not d.exists():
            continue
        for name in (base, base + "_업로드용"):
            p = d / (name + ".eaf")
            if p.is_file():
                return p
    for p in _scan_oral_files():
        if re.sub(r"_업로드용$", "", p.stem) == base:
            return p
    return None


def _headword_topic(con, headword_no) -> dict:
    """표제어번호 → {topic(주제), subTopic(소주제), headword(표제어/질문)}. 없으면 빈 dict."""
    hn = str(headword_no or "").strip()
    if not hn.isdigit():
        return {}
    row = con.execute(
        """SELECT topic, sub_topic FROM kd_topic
           WHERE CAST(? AS INTEGER) BETWEEN CAST(headword_start_no AS INTEGER) AND CAST(headword_end_no AS INTEGER)
           ORDER BY (CASE WHEN IFNULL(sub_topic,'')<>'' THEN 0 ELSE 1 END),
                    (CAST(headword_end_no AS INTEGER) - CAST(headword_start_no AS INTEGER))
           LIMIT 1""",
        (hn,)).fetchone()
    hw = con.execute("SELECT meaning FROM kd_headword_no WHERE headword_no=?", (hn,)).fetchone()
    out = {}
    if row and (row["topic"] or row["sub_topic"]):
        out["topic"] = row["topic"] or ""
        out["subTopic"] = row["sub_topic"] or ""
    if hw and hw["meaning"]:
        out["headword"] = hw["meaning"]
    return out


def _overlay_at(spans: list, start_ms: int) -> str:
    """항목번호처럼 '시작 이하의 마지막 값' 매칭 (구간 라벨)."""
    cur = ""
    for s in spans:
        if s["start"] <= start_ms:
            cur = s["text"]
    return cur


def _overlays_over(spans: list, start_ms: int, end_ms: int) -> list:
    """세그먼트 [start,end] 와 시간이 겹치는 스팬들의 텍스트."""
    out = []
    for s in spans:
        if s["start"] < end_ms and s["end"] > start_ms:
            if s["text"] and s["text"] not in out:
                out.append(s["text"])
    return out


def api_oral_detail(qs: dict) -> dict:
    """구술발화 상세 — wb_trs_line_talk 라인을 세그먼트로 반환."""
    trs_id = (qs.get("id") or qs.get("trsId") or qs.get("oralId") or [""])[0].strip()
    if not trs_id:
        return {"ok": False, "message": "id 파라미터가 필요합니다."}

    with db_connect() as con:
        f = con.execute(
            """
            SELECT f.trs_id, f.research_region_id, f.trs_file_nm, f.wave_file_nm,
                   f.audio_filename, f.upper_headword, f.headword, f.use_yn, f.remark,
                   rr.region_nm, rr.sigungu_nm, rr.research_year,
                   rr.researcher, rr.transcriber, rr.mic, rr.recorder
            FROM wb_trs_file_talk f
            LEFT JOIN wb_research_region rr ON rr.research_region_id = f.research_region_id
            WHERE f.trs_id = ?
            """,
            (trs_id,),
        ).fetchone()
        if not f:
            return {"ok": False, "message": f"해당 자료를 찾을 수 없습니다: {trs_id}"}
        src = con.execute(
            """SELECT name, sex, age, birth FROM wb_source
               WHERE research_region_id = ?
               ORDER BY CAST(source_id AS INTEGER) LIMIT 1""",
            (f["research_region_id"],),
        ).fetchone()
        lines = con.execute(
            """SELECT trs_line_id, trs_line_no, headword_no, trs_line, start_time, end_time
               FROM wb_trs_line_talk
               WHERE trs_id = ? AND trs_line_se = 'text'
               ORDER BY CAST(trs_line_no AS INTEGER)""",
            (trs_id,),
        ).fetchall()

    def to_ms(v):
        try:
            return int(float(v) * 1000)
        except (TypeError, ValueError):
            return 0

    segments = []
    for i, ln in enumerate(lines, 1):
        p = _parse_trs_line(ln["trs_line"])
        segments.append({
            "seq": i,
            "trsLineId": str(ln["trs_line_id"]) if ln["trs_line_id"] is not None else "",
            "speaker": p["speaker"],
            "startMs": to_ms(ln["start_time"]),
            "endMs": to_ms(ln["end_time"]),
            "form": p["form"],
            "std": p["std"],
            "singleTurn": p["singleTurn"],
            "raw": ln["trs_line"] or "",
            "item": ln["trs_line_no"] or "",
            "headwordNo": str(ln["headword_no"]).strip() if ln["headword_no"] not in (None, "") else "",
        })
    dur_ms = max((s["endMs"] for s in segments), default=0)

    # ── 부가정보 보강: 표제어번호(항목번호)·표제어·주제·소주제 + 음성상태·비고 ──
    #  · 표제어번호: DB headword_no(구형 .trs) 우선, 없으면 원천 EAF의 '항목번호' 계층(시간정렬)
    #  · 음성상태/비고: 원천 EAF의 해당 계층(시간정렬)을 세그먼트 시간과 겹쳐 매핑
    eaf_path = _find_source_eaf(f["trs_file_nm"] or "")
    item_spans, voice_spans, remark_spans = [], [], []
    if eaf_path:
        try:
            rec = _load_oral(eaf_path)
            item_spans = rec.get("itemSpans", [])
            voice_spans = rec.get("voiceSpans", [])
            remark_spans = rec.get("remarkSpans", [])
        except Exception:
            pass
    with db_connect() as con2:
        for s in segments:
            hw = s["headwordNo"] or _overlay_at(item_spans, s["startMs"])
            hw = re.sub(r"[^0-9]", "", str(hw))   # '10201 ' 같은 공백 제거
            s["headwordNo"] = hw
            if hw:
                info = _headword_topic(con2, hw)
                s["headword"] = info.get("headword", "")
                s["subject"] = info.get("topic", "")
                s["subTopic"] = info.get("subTopic", "")
            else:
                s["headword"] = s["subject"] = s["subTopic"] = ""
            s["voiceStatus"] = "; ".join(_overlays_over(voice_spans, s["startMs"], s["endMs"]))
            s["remark"] = "; ".join(_overlays_over(remark_spans, s["startMs"], s["endMs"]))

    sex_map = {"0": "여", "1": "남"}  # wb_source: 0=여,1=남
    sex_raw = (src["sex"] if src else "") or ""
    audio_id = _oral_media_id(f["audio_filename"], f["trs_file_nm"])
    has_media = _find_oral_media(audio_id, "wav") is not None

    return {
        "ok": True,
        "oralId": str(f["trs_id"]),
        "region": f["region_nm"] or f["sigungu_nm"] or "-",
        "year": f["research_year"] or "",
        "topic": f["upper_headword"] or f["headword"] or "-",
        "informant": (src["name"] if src else "") or "",
        "sex": sex_map.get(sex_raw, sex_raw),
        "age": (src["age"] if src else "") or "",
        "birth": (src["birth"] if src else "") or "",
        "researcher": f["researcher"] or "",
        "transcriber": f["transcriber"] or "",
        "duration": _sec_to_hms(dur_ms / 1000),
        "durationMs": dur_ms,
        "segmentCount": len(segments),
        "hasMedia": has_media,
        "mediaUrl": f"/oral-media/{audio_id}.wav" if has_media else "",
        "fileName": f["trs_file_nm"] or (audio_id + ".eaf"),
        "wavFileName": f["wave_file_nm"] or (audio_id + ".wav"),
        "segments": segments,
    }


def api_oral_meta(qs: dict) -> dict:
    """구술발화 수정폼 메타 — 조사지역/제보자/지역내 파일목록 (wb_research_region 등)."""
    trs_id = (qs.get("id") or qs.get("trsId") or qs.get("oralId") or [""])[0].strip()
    if not trs_id:
        return {"ok": False, "message": "id 파라미터가 필요합니다."}

    with db_connect() as con:
        f = con.execute(
            """SELECT trs_id, research_region_id, upper_headword, headword,
                      use_yn, trs_file_nm, audio_filename, start_dialect_no, end_dialect_no
               FROM wb_trs_file_talk WHERE trs_id = ?""",
            (trs_id,),
        ).fetchone()
        if not f:
            return {"ok": False, "message": f"해당 자료를 찾을 수 없습니다: {trs_id}"}
        rrid = f["research_region_id"]
        rr = con.execute(
            "SELECT * FROM wb_research_region WHERE research_region_id = ?", (rrid,)
        ).fetchone()
        sources = con.execute(
            """SELECT source_id, se, name, sex, age, birth,
                      IFNULL(residence,'') residence, IFNULL(birth_place,'') birth_place,
                      IFNULL(job,'') job, IFNULL(education,'') education
               FROM wb_source
               WHERE research_region_id = ? ORDER BY se, CAST(source_id AS INTEGER)""",
            (rrid,),
        ).fetchall()
        files = con.execute(
            """SELECT trs_id, trs_file_nm, trs_time, wave_file_nm, wave_time, ver, upper_headword
               FROM wb_trs_file_talk WHERE research_region_id = ?
               ORDER BY CAST(trs_id AS INTEGER)""",
            (rrid,),
        ).fetchall()
        # 항목번호: 이 파일 전사 라인의 headword_no (distinct), 없으면 파일 범위(start~end)
        item_nos = [str(r[0]) for r in con.execute(
            """SELECT DISTINCT headword_no FROM wb_trs_line_talk
               WHERE trs_id = ? AND IFNULL(headword_no,'')<>''
               ORDER BY CAST(headword_no AS INTEGER)""", (trs_id,)).fetchall()]
        if not item_nos and (f["start_dialect_no"] or f["end_dialect_no"]):
            item_nos = [x for x in [f["start_dialect_no"], f["end_dialect_no"]] if x]

    sex_map = {"0": "여", "1": "남"}  # wb_source: 0=여,1=남

    def src_dict(s):
        return {
            "sourceId": str(s["source_id"]) if s["source_id"] is not None else "",
            "name": s["name"] or "",
            "sex": sex_map.get(s["sex"] or "", s["sex"] or ""),
            "age": s["age"] or "",
            "birth": s["birth"] or "",
            "residence": s["residence"] or "",
            "birthPlace": s["birth_place"] or "",
            "job": s["job"] or "",
            "education": s["education"] or "",
            "label": f"{s['name'] or ''} ( {sex_map.get(s['sex'] or '', '')} {s['age'] or ''} )".strip(),
        }

    mains = [src_dict(s) for s in sources if (s["se"] in ("0", None, ""))]
    subs = [src_dict(s) for s in sources if (s["se"] not in ("0", None, ""))]

    file_list = []
    for r in files:
        file_list.append({
            "trsId": str(r["trs_id"]),
            "trsFileNm": r["trs_file_nm"] or "",
            "trsTime": _sec_to_hms(r["trs_time"]) if r["trs_time"] else "00:00:00",
            "waveFileNm": r["wave_file_nm"] or "",
            "waveTime": _sec_to_hms(r["wave_time"]) if r["wave_time"] else "00:00:00",
            "ver": r["ver"] or "-",
            "topic": r["upper_headword"] or "",
        })

    region = {}
    if rr:
        keys = rr.keys()
        sigungu = rr["sigungu_nm"] if "sigungu_nm" in keys else ""
        region = {
            "researchRegionId": str(rrid or ""),
            "regionNm": rr["region_nm"] or "",
            "sigunguNm": sigungu or "",
            "researchYear": rr["research_year"] or "",
            "regionRemark": rr["region_remark"] or "",
            "researchPlace": rr["research_place"] or "",
            "firstPlace": rr["first_place"] or "",
            "researcher": rr["researcher"] or "",
            "transcriber": rr["transcriber"] or "",
            "mic": rr["mic"] or "",
            "recorder": rr["recorder"] or "",
            "fileUniqueness": rr["file_uniqueness"] or "",
            "legalRegionCode": rr["legal_region_code"] or "",
            "regionNmYn": rr["region_nm_yn"] or "",
            "startDate": _fmt_epoch_ms(rr["start_date"]) or "",
            "endDate": _fmt_epoch_ms(rr["end_date"]) or "",
        }

    return {
        "ok": True,
        "trsId": str(f["trs_id"]),
        "headword": f["upper_headword"] or f["headword"] or "",
        "useYn": f["use_yn"] or "Y",
        "fileName": f["trs_file_nm"] or "",
        "region": region,
        "mainSources": mains,
        "subSources": subs,
        "files": file_list,
        "itemNos": item_nos,
    }


def _date_to_epoch_ms(v) -> str:
    """'YYYY-MM-DD' → epoch ms 문자열. 이미 epoch ms거나 비면 그대로/빈값."""
    s = str(v or "").strip()
    if not s:
        return ""
    if s.isdigit():
        return s
    m = re.match(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", s)
    if not m:
        return s
    try:
        dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        return str(int(dt.timestamp() * 1000))
    except ValueError:
        return s


def api_oral_save_meta(body: dict) -> dict:
    """구술발화 수정폼 메타 저장 — wb_trs_file_talk + wb_research_region UPDATE."""
    trs_id = str(body.get("trsId") or "").strip()
    if not trs_id:
        return {"ok": False, "message": "trsId가 필요합니다."}
    region = body.get("region") or {}
    headword = (body.get("headword") or "").strip()
    use_yn = "N" if str(body.get("useYn") or "Y").upper() == "N" else "Y"

    with db_connect() as con:
        f = con.execute(
            "SELECT research_region_id FROM wb_trs_file_talk WHERE trs_id = ?", (trs_id,)
        ).fetchone()
        if not f:
            return {"ok": False, "message": f"해당 자료를 찾을 수 없습니다: {trs_id}"}
        rrid = f["research_region_id"]

        con.execute(
            "UPDATE wb_trs_file_talk SET upper_headword = ?, use_yn = ? WHERE trs_id = ?",
            (headword, use_yn, trs_id),
        )
        if rrid is not None:
            con.execute(
                """UPDATE wb_research_region SET
                       research_year = ?, region_remark = ?, research_place = ?,
                       first_place = ?, researcher = ?, transcriber = ?,
                       mic = ?, recorder = ?, file_uniqueness = ?,
                       start_date = ?, end_date = ?
                   WHERE research_region_id = ?""",
                (
                    (region.get("researchYear") or "").strip(),
                    (region.get("regionRemark") or "").strip(),
                    (region.get("researchPlace") or "").strip(),
                    (region.get("firstPlace") or "").strip(),
                    (region.get("researcher") or "").strip(),
                    (region.get("transcriber") or "").strip(),
                    (region.get("mic") or "").strip(),
                    (region.get("recorder") or "").strip(),
                    (region.get("fileUniqueness") or "").strip(),
                    _date_to_epoch_ms(region.get("startDate")),
                    _date_to_epoch_ms(region.get("endDate")),
                    rrid,
                ),
            )
        # 제보자(주/부) 저장 — 요청에 sources가 있으면 해당 조사지역 제보자를 교체
        if rrid is not None and ("mainSources" in body or "subSources" in body):
            sex_code = {"남": "1", "여": "0", "man": "1", "woman": "0", "wom": "0"}
            con.execute("DELETE FROM wb_source WHERE research_region_id = ?", (rrid,))
            sid = _next_id(con, "wb_source", "source_id")

            def ins(lst, base_se):
                nonlocal sid
                for i, s in enumerate(lst or []):
                    nm = (s.get("name") or "").strip()
                    if not nm:
                        continue
                    con.execute(
                        """INSERT INTO wb_source
                           (source_id, research_region_id, se, name, sex, age, birth,
                            residence, birth_place, job, education)
                           VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                        (
                            str(sid), str(rrid), str(base_se + i), nm,
                            sex_code.get((s.get("sex") or "").strip(), (s.get("sex") or "").strip()),
                            (s.get("age") or "").strip(), (s.get("birth") or "").strip(),
                            (s.get("residence") or "").strip(), (s.get("birthPlace") or "").strip(),
                            (s.get("job") or "").strip(), (s.get("education") or "").strip(),
                        ),
                    )
                    sid += 1

            ins(body.get("mainSources"), 0)
            ins(body.get("subSources"), 1)
        con.commit()
    return {"ok": True, "trsId": trs_id, "message": "저장되었습니다."}


def api_oral_raw(qs: dict) -> dict:
    """trs 보기 — 레거시와 동일하게 원본 .eaf/.trs 파일 소스를 그대로 반환.

    원본 파일이 로컬(ORAL_DATA_ROOT)에 있으면 파일 내용을 그대로(source=file),
    없으면 wb_trs_line_talk 행을 이어붙인 DB 재구성본(source=db)을 반환한다.
    """
    trs_id = (qs.get("id") or qs.get("trsId") or qs.get("oralId") or [""])[0].strip()
    if not trs_id:
        return {"ok": False, "message": "id 파라미터가 필요합니다."}
    with db_connect() as con:
        f = con.execute(
            "SELECT trs_file_nm, audio_filename FROM wb_trs_file_talk WHERE trs_id = ?",
            (trs_id,),
        ).fetchone()
        if not f:
            return {"ok": False, "message": f"해당 자료를 찾을 수 없습니다: {trs_id}"}
        file_name = f["trs_file_nm"] or ("trs " + trs_id)
        audio_id = _oral_media_id(f["audio_filename"], f["trs_file_nm"])

        # 1) 원본 파일 우선 (.eaf → .trs)
        for ext in ("eaf", "trs"):
            p = _find_oral_media(audio_id, ext)
            if p:
                try:
                    raw = p.read_text(encoding="utf-8", errors="replace")
                except OSError:
                    continue
                return {
                    "ok": True,
                    "trsId": trs_id,
                    "fileName": p.name,
                    "source": "file",
                    "format": ext,
                    "lineCount": raw.count("\n") + 1,
                    "raw": raw,
                }

        # 2) 폴백: DB 라인 재구성
        rows = con.execute(
            """SELECT trs_line FROM wb_trs_line_talk
               WHERE trs_id = ? ORDER BY CAST(trs_line_id AS INTEGER)""",
            (trs_id,),
        ).fetchall()
    raw = "\n".join((r["trs_line"] or "") for r in rows)
    fmt = "xml" if raw.lstrip().startswith("<?xml") else "text"
    return {
        "ok": True,
        "trsId": trs_id,
        "fileName": file_name,
        "source": "db",
        "format": fmt,
        "lineCount": len(rows),
        "raw": raw,
    }


def api_oral_save_raw(body: dict) -> dict:
    """trs 편집 저장 — 원본 .eaf/.trs 파일을 그대로 덮어쓴다(백업 후).

    안전장치: 최초 1회 <파일>.orig(원본 보존) + 매 저장 <파일>.bak(직전본) 생성.
    로컬에 원본 파일이 없는 레코드(source=db)는 저장 불가.
    """
    trs_id = str(body.get("trsId") or "").strip()
    raw = body.get("raw")
    if not trs_id:
        return {"ok": False, "message": "trsId가 필요합니다."}
    if raw is None:
        return {"ok": False, "message": "저장할 내용이 없습니다."}

    with db_connect() as con:
        f = con.execute(
            "SELECT trs_file_nm, audio_filename FROM wb_trs_file_talk WHERE trs_id = ?",
            (trs_id,),
        ).fetchone()
        if not f:
            return {"ok": False, "message": f"해당 자료를 찾을 수 없습니다: {trs_id}"}
    audio_id = _oral_media_id(f["audio_filename"], f["trs_file_nm"])

    target = None
    for ext in ("eaf", "trs"):
        p = _find_oral_media(audio_id, ext)
        if p:
            target = p
            break
    if target is None:
        return {"ok": False, "message": "원본 파일이 로컬에 없어 저장할 수 없습니다(DB 재구성본은 편집 불가)."}

    data = str(raw)
    # 백업: 원본 보존(최초 1회) + 직전본(매 저장)
    orig = Path(str(target) + ".orig")
    if not orig.exists():
        orig.write_bytes(target.read_bytes())
    Path(str(target) + ".bak").write_bytes(target.read_bytes())
    target.write_text(data, encoding="utf-8")
    return {
        "ok": True,
        "trsId": trs_id,
        "fileName": target.name,
        "bytes": len(data.encode("utf-8")),
        "message": "원본 파일에 저장되었습니다.",
    }


ORAL_UPLOAD_DIR = ORAL_DATA_ROOT / "uploads"


# ── .xlsx 리더 (외부 의존성 없이 표준 라이브러리만 사용) ────────────────
_XL_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_XL_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"


def _xl_col_index(letters: str) -> int:
    """엑셀 열 문자('A','B','AA')를 0부터 시작하는 인덱스로 변환."""
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch.upper()) - 64)
    return max(0, n - 1)


def _read_xlsx(data: bytes) -> list:
    """.xlsx → [(시트명, [[셀값, ...], ...]), ...]"""
    zf = zipfile.ZipFile(io.BytesIO(data))
    names = zf.namelist()

    shared = []
    if "xl/sharedStrings.xml" in names:
        for si in ET.fromstring(zf.read("xl/sharedStrings.xml")):
            shared.append("".join(t.text or "" for t in si.iter(_XL_NS + "t")))

    rels = {}
    if "xl/_rels/workbook.xml.rels" in names:
        for r in ET.fromstring(zf.read("xl/_rels/workbook.xml.rels")):
            rels[r.get("Id")] = r.get("Target") or ""

    sheets = []
    for sh in ET.fromstring(zf.read("xl/workbook.xml")).iter(_XL_NS + "sheet"):
        target = rels.get(sh.get(_XL_REL + "id"), "")
        path = "xl/" + target.lstrip("/")
        path = path.replace("xl/xl/", "xl/")
        sheets.append((sh.get("name") or "", path))

    out = []
    for name, path in sheets:
        if path not in names:
            continue
        rows = []
        for row in ET.fromstring(zf.read(path)).iter(_XL_NS + "row"):
            cells = []
            for c in row.iter(_XL_NS + "c"):
                ref = c.get("r") or ""
                col = _xl_col_index("".join(ch for ch in ref if ch.isalpha()))
                ctype_attr = c.get("t")
                v = c.find(_XL_NS + "v")
                if ctype_attr == "s" and v is not None and (v.text or "").isdigit():
                    val = shared[int(v.text)] if int(v.text) < len(shared) else ""
                elif ctype_attr == "inlineStr":
                    node = c.find(_XL_NS + "is")
                    val = "".join(t.text or "" for t in node.iter(_XL_NS + "t")) if node is not None else ""
                else:
                    val = (v.text or "") if v is not None else ""
                while len(cells) < col:
                    cells.append("")
                cells.append(str(val).strip())
            rows.append(cells)
        out.append((name, rows))
    return out


# ── 검색 표준어 어휘 일괄 등록 ──────────────────────────────────────────
STD_VOCAB_JSON = ROOT / "mariadb/neibis/survey/data/search-std-vocab.json"


def _find_col(header: list, *keywords) -> int:
    """헤더 행에서 키워드를 포함하는 열의 인덱스를 찾는다. 없으면 -1."""
    for i, cell in enumerate(header):
        flat = str(cell or "").replace(" ", "")
        if any(k in flat for k in keywords):
            return i
    return -1


def api_std_vocab_bulk(raw: bytes, ctype: str) -> dict:
    """업로드된 엑셀에서 '표준어' 열을 읽어 검색 어휘 목록을 통째로 교체한다."""
    mp = _parse_multipart(raw, ctype)
    files = mp.get("files") or []
    if not files:
        return {"ok": False, "message": "엑셀 파일이 첨부되지 않았습니다."}

    up = files[0]
    if not str(up.get("filename", "")).lower().endswith((".xlsx", ".xlsm")):
        return {"ok": False, "message": "xlsx 형식의 파일만 등록할 수 있습니다."}

    try:
        sheets = _read_xlsx(up["data"])
    except Exception as e:
        return {"ok": False, "message": f"엑셀을 읽을 수 없습니다: {e}"}

    # '표준어' 열을 가진 첫 시트를 대상으로 삼는다.
    target = None
    skipped_sheets = []
    for name, rows in sheets:
        hit = None
        for idx, row in enumerate(rows[:10]):
            if _find_col(row, "표준어") >= 0:
                hit = idx
                break
        if hit is not None and target is None:
            target = (name, rows, hit)
        elif any(any(c for c in r) for r in rows):
            skipped_sheets.append(name)

    if target is None:
        return {"ok": False, "message": "'표준어' 열이 있는 시트를 찾지 못했습니다."}

    name, rows, hrow = target
    header = rows[hrow]
    c_word = _find_col(header, "표준어")
    c_no = _find_col(header, "항목번호", "항목No")

    today = datetime.now().strftime("%Y-%m-%d")
    items, seen = [], {}
    blank = dup = 0
    for row in rows[hrow + 1:]:
        word = row[c_word].strip() if c_word < len(row) else ""
        if not word:
            blank += 1
            continue
        if word in seen:
            dup += 1
            continue
        seen[word] = True
        item_no = row[c_no].strip() if 0 <= c_no < len(row) else ""
        items.append({
            "id": len(items) + 1,
            "itemNo": item_no,
            "word": word,
            "useYn": "Y",
            "sortOrdr": len(items) + 1,
            "rmrk": "",
            "regDt": today,
            "updDt": today,
        })

    if not items:
        return {"ok": False, "message": "등록할 어휘가 없습니다."}

    # 기존 목록은 .bak 한 부만 남긴다.
    if STD_VOCAB_JSON.is_file():
        STD_VOCAB_JSON.with_suffix(".json.bak").write_bytes(STD_VOCAB_JSON.read_bytes())

    STD_VOCAB_JSON.parent.mkdir(parents=True, exist_ok=True)
    STD_VOCAB_JSON.write_text(json.dumps({
        "source": up["filename"],
        "updated": today,
        "items": items,
    }, ensure_ascii=False, indent=1), encoding="utf-8")

    return {
        "ok": True,
        "count": len(items),
        "sheet": name,
        "blank": blank,
        "dup": dup,
        "skippedSheets": skipped_sheets,
        "file": up["filename"],
    }


def _parse_multipart(raw: bytes, ctype: str) -> dict:
    """multipart/form-data 최소 파서 → {files:[{field,filename,data}], fields:{}}."""
    out = {"files": [], "fields": {}}
    if "multipart/form-data" not in (ctype or "") or "boundary=" not in ctype:
        return out
    boundary = ctype.split("boundary=", 1)[1].strip().strip('"')
    sep = b"--" + boundary.encode()
    for part in raw.split(sep):
        part = part.strip(b"\r\n")
        if not part or part == b"--":
            continue
        if b"\r\n\r\n" not in part:
            continue
        head, data = part.split(b"\r\n\r\n", 1)
        data = data.rstrip(b"\r\n")
        head_txt = head.decode("utf-8", "replace")
        nm = re.search(r'name="([^"]*)"', head_txt)
        fm = re.search(r'filename="([^"]*)"', head_txt)
        name = nm.group(1) if nm else ""
        if fm and fm.group(1):
            out["files"].append({
                "field": name,
                "filename": unicodedata.normalize("NFC", fm.group(1)),
                "data": data,
            })
        else:
            out["fields"][name] = data.decode("utf-8", "replace")
    return out


def _wav_duration_sec(path: Path) -> float:
    try:
        with wave.open(str(path), "rb") as w:
            fr = w.getframerate() or 0
            return (w.getnframes() / fr) if fr else 0.0
    except (wave.Error, OSError, EOFError):
        return 0.0


def api_oral_upload(raw: bytes, ctype: str) -> dict:
    """등록 ① 업로드+파싱 — .eaf/.trs/.wav 저장 후 파일별 정보 반환."""
    mp = _parse_multipart(raw, ctype)
    if not mp["files"]:
        return {"ok": False, "message": "업로드된 파일이 없습니다."}
    ORAL_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    results = []
    for f in mp["files"]:
        fname = re.sub(r"[^\w.\-가-힣]", "_", f["filename"])
        if not fname:
            continue
        dest = ORAL_UPLOAD_DIR / fname
        dest.write_bytes(f["data"])
        ext = fname.rsplit(".", 1)[-1].lower() if "." in fname else ""
        base = _oral_id_of(dest)
        if ext in ("eaf", "trs"):
            try:
                rec = parse_oral_transcript(dest)
            except Exception as e:
                label = "TRS" if ext == "trs" else "EAF"
                results.append({"ok": False, "fileName": fname, "message": f"{label} 파싱 실패: {e}"})
                continue
            results.append({
                "ok": True, "kind": ext, "fileName": fname, "baseId": base,
                "durationSec": round(rec["durationMs"] / 1000, 3),
                "durationHms": _ms_to_hms(rec["durationMs"]),
                "segmentCount": rec["segmentCount"],
                "version": "1",
                "meta": {
                    "region": rec["region"], "sidoCd": rec["sidoCd"],
                    "year": rec["year"], "contentCode": rec["contentCode"],
                    "topic": rec["topic"], "informant": rec["informant"],
                    "sex": rec["sex"], "birth": rec["birth"],
                    "investigator": rec.get("investigator", ""),
                    "transcriber": rec.get("transcriber", ""),
                    # 제보자가 2명 이상인지는 이 목록으로만 알 수 있다
                    "speakers": rec.get("speakers", []),
                },
            })
        elif ext == "wav":
            sec = _wav_duration_sec(dest)
            results.append({
                "ok": True, "kind": "wav", "fileName": fname, "baseId": base,
                "waveSec": round(sec, 3), "waveTimeHms": _sec_to_hms(sec),
            })
        else:
            results.append({"ok": False, "fileName": fname, "message": "지원하지 않는 형식(.eaf/.trs/.wav)"})
    return {"ok": True, "results": results}


def _next_id(con, table: str, col: str) -> int:
    row = con.execute(f"SELECT MAX(CAST({col} AS INTEGER)) FROM {table}").fetchone()
    return (row[0] or 0) + 1


def api_oral_create(body: dict) -> dict:
    """등록 ② insert — 업로드된 eaf별 wb_trs_file_talk + wb_trs_line_talk,
    조사지역 wb_research_region, 제보자 wb_source 생성."""
    files = body.get("files") or []
    if not files:
        return {"ok": False, "message": "등록할 전사자료(.eaf/.trs)가 없습니다."}
    region = body.get("region") or {}
    headword = (body.get("headword") or "").strip()
    use_yn = "N" if str(body.get("useYn") or "Y").upper() == "N" else "Y"
    year = (region.get("researchYear") or "").strip()
    degree = "2" if (year.isdigit() and int(year) >= 2022) else "1"
    sex_code = {"남": "1", "여": "0", "man": "1", "woman": "0", "wom": "0"}  # wb_source: 0=여,1=남

    def src_rows(lst, base_se):
        out = []
        for i, s in enumerate(lst or []):
            nm = (s.get("name") or "").strip()
            if not nm:
                continue
            out.append({
                "se": str(base_se + i),
                "sourceId": str(s.get("sourceId") or "").strip(),
                "name": nm,
                "sex": sex_code.get((s.get("sex") or "").strip(), (s.get("sex") or "").strip()),
                "age": (s.get("age") or "").strip(),
                "birth": (s.get("birth") or "").strip(),
                "residence": (s.get("residence") or "").strip(),
                "birth_place": (s.get("birthPlace") or "").strip(),
                "job": (s.get("job") or "").strip(),
                "education": (s.get("education") or "").strip(),
            })
        return out

    created = []
    with db_connect() as con:
        rrid = _next_id(con, "wb_research_region", "research_region_id")
        con.execute(
            """INSERT INTO wb_research_region
               (research_region_id, region_nm, sigungu_nm, legal_region_code, region_nm_yn,
                region_remark, research_place, first_place, researcher, transcriber,
                mic, recorder, file_uniqueness, research_year, start_date, end_date)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                str(rrid),
                (region.get("regionNm") or "").strip(),
                (region.get("sigunguNm") or "").strip(),
                (region.get("legalRegionCode") or "").strip(),
                (region.get("regionNmYn") or "").strip(),
                (region.get("regionRemark") or "").strip(),
                (region.get("researchPlace") or "").strip(),
                (region.get("firstPlace") or "").strip(),
                (region.get("researcher") or "").strip(),
                (region.get("transcriber") or "").strip(),
                (region.get("mic") or "").strip(),
                (region.get("recorder") or "").strip(),
                (region.get("fileUniqueness") or "").strip(),
                year,
                _date_to_epoch_ms(region.get("startDate")),
                _date_to_epoch_ms(region.get("endDate")),
            ),
        )
        # 제보자 (주=se 0.., 부=se 1..). sourceId 있으면(제보자 관리에서 등록된 것) 링크, 없으면 신규 INSERT.
        sid = _next_id(con, "wb_source", "source_id")
        for s in src_rows(body.get("mainSources"), 0) + src_rows(body.get("subSources"), 1):
            if s["sourceId"]:
                con.execute(
                    "UPDATE wb_source SET research_region_id=?, se=? WHERE source_id=?",
                    (str(rrid), s["se"], s["sourceId"]),
                )
                continue
            con.execute(
                """INSERT INTO wb_source
                   (source_id, research_region_id, se, name, sex, age, birth,
                    residence, birth_place, job, education)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (str(sid), str(rrid), s["se"], s["name"], s["sex"], s["age"], s["birth"],
                 s["residence"], s["birth_place"], s["job"], s["education"]),
            )
            sid += 1

        next_trs = _next_id(con, "wb_trs_file_talk", "trs_id")
        next_line = _next_id(con, "wb_trs_line_talk", "trs_line_id")
        for fname in files:
            safe = re.sub(r"[^\w.\-가-힣]", "_", str(fname))
            p = ORAL_UPLOAD_DIR / safe
            if not p.is_file():
                return {"ok": False, "message": f"업로드 파일을 찾을 수 없습니다: {safe}"}
            rec = parse_oral_transcript(p)
            base = _oral_id_of(p)
            stored_nm = (base + ".eaf") if p.suffix.lower() == ".eaf" else (base + p.suffix.lower())
            wav = base + ".wav"
            wav_sec = _wav_duration_sec(ORAL_UPLOAD_DIR / wav) if (ORAL_UPLOAD_DIR / wav).is_file() else 0.0
            trs_id = str(next_trs)
            next_trs += 1
            con.execute(
                """INSERT INTO wb_trs_file_talk
                   (trs_id, research_region_id, upper_headword, trs_file_nm, trs_time,
                    wave_file_nm, wave_time, research_degree, audio_filename, ver, use_yn)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    trs_id, str(rrid), headword, stored_nm,
                    str(round(rec["durationMs"] / 1000, 3)),
                    wav, str(round(wav_sec, 3)) if wav_sec else "",
                    degree, base, "1", use_yn,
                ),
            )
            for i, sg in enumerate(rec["segments"], 1):
                line = _compose_trs_line(sg["speaker"], sg["form"], sg["std"])
                con.execute(
                    """INSERT INTO wb_trs_line_talk
                       (trs_line_id, trs_id, trs_line_no, trs_line_se, trs_line_text_no,
                        trs_line, start_time, end_time)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    (
                        str(next_line), trs_id, str(i), "text", str(i), line,
                        str(round(sg["startMs"] / 1000, 3)), str(round(sg["endMs"] / 1000, 3)),
                    ),
                )
                next_line += 1
            created.append({"trsId": trs_id, "fileName": stored_nm, "segments": rec["segmentCount"]})
        con.commit()

    return {"ok": True, "researchRegionId": str(rrid), "created": created,
            "message": f"{len(created)}개 전사자료가 등록되었습니다."}


def _speaker_to_marker(speaker: str) -> str:
    s = (speaker or "").strip()
    if s == "조사자":
        return "@"
    m = re.match(r"제보자(\d*)", s)
    if m:
        return "#" + (m.group(1) or "1")
    return "@"


def _compose_trs_line(speaker: str, form: str, std: str) -> str:
    """화자+형태음소전사+표준어대역 → .trs 라인 재구성."""
    mark = _speaker_to_marker(speaker)
    form = (form or "").strip()
    std = (std or "").strip()
    line = mark + " %1" + form
    if std:
        line += " %2 {" + std + "}"
    return line


def api_oral_save_lines(body: dict) -> dict:
    """전사 라인 저장 — wb_trs_line_talk.trs_line UPDATE (화자+형태음소+표준어대역 재구성)."""
    trs_id = str(body.get("trsId") or "").strip()
    lines = body.get("lines") or []
    if not trs_id:
        return {"ok": False, "message": "trsId가 필요합니다."}
    if not isinstance(lines, list) or not lines:
        return {"ok": False, "message": "저장할 라인이 없습니다."}

    updated = 0
    skipped = 0
    with db_connect() as con:
        exists = con.execute(
            "SELECT 1 FROM wb_trs_file_talk WHERE trs_id = ?", (trs_id,)
        ).fetchone()
        if not exists:
            return {"ok": False, "message": f"해당 자료를 찾을 수 없습니다: {trs_id}"}
        for ln in lines:
            line_id = str((ln or {}).get("trsLineId") or "").strip()
            if not line_id:
                continue
            cur_row = con.execute(
                "SELECT trs_line FROM wb_trs_line_talk WHERE trs_line_id = ? AND trs_id = ?",
                (line_id, trs_id),
            ).fetchone()
            if not cur_row:
                continue
            # 안전장치: 단일 턴(%1 정확히 1개)만 구조화 저장 허용.
            # 헤더(0개)·복합 멀티턴(2개 이상)은 재구성 시 손실되므로 제외.
            if (cur_row["trs_line"] or "").count("%1") != 1:
                skipped += 1
                continue
            new_text = _compose_trs_line(
                ln.get("speaker", ""), ln.get("form", ""), ln.get("std", "")
            )
            cur = con.execute(
                "UPDATE wb_trs_line_talk SET trs_line = ? WHERE trs_line_id = ? AND trs_id = ?",
                (new_text, line_id, trs_id),
            )
            updated += cur.rowcount
        con.commit()
    msg = f"{updated}개 라인 저장됨"
    if skipped:
        msg += f" (복합 라인 {skipped}개 제외)"
    return {"ok": True, "trsId": trs_id, "updated": updated, "skipped": skipped, "message": msg}


def api_source_list(qs: dict) -> dict:
    """제보자 관리 목록 — wb_source."""
    def q1(k, d=""):
        return (qs.get(k) or [d])[0].strip()
    search = q1("searchText") or q1("q")
    try:
        page = max(1, int(q1("page", "1")))
    except ValueError:
        page = 1
    try:
        page_size = int(q1("pageSize", "10"))
    except ValueError:
        page_size = 10
    if page_size not in (10, 50, 100, 200, 300):
        page_size = 10
    region = q1("region")      # 조사 지역 접두 매칭 (예: "충청북도" 또는 "충청북도 청주시")
    age_group = q1("ageGroup")  # 연령대(10년 단위) 숫자 (예: "2" → 20대)
    where, params = [], []
    if search:
        where.append("(name LIKE ? OR IFNULL(region_nm,'') LIKE ?)")
        params += ["%" + search + "%", "%" + search + "%"]
    if region and region != "*":
        where.append("IFNULL(region_nm,'') LIKE ?")
        params.append(region + "%")
    if age_group and age_group not in ("*", "0"):
        try:
            d = int(age_group)
            where.append("CAST(NULLIF(TRIM(age),'') AS INTEGER) BETWEEN ? AND ?")
            params += [d * 10, d * 10 + 9]
        except ValueError:
            pass
    wh = ("WHERE " + " AND ".join(where)) if where else ""
    sex_lbl = {"0": "여", "1": "남"}
    with db_connect() as con:
        total = con.execute(f"SELECT COUNT(*) FROM wb_source {wh}", params).fetchone()[0]
        offset = (page - 1) * page_size
        rows = con.execute(
            f"""SELECT source_id, IFNULL(region_nm,'') region_nm, name, IFNULL(sex,'') sex,
                       IFNULL(birth,'') birth, IFNULL(age,'') age
                FROM wb_source {wh}
                ORDER BY CAST(source_id AS INTEGER) DESC
                LIMIT ? OFFSET ?""",
            params + [page_size, offset],
        ).fetchall()
    items = [
        {
            "sourceId": str(r["source_id"]),
            "regionNm": r["region_nm"],
            "name": r["name"] or "",
            "sex": sex_lbl.get(r["sex"], r["sex"]),
            "birth": r["birth"],
            "age": r["age"],
        }
        for r in rows
    ]
    total_pages = max(1, (total + page_size - 1) // page_size)
    return {"ok": True, "total": total, "page": page, "pageSize": page_size,
            "totalPages": total_pages, "list": items}


def api_source_regions(qs: dict) -> dict:
    """제보자 검색용 조사 지역 캐스케이드(시/도 → 시군구).
    wb_source.region_nm 에 실제 존재하는 값을 토큰 분해해 계층 구성."""
    with db_connect() as con:
        rows = con.execute(
            "SELECT DISTINCT region_nm FROM wb_source WHERE IFNULL(region_nm,'') != ''"
        ).fetchall()
    sido_set = {}          # sido -> True (순서 보존용 dict)
    sigungu = {}           # sido -> { "sido sigungu": sigungu_label }
    for r in rows:
        toks = str(r["region_nm"]).split()
        if not toks:
            continue
        sido = toks[0]
        sido_set.setdefault(sido, True)
        if len(toks) >= 2:
            val = sido + " " + toks[1]
            sigungu.setdefault(sido, {})[val] = toks[1]
    sido_list = sorted(sido_set.keys())
    sigungu_out = {
        sido: [{"value": v, "label": lbl} for v, lbl in sorted(pairs.items())]
        for sido, pairs in sigungu.items()
    }
    return {"ok": True, "sido": sido_list, "sigungu": sigungu_out}


def api_source_detail(qs: dict) -> dict:
    sid = (qs.get("id") or qs.get("sourceId") or [""])[0].strip()
    if not sid:
        return {"ok": False, "message": "source_id 가 필요합니다."}
    with db_connect() as con:
        row = con.execute("SELECT * FROM wb_source WHERE source_id=? LIMIT 1", (sid,)).fetchone()
    if not row:
        return {"ok": False, "message": "제보자를 찾을 수 없습니다."}
    d = dict(row)
    # 연동된 전사(.eaf) 파일 목록 — 제보자가 속한 조사지역(research_region)의 전사파일.
    # 로컬 DB는 같은 지역(동일 region_nm/법정동코드)에 대해 조사지역(research_region) 레코드가
    # 여러 개로 중복되어 있고, 제보자가 연결된 research_region_id 와 전사파일이 등록된
    # research_region_id 가 서로 다른 중복 레코드인 경우가 많다. 그래서 단일 id 로만 조인하면
    # 실제로는 그 지역에 전사파일이 있어도 0건이 된다.
    # → 제보자의 지역 정체성(region_nm / 법정동코드)을 먼저 확정한 뒤, 그 지역명·코드를
    #   공유하는 '모든' 중복 조사지역의 전사파일을 합쳐서 보여준다.
    trs_files = []
    with db_connect() as con:
        rrid_direct = str(d.get("research_region_id") or "").strip()
        region_nm = (d.get("region_nm") or "").strip()
        legal_code = (d.get("legal_region_code") or "").strip()
        # 제보자가 조사지역에 직접 연결돼 있으면 그 지역의 지역명/코드를 채운다.
        if rrid_direct:
            rr = con.execute(
                "SELECT region_nm, legal_region_code FROM wb_research_region WHERE research_region_id=?",
                (rrid_direct,),
            ).fetchone()
            if rr:
                region_nm = region_nm or (rr["region_nm"] or "")
                legal_code = legal_code or (rr["legal_region_code"] or "")
        # 지역명이 없으면 거주지 문자열로 최후 매칭 시도.
        if not region_nm:
            region_nm = (d.get("residence") or "").strip()
        # 지역명/코드를 공유하는 모든 조사지역 + 제보자가 직접 연결된 조사지역을 대상으로 합집합.
        rrids = set()
        if rrid_direct:
            rrids.add(rrid_direct)
        if region_nm or legal_code:
            for r in con.execute(
                """SELECT research_region_id FROM wb_research_region
                   WHERE (? != '' AND region_nm = ?) OR (? != '' AND legal_region_code = ?)""",
                (region_nm, region_nm, legal_code, legal_code),
            ).fetchall():
                rrids.add(str(r["research_region_id"]))
        if rrids:
            ph = ",".join("?" * len(rrids))
            for r in con.execute(
                f"""SELECT DISTINCT trs_file_nm FROM wb_trs_file_talk
                    WHERE research_region_id IN ({ph}) AND IFNULL(trs_file_nm,'') != ''
                    ORDER BY trs_file_nm""",
                tuple(rrids),
            ).fetchall():
                trs_files.append(r["trs_file_nm"])
    return {"ok": True, "data": {
        "sourceId": d.get("source_id") or "", "name": d.get("name") or "",
        "sex": d.get("sex") or "1", "age": d.get("age") or "", "birth": d.get("birth") or "",
        "birthPlace": d.get("birth_place") or "", "residence": d.get("residence") or "",
        "parentResidence": d.get("parent_residence") or "", "job": d.get("job") or "",
        "education": d.get("education") or "", "introductionDetail": d.get("introduction_detail") or "",
        "consentForm": d.get("consent_form") or "0", "voiceRemark": d.get("voice_remark") or "",
        "remark": d.get("remark") or "", "sourcePlace": d.get("source_place") or "",
        "regionNm": d.get("region_nm") or "",
        "legalRegionCode": d.get("legal_region_code") or "",
        "researchRegionId": str(d.get("research_region_id") or ""),
        "trsFiles": trs_files,
    }}


_SOURCE_FIELDS = [
    ("name", "name"), ("sex", "sex"), ("age", "age"), ("birth", "birth"),
    ("birth_place", "birthPlace"), ("residence", "residence"),
    ("parent_residence", "parentResidence"), ("job", "job"), ("education", "education"),
    ("introduction_detail", "introductionDetail"), ("consent_form", "consentForm"),
    ("voice_remark", "voiceRemark"), ("remark", "remark"), ("source_place", "sourcePlace"),
    ("region_nm", "regionNm"),
    # 거주지 지역찾기로 선택한 법정동 코드 (대표 지역코드로 저장)
    ("legal_region_code", "residenceCode"),
]


def api_source_save(body: dict) -> dict:
    """제보자 등록(C)/수정(M) — wb_source."""
    mode = (body.get("mode") or "C").strip().upper()
    name = (body.get("name") or "").strip()
    if not name:
        return {"ok": False, "message": "제보자 이름을 입력하세요."}
    vals = {col: str(body.get(key) or "").strip() for col, key in _SOURCE_FIELDS}
    with db_connect() as con:
        if mode == "M":
            sid = str(body.get("sourceId") or body.get("id") or "").strip()
            if not sid:
                return {"ok": False, "message": "수정 대상 source_id 가 없습니다."}
            sets = ", ".join(c + "=?" for c, _ in _SOURCE_FIELDS)
            cur = con.execute(
                f"UPDATE wb_source SET {sets} WHERE source_id=?",
                [vals[c] for c, _ in _SOURCE_FIELDS] + [sid],
            )
            con.commit()
            if cur.rowcount == 0:
                return {"ok": False, "message": "수정할 제보자를 찾지 못했습니다."}
            return {"ok": True, "message": "저장되었습니다.", "mode": "M", "sourceId": sid}
        sid = _next_id(con, "wb_source", "source_id")
        cols = ["source_id", "research_region_id", "se"] + [c for c, _ in _SOURCE_FIELDS]
        ph = ",".join(["?"] * len(cols))
        con.execute(
            f"INSERT INTO wb_source ({','.join(cols)}) VALUES ({ph})",
            [str(sid), str(body.get("researchRegionId") or "").strip(), "0"]
            + [vals[c] for c, _ in _SOURCE_FIELDS],
        )
        con.commit()
        return {"ok": True, "message": "제보자가 등록되었습니다.", "mode": "C", "sourceId": str(sid)}


def api_source_delete(body: dict) -> dict:
    """제보자 삭제 — wb_source."""
    sid = str(body.get("sourceId") or body.get("id") or "").strip()
    if not sid:
        return {"ok": False, "message": "삭제 대상 source_id 가 없습니다."}
    with db_connect() as con:
        cur = con.execute("DELETE FROM wb_source WHERE source_id=?", (sid,))
        con.commit()
        if cur.rowcount == 0:
            return {"ok": False, "message": "삭제할 제보자를 찾지 못했습니다."}
    return {"ok": True, "message": "삭제되었습니다.", "sourceId": sid}


# ─────────────────────────────────────────────────────────────────────────
# 설문조사 관리 — tb_survey_new(+ tb_survey_question_new / tb_survey_example_new /
# tb_survey_answer_new). 기존 DB 구조 유지. 날짜는 epoch millisecond(문자열) 저장.
# ─────────────────────────────────────────────────────────────────────────
def _epoch_ms_to_date(ms) -> str:
    """epoch millisecond → 'YYYY-MM-DD' (로컬 시간)."""
    s = str(ms or "").strip()
    if not s:
        return ""
    try:
        return datetime.fromtimestamp(int(s) / 1000).strftime("%Y-%m-%d")
    except (ValueError, OverflowError, OSError):
        return ""


def _date_to_epoch_ms(s, end=False) -> str:
    """'YYYY-MM-DD' → epoch millisecond 문자열. end=True 면 그날 23:59:59."""
    s = (s or "").strip()
    if not s:
        return ""
    try:
        dt = datetime.strptime(s[:10], "%Y-%m-%d")
        if end:
            dt = dt.replace(hour=23, minute=59, second=59)
        return str(int(dt.timestamp() * 1000))
    except ValueError:
        return ""


def _survey_status(start_ms, end_ms) -> str:
    """설문 상태 — 예정 / 진행중 / 종료 (현재 시각 기준)."""
    now = datetime.now().timestamp() * 1000
    try:
        s = int(str(start_ms).strip()) if str(start_ms).strip() else None
        e = int(str(end_ms).strip()) if str(end_ms).strip() else None
    except ValueError:
        return ""
    if s is not None and now < s:
        return "예정"
    if e is not None and now > e:
        return "종료"
    return "진행중"


def _survey_response_count(con, survey_no) -> int:
    """설문 응답자 수 — 문항에 달린 응답의 distinct answer_serial."""
    row = con.execute(
        """SELECT COUNT(DISTINCT a.answer_serial)
           FROM tb_survey_answer_new a
           JOIN tb_survey_question_new q ON a.question_no = q.question_no
           WHERE q.survey_no = ?""",
        (str(survey_no),),
    ).fetchone()
    return row[0] or 0


def api_survey_list(qs: dict) -> dict:
    """설문조사 목록 — tb_survey_new."""
    def q1(k, d=""):
        return (qs.get(k) or [d])[0].strip()
    search = q1("searchText") or q1("q")
    try:
        page = max(1, int(q1("page", "1")))
    except ValueError:
        page = 1
    try:
        page_size = int(q1("pageSize", "10"))
    except ValueError:
        page_size = 10
    if page_size not in (10, 20, 50, 100):
        page_size = 10
    status_filter = q1("status")  # 예정 / 진행중 / 종료 (빈값=전체)
    where, params = [], []
    if search:
        where.append("IFNULL(survey_title,'') LIKE ?")
        params.append("%" + search + "%")
    wh = ("WHERE " + " AND ".join(where)) if where else ""
    # 상태는 기간으로 계산되므로 제목 필터로 전량 조회 후 상태 필터 + 파이썬 페이징.
    all_items = []
    with db_connect() as con:
        rows = con.execute(
            f"""SELECT survey_no, survey_title, start_date, end_date, question_cnt
                FROM tb_survey_new {wh}
                ORDER BY CAST(survey_no AS INTEGER) DESC""",
            params,
        ).fetchall()
        for r in rows:
            st = _survey_status(r["start_date"], r["end_date"])
            if status_filter and st != status_filter:
                continue
            all_items.append({
                "surveyNo": str(r["survey_no"]),
                "surveyTitle": r["survey_title"] or "",
                "startDate": _epoch_ms_to_date(r["start_date"]),
                "endDate": _epoch_ms_to_date(r["end_date"]),
                "questionCnt": r["question_cnt"] or "0",
                "responseCount": _survey_response_count(con, r["survey_no"]),
                "status": st,
            })
    total = len(all_items)
    total_pages = max(1, (total + page_size - 1) // page_size)
    if page > total_pages:
        page = total_pages
    offset = (page - 1) * page_size
    items = all_items[offset:offset + page_size]
    return {"ok": True, "total": total, "page": page, "pageSize": page_size,
            "totalPages": total_pages, "list": items}


def api_survey_detail(qs: dict) -> dict:
    """설문조사 상세 — tb_survey_new + 문항/보기(읽기)."""
    sid = (qs.get("id") or qs.get("surveyNo") or [""])[0].strip()
    if not sid:
        return {"ok": False, "message": "survey_no 가 필요합니다."}
    with db_connect() as con:
        row = con.execute("SELECT * FROM tb_survey_new WHERE survey_no=? LIMIT 1", (sid,)).fetchone()
        if not row:
            return {"ok": False, "message": "설문을 찾을 수 없습니다."}
        d = dict(row)
        questions = []
        for q in con.execute(
            """SELECT question_no, question_title, question_order
               FROM tb_survey_question_new WHERE survey_no=?
               ORDER BY CAST(question_order AS INTEGER), CAST(question_no AS INTEGER)""",
            (sid,),
        ).fetchall():
            examples = []
            for e in con.execute(
                """SELECT example_no, example_title FROM tb_survey_example_new
                   WHERE question_no=? ORDER BY CAST(example_no AS INTEGER)""",
                (q["question_no"],),
            ).fetchall():
                cnt = con.execute(
                    "SELECT COUNT(*) FROM tb_survey_answer_new WHERE example_no=? AND question_no=?",
                    (e["example_no"], q["question_no"]),
                ).fetchone()[0]
                examples.append({
                    "exampleNo": str(e["example_no"]),
                    "exampleTitle": e["example_title"] or "",
                    "count": cnt,
                })
            questions.append({
                "questionNo": str(q["question_no"]),
                "questionTitle": q["question_title"] or "",
                "examples": examples,
            })
        resp = _survey_response_count(con, sid)
    return {"ok": True, "data": {
        "surveyNo": str(d.get("survey_no") or ""),
        "surveyTitle": d.get("survey_title") or "",
        "surveyCntnts": d.get("survey_cntnts") or "",
        "startDate": _epoch_ms_to_date(d.get("start_date")),
        "endDate": _epoch_ms_to_date(d.get("end_date")),
        "prsnlInputYn": d.get("prsnl_input_yn") or "N",
        "prsnlInfoCntnts": d.get("prsnl_info_cntnts") or "",
        "questionCnt": d.get("question_cnt") or "0",
        "status": _survey_status(d.get("start_date"), d.get("end_date")),
        "responseCount": resp,
        "questions": questions,
    }}


def _epoch_ms_to_datetime(ms) -> str:
    """epoch millisecond → 'YYYY.MM.DD HH:MM:SS'."""
    s = str(ms or "").strip()
    if not s:
        return ""
    try:
        return datetime.fromtimestamp(int(s) / 1000).strftime("%Y.%m.%d %H:%M:%S")
    except (ValueError, OverflowError, OSError):
        return ""


def api_survey_respondents(qs: dict) -> dict:
    """응답자별 보기 — 설문 응답자(answer_serial) 목록 + 최종 응답 일시."""
    sid = (qs.get("id") or qs.get("surveyNo") or [""])[0].strip()
    if not sid:
        return {"ok": False, "message": "survey_no 가 필요합니다."}
    items = []
    with db_connect() as con:
        srv = con.execute(
            "SELECT survey_title, survey_cntnts, start_date, end_date, prsnl_input_yn FROM tb_survey_new WHERE survey_no=? LIMIT 1",
            (sid,)).fetchone()
        if not srv:
            return {"ok": False, "message": "설문을 찾을 수 없습니다."}
        rows = con.execute(
            """SELECT a.answer_serial, MAX(CAST(a.answer_dt AS INTEGER)) mdt
               FROM tb_survey_answer_new a
               JOIN tb_survey_question_new q ON a.question_no = q.question_no
               WHERE q.survey_no = ?
               GROUP BY a.answer_serial
               ORDER BY mdt DESC""",
            (sid,)).fetchall()
        for i, r in enumerate(rows, start=1):
            items.append({
                "no": i,
                "answerSerial": r["answer_serial"] or "",
                "answerDate": _epoch_ms_to_datetime(r["mdt"]),
            })
    return {"ok": True, "total": len(items), "list": items, "surveyNo": sid,
            "surveyTitle": srv["survey_title"] or "",
            "surveyCntnts": srv["survey_cntnts"] or "",
            "startDate": _epoch_ms_to_date(srv["start_date"]),
            "endDate": _epoch_ms_to_date(srv["end_date"]),
            "status": _survey_status(srv["start_date"], srv["end_date"])}


def api_survey_answer_detail(qs: dict) -> dict:
    """답변 보기 팝업 — 특정 응답자(answer_serial)의 문항별 선택 답변."""
    def q1(k):
        return (qs.get(k) or [""])[0].strip()
    sid = q1("id") or q1("surveyNo")
    serial = q1("serial") or q1("answerSerial")
    if not sid or not serial:
        return {"ok": False, "message": "survey_no 와 answer_serial 이 필요합니다."}
    answers = []
    with db_connect() as con:
        questions = con.execute(
            """SELECT question_no, question_title, question_order
               FROM tb_survey_question_new WHERE survey_no=?
               ORDER BY CAST(question_order AS INTEGER), CAST(question_no AS INTEGER)""",
            (sid,)).fetchall()
        picked = {}
        for a in con.execute(
            """SELECT a.question_no, e.example_title
               FROM tb_survey_answer_new a
               LEFT JOIN tb_survey_example_new e ON a.example_no = e.example_no
               WHERE a.answer_serial = ?""",
            (serial,)).fetchall():
            picked[str(a["question_no"])] = a["example_title"] or ""
        for i, q in enumerate(questions, start=1):
            answers.append({
                "no": i,
                "questionTitle": q["question_title"] or "",
                "answerLabel": picked.get(str(q["question_no"]), ""),
            })
    return {"ok": True, "answerSerial": serial, "answers": answers}


def build_survey_excel(sid: str):
    """설문 응답 엑셀(.xlsx) 생성 — (bytes, 파일명). 두 시트: 질문별 응답 / 응답자별 응답."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    sid = str(sid).strip()
    with db_connect() as con:
        srv = con.execute(
            "SELECT survey_title FROM tb_survey_new WHERE survey_no=? LIMIT 1", (sid,)).fetchone()
        if not srv:
            return None, None
        title = srv["survey_title"] or ("survey_" + sid)
        questions = con.execute(
            """SELECT question_no, question_title FROM tb_survey_question_new
               WHERE survey_no=? ORDER BY CAST(question_order AS INTEGER), CAST(question_no AS INTEGER)""",
            (sid,)).fetchall()
        qlist = [{"no": str(q["question_no"]), "title": q["question_title"] or ""} for q in questions]
        # 시트1용 보기/응답수
        ex_by_q = {}
        for q in qlist:
            exs = con.execute(
                """SELECT example_no, example_title FROM tb_survey_example_new
                   WHERE question_no=? ORDER BY CAST(example_no AS INTEGER)""", (q["no"],)).fetchall()
            counts = {str(r["example_no"]): r["c"] for r in con.execute(
                "SELECT example_no, COUNT(*) c FROM tb_survey_answer_new WHERE question_no=? GROUP BY example_no",
                (q["no"],)).fetchall()}
            ex_by_q[q["no"]] = [{"title": e["example_title"] or "",
                                 "count": counts.get(str(e["example_no"]), 0)} for e in exs]
        # 시트2용 응답자 + 답변
        respondents = con.execute(
            """SELECT a.answer_serial, MAX(CAST(a.answer_dt AS INTEGER)) mdt
               FROM tb_survey_answer_new a JOIN tb_survey_question_new q ON a.question_no=q.question_no
               WHERE q.survey_no=? GROUP BY a.answer_serial ORDER BY mdt DESC""", (sid,)).fetchall()
        picked = {}  # serial -> {qno: label}
        for a in con.execute(
            """SELECT a.answer_serial, a.question_no, e.example_title
               FROM tb_survey_answer_new a
               JOIN tb_survey_question_new q ON a.question_no=q.question_no
               LEFT JOIN tb_survey_example_new e ON a.example_no=e.example_no
               WHERE q.survey_no=?""", (sid,)).fetchall():
            picked.setdefault(a["answer_serial"], {})[str(a["question_no"])] = a["example_title"] or ""

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2563EB")
    center = Alignment(horizontal="center", vertical="center")

    def style_header(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = center

    # 시트1: 질문별 응답
    ws1 = wb.active
    ws1.title = "질문별 응답"
    ws1.append(["문항번호", "문항", "보기", "응답수", "비율(%)"])
    for i, q in enumerate(qlist, start=1):
        exs = ex_by_q[q["no"]]
        total = sum(e["count"] for e in exs) or 0
        if not exs:
            ws1.append([i, q["title"], "", 0, 0])
        for e in exs:
            pct = round(e["count"] / total * 100) if total else 0
            ws1.append([i, q["title"], e["title"], e["count"], pct])
    style_header(ws1, 5)
    for col, w in zip("ABCDE", (10, 46, 22, 10, 10)):
        ws1.column_dimensions[col].width = w

    # 시트2: 응답자별 응답
    ws2 = wb.create_sheet("응답자별 응답")
    header = ["번호", "응답자 일련번호", "응답 일시"] + [f"Q{i}. {q['title']}" for i, q in enumerate(qlist, start=1)]
    ws2.append(header)
    for idx, r in enumerate(respondents, start=1):
        serial = r["answer_serial"]
        row = [idx, serial, _epoch_ms_to_datetime(r["mdt"])]
        ans = picked.get(serial, {})
        row += [ans.get(q["no"], "") for q in qlist]
        ws2.append(row)
    style_header(ws2, len(header))
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 20
    for c in range(4, len(header) + 1):
        ws2.column_dimensions[ws2.cell(row=1, column=c).column_letter].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    safe_title = re.sub(r'[\\/:*?"<>|]', "_", title).strip() or ("survey_" + sid)
    return buf.getvalue(), safe_title + "_설문응답.xlsx"


def _save_survey_examples(con, question_no, examples):
    """문항의 보기 upsert — 기존 example_no 는 보존(응답 연결 유지), 제거분만 삭제."""
    question_no = str(question_no)
    existing = {str(r["example_no"]) for r in con.execute(
        "SELECT example_no FROM tb_survey_example_new WHERE question_no=?", (question_no,)).fetchall()}
    seen = set()
    for e in examples:
        eno = str(e.get("exampleNo") or "").strip()
        etitle = (e.get("exampleTitle") or "").strip()
        if eno and eno in existing:
            con.execute("UPDATE tb_survey_example_new SET example_title=? WHERE example_no=?", (etitle, eno))
            seen.add(eno)
        else:
            eno = str(_next_id(con, "tb_survey_example_new", "example_no"))
            con.execute(
                "INSERT INTO tb_survey_example_new (example_no, question_no, example_title) VALUES (?,?,?)",
                (eno, question_no, etitle))
    for eno in existing - seen:
        con.execute("DELETE FROM tb_survey_example_new WHERE example_no=?", (eno,))


def _save_survey_questions(con, survey_no, questions):
    """설문 문항 upsert — 기존 question_no/example_no 는 보존(응답 연결 유지).
    반환: 저장된 유효 문항 수."""
    survey_no = str(survey_no)
    existing_q = {str(r["question_no"]) for r in con.execute(
        "SELECT question_no FROM tb_survey_question_new WHERE survey_no=?", (survey_no,)).fetchall()}
    seen_q = set()
    order = 0
    for q in questions:
        qtitle = (q.get("questionTitle") or "").strip()
        examples = [e for e in (q.get("examples") or []) if (e.get("exampleTitle") or "").strip()]
        if not qtitle:      # 제목 없는 문항은 저장하지 않음
            continue
        order += 1
        qno = str(q.get("questionNo") or "").strip()
        if qno and qno in existing_q:
            con.execute(
                "UPDATE tb_survey_question_new SET question_title=?, example_cnt=?, question_order=? WHERE question_no=?",
                (qtitle, str(len(examples)), str(order), qno))
            seen_q.add(qno)
        else:
            qno = str(_next_id(con, "tb_survey_question_new", "question_no"))
            con.execute(
                "INSERT INTO tb_survey_question_new (question_no, survey_no, question_title, example_cnt, question_order) VALUES (?,?,?,?,?)",
                (qno, survey_no, qtitle, str(len(examples)), str(order)))
        _save_survey_examples(con, qno, examples)
    for qno in existing_q - seen_q:
        con.execute("DELETE FROM tb_survey_example_new WHERE question_no=?", (qno,))
        con.execute("DELETE FROM tb_survey_question_new WHERE question_no=?", (qno,))
    return order


def api_survey_save(body: dict) -> dict:
    """설문조사 등록(C)/수정(M) — tb_survey_new + 문항(tb_survey_question_new)/보기(tb_survey_example_new)."""
    mode = (body.get("mode") or "C").strip().upper()
    title = (body.get("surveyTitle") or "").strip()
    if not title:
        return {"ok": False, "message": "설문 제목을 입력하세요."}
    start_raw = (body.get("startDate") or "").strip()
    end_raw = (body.get("endDate") or "").strip()
    if not start_raw or not end_raw:
        return {"ok": False, "message": "설문 기간을 입력하세요."}
    start_ms = _date_to_epoch_ms(start_raw)
    end_ms = _date_to_epoch_ms(end_raw, end=True)
    if not start_ms or not end_ms:
        return {"ok": False, "message": "설문 기간 형식이 올바르지 않습니다."}
    if int(end_ms) < int(start_ms):
        return {"ok": False, "message": "종료일이 시작일보다 빠릅니다."}
    cntnts = (body.get("surveyCntnts") or "").strip()
    prsnl_yn = "Y" if str(body.get("prsnlInputYn") or "N").upper() == "Y" else "N"
    prsnl_cntnts = (body.get("prsnlInfoCntnts") or "").strip() if prsnl_yn == "Y" else ""
    now_ms = str(int(datetime.now().timestamp() * 1000))
    actor = (body.get("actor") or "neibis").strip()
    questions = body.get("questions")  # None 이면 문항 미변경, list 면 반영
    with db_connect() as con:
        if mode == "M":
            sid = str(body.get("surveyNo") or body.get("id") or "").strip()
            if not sid:
                return {"ok": False, "message": "수정 대상 survey_no 가 없습니다."}
            cur = con.execute(
                """UPDATE tb_survey_new
                   SET survey_title=?, survey_cntnts=?, start_date=?, end_date=?,
                       prsnl_input_yn=?, prsnl_info_cntnts=?, update_id=?, update_dt=?
                   WHERE survey_no=?""",
                (title, cntnts, start_ms, end_ms, prsnl_yn, prsnl_cntnts, actor, now_ms, sid),
            )
            if cur.rowcount == 0:
                return {"ok": False, "message": "수정할 설문을 찾지 못했습니다."}
            if isinstance(questions, list):
                qcnt = _save_survey_questions(con, sid, questions)
                con.execute("UPDATE tb_survey_new SET question_cnt=? WHERE survey_no=?", (str(qcnt), sid))
            con.commit()
            return {"ok": True, "message": "저장되었습니다.", "mode": "M", "surveyNo": sid}
        sid = _next_id(con, "tb_survey_new", "survey_no")
        con.execute(
            """INSERT INTO tb_survey_new
               (survey_no, survey_title, survey_cntnts, start_date, end_date,
                prsnl_input_yn, prsnl_info_cntnts, question_cnt,
                create_id, create_dt, update_id, update_dt)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (str(sid), title, cntnts, start_ms, end_ms, prsnl_yn, prsnl_cntnts,
             "0", actor, now_ms, actor, now_ms),
        )
        qcnt = 0
        if isinstance(questions, list):
            qcnt = _save_survey_questions(con, sid, questions)
            con.execute("UPDATE tb_survey_new SET question_cnt=? WHERE survey_no=?", (str(qcnt), str(sid)))
        con.commit()
        return {"ok": True, "message": "설문이 등록되었습니다.", "mode": "C", "surveyNo": str(sid)}


def api_survey_delete(body: dict) -> dict:
    """설문조사 삭제 — tb_survey_new (+ 문항/보기 정리). 단건 또는 다건(surveyNos)."""
    ids = body.get("surveyNos")
    if not ids:
        one = str(body.get("surveyNo") or body.get("id") or "").strip()
        ids = [one] if one else []
    ids = [str(x).strip() for x in ids if str(x).strip()]
    if not ids:
        return {"ok": False, "message": "삭제 대상 survey_no 가 없습니다."}
    deleted = 0
    with db_connect() as con:
        for sid in ids:
            qnos = [str(r["question_no"]) for r in con.execute(
                "SELECT question_no FROM tb_survey_question_new WHERE survey_no=?", (sid,)).fetchall()]
            for qno in qnos:
                con.execute("DELETE FROM tb_survey_example_new WHERE question_no=?", (qno,))
            con.execute("DELETE FROM tb_survey_question_new WHERE survey_no=?", (sid,))
            cur = con.execute("DELETE FROM tb_survey_new WHERE survey_no=?", (sid,))
            deleted += cur.rowcount
        con.commit()
    if deleted == 0:
        return {"ok": False, "message": "삭제할 설문을 찾지 못했습니다."}
    return {"ok": True, "message": f"{deleted}건이 삭제되었습니다.", "deleted": deleted}


def api_survey_legacy_list(qs: dict) -> dict:
    """이전(레거시) 설문 목록 — 구버전 tb_survey(+ tb_survey_question/tb_survey_answer).
    읽기 전용 과거 데이터. 참여자 수 = 응답의 distinct sbj_idx."""
    items = []
    with db_connect() as con:
        rows = con.execute(
            """SELECT srv_seq, srv_nm, start_dt, end_dt FROM tb_survey
               ORDER BY CAST(srv_seq AS INTEGER) DESC"""
        ).fetchall()
        for r in rows:
            part = con.execute(
                """SELECT COUNT(DISTINCT a.sbj_idx)
                   FROM tb_survey_answer a
                   JOIN tb_survey_question q ON a.qst_seq = q.qst_seq
                   WHERE q.srv_seq = ?""",
                (r["srv_seq"],),
            ).fetchone()[0]
            items.append({
                "srvSeq": str(r["srv_seq"]),
                "srvNm": r["srv_nm"] or "",
                "startDate": _epoch_ms_to_date(r["start_dt"]),
                "endDate": _epoch_ms_to_date(r["end_dt"]),
                "participants": part or 0,
                "status": _survey_status(r["start_dt"], r["end_dt"]),
            })
    return {"ok": True, "total": len(items), "list": items}


def api_survey_legacy_detail(qs: dict) -> dict:
    """이전(레거시) 설문 상세 — tb_survey + 문항별 응답 통계.
    T(머리말)/C(선택문항, qst_cls_cd 코드그룹 → tb_code_detail 라벨)."""
    sid = (qs.get("id") or qs.get("srvSeq") or [""])[0].strip()
    if not sid:
        return {"ok": False, "message": "srv_seq 가 필요합니다."}
    with db_connect() as con:
        srv = con.execute("SELECT * FROM tb_survey WHERE srv_seq=? LIMIT 1", (sid,)).fetchone()
        if not srv:
            return {"ok": False, "message": "이전 설문을 찾을 수 없습니다."}
        s = dict(srv)
        participants = con.execute(
            """SELECT COUNT(DISTINCT a.sbj_idx) FROM tb_survey_answer a
               JOIN tb_survey_question q ON a.qst_seq = q.qst_seq
               WHERE q.srv_seq = ?""", (sid,)).fetchone()[0]
        items = []
        for q in con.execute(
            """SELECT qst_seq, qst_grp, qst_odr, qst_ty, qst_cls_cd, qst_cntnts
               FROM tb_survey_question WHERE srv_seq=?
               ORDER BY CAST(qst_grp AS INTEGER), CAST(qst_odr AS INTEGER), CAST(qst_seq AS INTEGER)""",
            (sid,),
        ).fetchall():
            title = (q["qst_cntnts"] or "").strip()
            if (q["qst_ty"] or "") == "T":
                if title:
                    items.append({"type": "header", "title": title})
                continue
            # 선택 문항 — 코드그룹의 보기 라벨 + 응답 집계
            grp = q["qst_cls_cd"]
            choices = []
            if grp:
                choices = con.execute(
                    """SELECT code_code, code_name FROM tb_code_detail
                       WHERE group_code=? AND IFNULL(delete_yn,'N')<>'Y' AND IFNULL(use_yn,'Y')<>'N'
                       ORDER BY CAST(code_sort_seq AS INTEGER)""", (grp,)).fetchall()
            counts = {r["ans_cls_cd"]: r["c"] for r in con.execute(
                "SELECT ans_cls_cd, COUNT(*) c FROM tb_survey_answer WHERE qst_seq=? GROUP BY ans_cls_cd",
                (q["qst_seq"],)).fetchall()}
            opts = []
            for c in choices:
                opts.append({"label": c["code_name"] or c["code_code"],
                             "count": counts.pop(c["code_code"], 0)})
            for code, cnt in counts.items():   # 코드표에 없는 응답도 표기
                if code:
                    opts.append({"label": str(code), "count": cnt})
            items.append({"type": "question", "title": title, "options": opts})
    return {"ok": True, "data": {
        "srvSeq": str(s.get("srv_seq") or ""),
        "srvNm": s.get("srv_nm") or "",
        "startDate": _epoch_ms_to_date(s.get("start_dt")),
        "endDate": _epoch_ms_to_date(s.get("end_dt")),
        "status": _survey_status(s.get("start_dt"), s.get("end_dt")),
        "participants": participants or 0,
        "items": items,
    }}


def api_survey_legacy_respondents(qs: dict) -> dict:
    """이전 설문 응답자별 보기 — 응답자(sbj_idx) 목록 + 최종 응답 일시."""
    sid = (qs.get("id") or qs.get("srvSeq") or [""])[0].strip()
    if not sid:
        return {"ok": False, "message": "srv_seq 가 필요합니다."}
    items = []
    with db_connect() as con:
        srv = con.execute("SELECT srv_nm FROM tb_survey WHERE srv_seq=? LIMIT 1", (sid,)).fetchone()
        if not srv:
            return {"ok": False, "message": "이전 설문을 찾을 수 없습니다."}
        rows = con.execute(
            """SELECT a.sbj_idx, MAX(CAST(a.ans_dt AS INTEGER)) mdt
               FROM tb_survey_answer a
               JOIN tb_survey_question q ON a.qst_seq = q.qst_seq
               WHERE q.srv_seq = ?
               GROUP BY a.sbj_idx
               ORDER BY mdt DESC""", (sid,)).fetchall()
        for i, r in enumerate(rows, start=1):
            items.append({
                "no": i,
                "answerSerial": r["sbj_idx"] or "",
                "answerDate": _epoch_ms_to_datetime(r["mdt"]),
            })
    return {"ok": True, "total": len(items), "list": items,
            "srvNm": srv["srv_nm"] or ""}


def api_survey_legacy_answer_detail(qs: dict) -> dict:
    """이전 설문 답변 보기 팝업 — 특정 응답자(sbj_idx)의 문항별 답변(코드→라벨)."""
    def q1(k):
        return (qs.get(k) or [""])[0].strip()
    sid = q1("id") or q1("srvSeq")
    serial = q1("serial") or q1("sbjIdx")
    if not sid or not serial:
        return {"ok": False, "message": "srv_seq 와 응답자 식별자가 필요합니다."}
    answers = []
    with db_connect() as con:
        rows = con.execute(
            """SELECT q.qst_seq, q.qst_cls_cd, q.qst_cntnts, a.ans_cls_cd, a.ans_opn,
                      (SELECT code_name FROM tb_code_detail d
                       WHERE d.group_code = q.qst_cls_cd AND d.code_code = a.ans_cls_cd) AS label
               FROM tb_survey_question q
               LEFT JOIN tb_survey_answer a ON a.qst_seq = q.qst_seq AND a.sbj_idx = ?
               WHERE q.srv_seq = ? AND q.qst_ty = 'C'
               ORDER BY CAST(q.qst_grp AS INTEGER), CAST(q.qst_odr AS INTEGER), CAST(q.qst_seq AS INTEGER)""",
            (serial, sid)).fetchall()
        for i, r in enumerate(rows, start=1):
            label = r["label"] or (r["ans_opn"] if (r["ans_opn"] and r["ans_opn"] != "None") else "") or (r["ans_cls_cd"] or "")
            answers.append({
                "no": i,
                "questionTitle": (r["qst_cntnts"] or "").strip(),
                "answerLabel": label,
            })
    return {"ok": True, "answerSerial": serial, "answers": answers}


def api_oral_source_search(qs: dict) -> dict:
    """제보자 검색 — 제보자 관리(wb_source)에서 후보를 중복 제거해 반환.
    EAF 자동입력된 이름/성별/생년으로 기존 제보자와 매칭·연결하는 데 사용."""
    def q1(k, d=""):
        return (qs.get(k) or [d])[0].strip()
    name = q1("name") or q1("q")
    sex = q1("sex")   # '여'/'남' 또는 코드
    birth = q1("birth")
    where, params = [], []
    if name:
        where.append("name LIKE ?"); params.append("%" + name + "%")
    if birth:
        where.append("IFNULL(birth,'') = ?"); params.append(birth)
    if sex in ("여", "0"):
        where.append("sex = '0'")
    elif sex in ("남", "1"):
        where.append("sex = '1'")
    wh = ("WHERE " + " AND ".join(where)) if where else ""

    sex_lbl = {"0": "여", "1": "남"}
    items = []
    with db_connect() as con:
        rows = con.execute(
            f"""SELECT name, sex, IFNULL(birth,'') birth, IFNULL(residence,'') residence,
                       IFNULL(birth_place,'') birth_place, IFNULL(job,'') job,
                       IFNULL(education,'') education,
                       MIN(CAST(source_id AS INTEGER)) source_id,
                       MAX(IFNULL(age,'')) age, MAX(IFNULL(region_nm,'')) region_nm,
                       COUNT(*) cnt
                FROM wb_source {wh}
                GROUP BY name, sex, IFNULL(birth,''), IFNULL(residence,'')
                ORDER BY cnt DESC, name
                LIMIT 40""",
            params,
        ).fetchall()
        for r in rows:
            items.append({
                "sourceId": str(r["source_id"]),
                "name": r["name"] or "",
                "sex": sex_lbl.get(r["sex"] or "", r["sex"] or ""),
                "age": r["age"] or "",
                "birth": r["birth"] or "",
                "residence": r["residence"] or "",
                "birthPlace": r["birth_place"] or "",
                "job": r["job"] or "",
                "education": r["education"] or "",
                "regionNm": r["region_nm"] or "",
                "surveyCount": r["cnt"],
            })
    return {"ok": True, "total": len(items), "list": items}


def _find_oral_media(oral_id: str, ext: str) -> Path | None:
    """id + 확장자로 원천 파일 경로 탐색 (.wav/.eaf)."""
    if not re.fullmatch(r"[A-Za-z0-9_\-]+", oral_id):
        return None
    for p in _scan_oral_files():
        if _oral_id_of(p) == oral_id:
            if ext == "eaf":
                return p if p.is_file() else None
            cand = p.with_name(oral_id + "." + ext)
            return cand if cand.is_file() else None
    return None


def api_vocab_detail(qs: dict) -> dict:
    """어휘조사자료 상세 (dialect_region_id 기준 + 동일 dialect_id 지역/용례)."""
    rid = (
        (qs.get("id") or qs.get("dialectRegionId") or qs.get("dialect_region_id") or [""])[0]
        .strip()
    )
    if not rid:
        return {"ok": False, "message": "id(dialect_region_id)가 필요합니다."}

    def sex_label(v) -> str:
        s = str(v or "").strip()
        if s == "0":
            return "남"
        if s == "1":
            return "여"
        return s or ""

    with db_connect() as con:
        row = con.execute(
            """
            SELECT * FROM tb_dialect_region
            WHERE dialect_region_id = ?
            LIMIT 1
            """,
            (rid,),
        ).fetchone()
        if not row:
            return {"ok": False, "message": f"자료를 찾을 수 없습니다. (id={rid})"}

        dialect_id = str(row["dialect_id"] or "").strip()
        new_row = None
        if dialect_id:
            new_row = con.execute(
                """
                SELECT * FROM tb_dialect_new
                WHERE dialect_id = ?
                LIMIT 1
                """,
                (dialect_id,),
            ).fetchone()

        # 동일 dialect_id 의 지역(음성) 블록 전체 (없으면 현재 1건)
        if dialect_id:
            regions = con.execute(
                """
                SELECT * FROM tb_dialect_region
                WHERE dialect_id = ?
                ORDER BY CAST(dialect_region_id AS INTEGER)
                """,
                (dialect_id,),
            ).fetchall()
        else:
            regions = [row]

        examples = []
        if dialect_id:
            examples = con.execute(
                """
                SELECT * FROM tb_dialect_example
                WHERE dialect_id = ?
                ORDER BY CAST(example_id AS INTEGER)
                """,
                (dialect_id,),
            ).fetchall()

    n = dict(new_row) if new_row else {}
    r0 = dict(row)

    # 표제어 계열: dialect_new 우선, 없으면 region
    dlt_tp = (n.get("dlt_tp") or r0.get("dlt_tp") or "") or ""
    std_tp = (n.get("std_tp") or r0.get("std_tp") or "") or ""
    item_nm = (n.get("item_nm") or r0.get("item_nm") or "") or ""
    use_yn = (n.get("use_yn") or r0.get("use_yn") or "") or ""

    region_items = []
    for r in regions:
        rd = dict(r)
        region_items.append(
            {
                "dialectRegionId": rd.get("dialect_region_id") or "",
                "dialectId": rd.get("dialect_id") or "",
                "serialNm": rd.get("serial_nm") or "",
                "source": rd.get("source") or "",
                "sidoCd": rd.get("sido_cd") or "",
                "sidoNm": rd.get("sido_nm") or "",
                "sigunguNm": rd.get("sigungu_nm") or "",
                "basisYear": rd.get("basis_year") or "",
                "sex": rd.get("sex") or "",
                "sexLabel": sex_label(rd.get("sex")),
                "age": rd.get("age") or "",
                "useYn": rd.get("use_yn") or "",
                "infoYn": (rd.get("info_yn") or "").strip(),
                "fileMemo": rd.get("file_memo") or "",
                "etc": rd.get("etc") or "",
                "dltTell": rd.get("dlt_tell") or "",
                "fileOpenYn": "",  # filled below if needed
                "itemNm": rd.get("item_nm") or "",
                "dltTp": rd.get("dlt_tp") or "",
                "stdTp": rd.get("std_tp") or "",
                "researchDegree": rd.get("research_degree") or "",
            }
        )

    # 파일공개: dialect_new 공통 값 (지역 단위 컬럼 없음)
    file_open = (n.get("file_open_yn") or "") or ""
    for it in region_items:
        it["fileOpenYn"] = file_open

    example_items = []
    for e in examples:
        ed = dict(e)
        example_items.append(
            {
                "exampleId": ed.get("example_id") or "",
                "dialectId": ed.get("dialect_id") or "",
                "dltExam": ed.get("dlt_exam") or "",
                "stdExam": ed.get("std_exam") or "",
                "sidoCodeExam": ed.get("sido_code_exam") or "",
                "sigunguCodeExam": ed.get("sigungu_code_exam") or "",
                "sidoNmExam": ed.get("sido_nm_exam") or "",
                "sigunguNmExam": ed.get("sigungu_nm_exam") or "",
                "sourceExam": ed.get("source_exam") or "",
            }
        )

    return {
        "ok": True,
        "dialectRegionId": r0.get("dialect_region_id") or "",
        "dialectId": dialect_id,
        "dltTp": dlt_tp,
        "dltSeg": (n.get("dlt_seg") or "") or "",
        "wordClass": (n.get("word_class") or "") or "",
        "stdTp": std_tp,
        "addMean": (n.get("add_mean") or "") or "",
        "mean": (n.get("mean") or "") or "",
        "itemNm": item_nm,
        "useYn": use_yn,
        "relDlt": (n.get("rel_dlt") or "") or "",
        "reference": (n.get("reference") or "") or "",
        "memo": (n.get("memo") or "") or "",
        "fileOpenYn": file_open,
        "regions": region_items,
        "examples": example_items,
        "db": str(DB_PATH),
    }


def api_vocab_list(qs: dict) -> dict:
    """어휘조사자료 목록 (tb_dialect_region + tb_dialect_new.file_open_yn/mean)."""
    try:
        page = max(1, int((qs.get("page") or ["1"])[0]))
    except ValueError:
        page = 1
    try:
        page_size = int((qs.get("pageSize") or ["10"])[0])
    except ValueError:
        page_size = 10
    if page_size not in (10, 50, 100, 200, 300):
        page_size = 10

    source = (qs.get("source") or [""])[0].strip()  # '' | dlt | sam
    use_yn = (qs.get("useYn") or qs.get("use_yn") or [""])[0].strip().upper()  # '' | Y | N
    sex = (qs.get("sex") or [""])[0].strip()  # '' | man | woman
    degree = (qs.get("degree") or [""])[0].strip()
    start_year = (qs.get("startYear") or [""])[0].strip()
    end_year = (qs.get("endYear") or [""])[0].strip()
    # ages: "2,5,7" (decade codes) or empty = all
    ages_raw = (qs.get("ages") or qs.get("age") or [""])[0].strip()
    # regions: "42:삼척시,41:"  (sido_cd:sigungu_nm, empty sigungu = whole sido)
    regions_raw = (qs.get("regions") or [""])[0].strip()
    # targets JSON: [{"field":"dlt|std|mean","op":"contains|match|startsWith|endsWith","value":"...","join":"in|notin"}]
    targets_raw = (qs.get("targets") or [""])[0].strip()
    # simple free-text fallback (방언형/표준어/일련번호/항목번호)
    q = (qs.get("q") or qs.get("searchValue") or [""])[0].strip()

    where: list[str] = []
    params: list = []
    need_mean_join = False

    if source == "dlt":
        where.append("(r.source LIKE '%지역어%' OR IFNULL(r.source,'') = '')")
    elif source == "sam":
        where.append("r.source LIKE '%우리말샘%'")

    if use_yn in ("Y", "N"):
        where.append("IFNULL(r.use_yn,'') = ?")
        params.append(use_yn)

    if sex == "man":
        where.append("r.sex = '0'")
    elif sex == "woman":
        where.append("r.sex = '1'")

    if degree in ("1", "2"):
        where.append("r.research_degree = ?")
        params.append(degree)

    if start_year.isdigit():
        where.append("CAST(IFNULL(r.basis_year,'0') AS INTEGER) >= ?")
        params.append(int(start_year))
    if end_year.isdigit():
        where.append("CAST(IFNULL(r.basis_year,'0') AS INTEGER) <= ?")
        params.append(int(end_year))

    if ages_raw:
        age_parts = [a.strip() for a in ages_raw.split(",") if a.strip().isdigit()]
        age_ors = []
        for a in age_parts:
            decade = int(a)
            if decade == 9:
                age_ors.append("CAST(IFNULL(r.age,'0') AS INTEGER) >= 90")
            elif decade > 0:
                lo = decade * 10
                hi = lo + 9
                age_ors.append(
                    "(CAST(IFNULL(r.age,'0') AS INTEGER) BETWEEN ? AND ?)"
                )
                params.extend([lo, hi])
        if age_ors:
            where.append("(" + " OR ".join(age_ors) + ")")

    if regions_raw:
        region_ors = []
        for part in regions_raw.split(","):
            part = part.strip()
            if not part:
                continue
            if ":" in part:
                sido_cd, sigungu_nm = part.split(":", 1)
            else:
                sido_cd, sigungu_nm = part, ""
            sido_cd = sido_cd.strip()
            sigungu_nm = sigungu_nm.strip()
            if not sido_cd and not sigungu_nm:
                continue
            bits = []
            if sido_cd:
                bits.append("r.sido_cd = ?")
                params.append(sido_cd)
            if sigungu_nm:
                bits.append("r.sigungu_nm = ?")
                params.append(sigungu_nm)
            if bits:
                region_ors.append("(" + " AND ".join(bits) + ")")
        if region_ors:
            where.append("(" + " OR ".join(region_ors) + ")")

    # multi-target text search
    targets: list = []
    if targets_raw:
        try:
            parsed = json.loads(targets_raw)
            if isinstance(parsed, list):
                targets = parsed
        except Exception:
            targets = []

    field_map = {
        "dlt": "r.dlt_tp",
        "std": "r.std_tp",
        "mean": "IFNULL(n.mean,'')",
        "serial": "r.serial_nm",
        "item": "IFNULL(r.item_nm,'')",
    }

    def _like_clause(col: str, op: str, value: str) -> tuple[str, list]:
        op = (op or "contains").strip()
        if op == "match":
            return f"{col} = ?", [value]
        if op == "startsWith":
            return f"{col} LIKE ?", [value + "%"]
        if op == "endsWith":
            return f"{col} LIKE ?", ["%" + value]
        return f"{col} LIKE ?", ["%" + value + "%"]

    for i, t in enumerate(targets):
        if not isinstance(t, dict):
            continue
        val = str(t.get("value") or "").strip()
        if not val:
            continue
        field = str(t.get("field") or "dlt").strip()
        col = field_map.get(field, "r.dlt_tp")
        if field == "mean":
            need_mean_join = True
        clause, cparams = _like_clause(col, str(t.get("op") or "contains"), val)
        join = str(t.get("join") or "in").strip()  # in = AND, notin = AND NOT
        if join == "notin":
            where.append("NOT (" + clause + ")")
        else:
            where.append(clause)
        params.extend(cparams)

    if q and not targets:
        where.append(
            "(r.dlt_tp LIKE ? OR r.std_tp LIKE ? OR r.serial_nm LIKE ? OR IFNULL(r.item_nm,'') LIKE ?)"
        )
        like = f"%{q}%"
        params.extend([like, like, like, like])

    wh = (" WHERE " + " AND ".join(where)) if where else ""
    # mean 검색 시에만 dialect_new 조인 (dialect_id 1건으로 축약해 중복 방지)
    if need_mean_join:
        join_sql = (
            " LEFT JOIN ("
            "   SELECT dialect_id, MAX(IFNULL(mean,'')) AS mean"
            "   FROM tb_dialect_new GROUP BY dialect_id"
            " ) n ON n.dialect_id = r.dialect_id "
        )
    else:
        join_sql = ""

    with db_connect() as con:
        total = con.execute(
            f"SELECT COUNT(*) FROM tb_dialect_region r{join_sql}{wh}", params
        ).fetchone()[0]
        offset = (page - 1) * page_size
        rows = con.execute(
            f"""
            SELECT r.dialect_region_id, r.dialect_id, r.serial_nm, r.item_nm,
                   r.dlt_tp, r.std_tp, r.use_yn, r.source,
                   r.sido_cd, r.sido_nm, r.sigungu_nm, r.basis_year,
                   r.sex, r.age, r.research_degree
            FROM tb_dialect_region r
            {join_sql}
            {wh}
            ORDER BY CAST(r.dialect_region_id AS INTEGER) DESC
            LIMIT ? OFFSET ?
            """,
            params + [page_size, offset],
        ).fetchall()

        # 페이지 단위 file_open_yn / mean 배치 조회 (JOIN 중복 회피)
        extra: dict[str, dict] = {}
        dialect_ids = [str(r["dialect_id"] or "") for r in rows if r["dialect_id"]]
        dialect_ids = list(dict.fromkeys(dialect_ids))  # unique, order-preserving
        if dialect_ids:
            ph = ",".join("?" * len(dialect_ids))
            for nr in con.execute(
                f"""
                SELECT dialect_id, file_open_yn, mean
                FROM tb_dialect_new
                WHERE dialect_id IN ({ph})
                """,
                dialect_ids,
            ).fetchall():
                did = str(nr["dialect_id"] or "")
                if did and did not in extra:
                    extra[did] = {
                        "file_open_yn": nr["file_open_yn"] or "",
                        "mean": nr["mean"] or "",
                    }

    def sex_label(v) -> str:
        s = str(v or "").strip()
        if s == "0":
            return "남"
        if s == "1":
            return "여"
        return s or "-"

    items = []
    for r in rows:
        did = str(r["dialect_id"] or "")
        ex = extra.get(did) or {}
        items.append(
            {
                "dialectRegionId": r["dialect_region_id"] or "",
                "dialectId": did,
                "serialNm": r["serial_nm"] or "",
                "itemNm": r["item_nm"] or "",
                "dltTp": r["dlt_tp"] or "",
                "stdTp": r["std_tp"] or "",
                "useYn": r["use_yn"] or "",
                "source": r["source"] or "",
                "sidoCd": r["sido_cd"] or "",
                "sidoNm": r["sido_nm"] or "",
                "sigunguNm": r["sigungu_nm"] or "",
                "basisYear": r["basis_year"] or "",
                "sex": r["sex"] or "",
                "sexLabel": sex_label(r["sex"]),
                "age": r["age"] or "",
                "researchDegree": r["research_degree"] or "",
                "fileOpenYn": ex.get("file_open_yn") or "",
                "mean": ex.get("mean") or "",
            }
        )

    total_pages = max(1, (total + page_size - 1) // page_size)
    return {
        "ok": True,
        "total": total,
        "page": page,
        "pageSize": page_size,
        "totalPages": total_pages,
        "list": items,
        "db": str(DB_PATH),
    }


# 부호 이미지 원본 위치 — DB icon 이 비어 있는 행(운영 CUBRID 에서 파일로만 관리)용
SYMBOL_ASSET_DIR = USER_MAP_ROOT / "symbol"
SYMBOL_MASK_DIR = USER_MAP_ROOT / "symbol_mask"


def _symbol_icon(icon, file_nm):
    """(iconSrc, iconKind) 반환.
    ① tb_map_symbol.icon(raw base64) → data URI
    ② 없으면 symbol/<file_nm> 실제 부호 이미지
    ③ 그것도 없으면 symbol_mask/<file_nm> 흰색 실루엣(화면에서 반전해 표시)"""
    v = str(icon or "")
    if v.startswith("data:"):
        return v, "icon"
    if len(v) > 20:
        return f"data:image/png;base64,{v}", "icon"
    fn = str(file_nm or "").strip()
    if fn and "/" not in fn and "\\" not in fn and ".." not in fn:
        if (SYMBOL_ASSET_DIR / fn).is_file():
            return "/user-map/symbol/" + urllib.parse.quote(fn), "file"
        if (SYMBOL_MASK_DIR / fn).is_file():
            return "/user-map/symbol_mask/" + urllib.parse.quote(fn), "mask"
    return "", ""


def _symbol_where(qs: dict):
    """검색조건(유형 A/B) → WHERE 절."""
    search = (qs.get("searchValue") or [""])[0].strip()
    if search in ("A", "B"):
        return " WHERE s.symbol_shape = ?", [search]
    return "", []


def _symbol_row(r) -> dict:
    src, kind = _symbol_icon(r["icon"], r["file_nm"])
    return {
        "symbolId": r["symbol_id"],
        "symbolNm": r["symbol_nm"] or "",
        "comment": r["comment"] or "",
        "useYn": r["use_yn"] or "",
        "regDt": fmt_reg_dt(r["reg_dt"], r["upt_dt"], r["file_nm"]),
        "symbolGb": r["symbol_gb"] or "",
        "symbolShape": r["symbol_shape"] or "",
        "fileNm": r["file_nm"] or "",
        "iconSrc": src,
        "iconKind": kind,
    }


SYMBOL_SELECT = """
    SELECT s.symbol_id, s.symbol_nm, s.comment, s.use_yn, s.reg_dt, s.upt_dt,
           s.symbol_gb, s.symbol_shape, s.map_symbol_id, m.file_nm, m.icon
    FROM tb_symbol s
    LEFT JOIN tb_map_symbol m ON m.map_symbol_id = s.map_symbol_id
"""


def api_symbol_list(qs: dict) -> dict:
    try:
        page = max(1, int((qs.get("page") or ["1"])[0]))
    except ValueError:
        page = 1
    try:
        page_size = int((qs.get("pageSize") or ["10"])[0])
    except ValueError:
        page_size = 10
    if page_size not in (10, 50, 100, 200, 300):
        page_size = 10

    wh, params = _symbol_where(qs)

    with db_connect() as con:
        total = con.execute(
            f"SELECT COUNT(*) FROM tb_symbol s{wh}", params
        ).fetchone()[0]
        offset = (page - 1) * page_size
        rows = con.execute(
            f"""{SYMBOL_SELECT}{wh}
            ORDER BY CAST(s.symbol_id AS INTEGER) DESC
            LIMIT ? OFFSET ?
            """,
            params + [page_size, offset],
        ).fetchall()

    items = [_symbol_row(r) for r in rows]
    total_pages = max(1, (total + page_size - 1) // page_size)
    return {
        "ok": True,
        "total": total,
        "page": page,
        "pageSize": page_size,
        "totalPages": total_pages,
        "list": items,
        "db": str(DB_PATH),
    }


def api_symbol_detail(qs: dict) -> dict:
    symbol_id = (qs.get("symbolId") or [""])[0].strip()
    if not symbol_id:
        return {"ok": False, "message": "symbolId 가 필요합니다."}
    with db_connect() as con:
        r = con.execute(
            f"""{SYMBOL_SELECT}
            WHERE CAST(s.symbol_id AS INTEGER) = CAST(? AS INTEGER)
            """,
            (symbol_id,),
        ).fetchone()
    if not r:
        return {"ok": False, "message": "해당 상징부호를 찾을 수 없습니다."}
    item = _symbol_row(r)
    item["mapSymbolId"] = r["map_symbol_id"] or ""
    return {"ok": True, "item": item}


def _symbol_next_id(con) -> int:
    a = con.execute("SELECT MAX(CAST(symbol_id AS INTEGER)) FROM tb_symbol").fetchone()[0]
    b = con.execute("SELECT MAX(CAST(map_symbol_id AS INTEGER)) FROM tb_map_symbol").fetchone()[0]
    return max(int(a or 0), int(b or 0)) + 1


def api_symbol_save(body: dict) -> dict:
    """상징부호 등록(C)/수정(M). 이미지는 base64(data URI 허용)로 받아
    tb_map_symbol.icon 에 원본과 동일한 형태(raw base64)로 저장한다."""
    mode = str(body.get("mode") or "C").upper()
    symbol_gb = str(body.get("symbolGb") or "I").upper()
    if symbol_gb not in ("I", "T"):
        symbol_gb = "I"
    symbol_shape = str(body.get("symbolShape") or "A").upper()
    if symbol_shape not in ("A", "B"):
        symbol_shape = "A"
    symbol_nm = str(body.get("symbolNm") or "").strip()
    comment = str(body.get("comment") or "").strip()
    use_yn = "N" if str(body.get("useYn") or "Y").upper() == "N" else "Y"
    reg_id = str(body.get("regId") or "admin").strip() or "admin"

    icon = str(body.get("icon") or "").strip()
    if icon.startswith("data:"):
        icon = icon.split(",", 1)[-1].strip()
    file_nm = str(body.get("fileNm") or "").strip()
    now_ms = str(int(time.time() * 1000))

    if symbol_gb == "T":
        if not symbol_nm:
            return {"ok": False, "message": "상징부호(텍스트)를 입력해 주세요."}
        icon, file_nm = "", ""
    elif mode == "C" and not icon:
        return {"ok": False, "message": "상징부호 화일을 선택하여 주십시요."}

    if icon and not file_nm:
        file_nm = now_ms + ".png"

    with db_connect() as con:
        if mode == "M":
            symbol_id = str(body.get("symbolId") or "").strip()
            if not symbol_id:
                return {"ok": False, "message": "symbolId 가 필요합니다."}
            row = con.execute(
                "SELECT symbol_id, map_symbol_id FROM tb_symbol"
                " WHERE CAST(symbol_id AS INTEGER) = CAST(? AS INTEGER)",
                (symbol_id,),
            ).fetchone()
            if not row:
                return {"ok": False, "message": "해당 상징부호를 찾을 수 없습니다."}
            map_symbol_id = row["map_symbol_id"] or symbol_id
            if icon:
                exists = con.execute(
                    "SELECT 1 FROM tb_map_symbol"
                    " WHERE CAST(map_symbol_id AS INTEGER) = CAST(? AS INTEGER)",
                    (map_symbol_id,),
                ).fetchone()
                if exists:
                    con.execute(
                        "UPDATE tb_map_symbol SET icon = ?, file_nm = ?"
                        " WHERE CAST(map_symbol_id AS INTEGER) = CAST(? AS INTEGER)",
                        (icon, file_nm, map_symbol_id),
                    )
                else:
                    con.execute(
                        "INSERT INTO tb_map_symbol (map_symbol_id, icon, file_nm)"
                        " VALUES (?,?,?)",
                        (str(map_symbol_id), icon, file_nm),
                    )
            con.execute(
                """UPDATE tb_symbol
                      SET symbol_shape = ?, comment = ?, symbol_nm = ?, use_yn = ?,
                          symbol_gb = ?, upt_id = ?, upt_dt = ?
                    WHERE CAST(symbol_id AS INTEGER) = CAST(? AS INTEGER)""",
                (symbol_shape, comment, symbol_nm, use_yn, symbol_gb,
                 reg_id, now_ms, symbol_id),
            )
            con.commit()
            return {"ok": True, "symbolId": int(symbol_id), "message": "수정되었습니다."}

        new_id = _symbol_next_id(con)
        if icon:
            con.execute(
                "INSERT INTO tb_map_symbol (map_symbol_id, icon, file_nm) VALUES (?,?,?)",
                (str(new_id), icon, file_nm),
            )
        con.execute(
            """INSERT INTO tb_symbol
                 (symbol_id, symbol_shape, comment, symbol_nm, use_yn,
                  reg_id, reg_dt, upt_id, upt_dt, symbol_gb, map_symbol_id)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (str(new_id), symbol_shape, comment, symbol_nm, use_yn,
             reg_id, now_ms, "", "", symbol_gb, str(new_id)),
        )
        con.commit()
        return {"ok": True, "symbolId": new_id, "message": "등록되었습니다."}


def api_symbol_delete(body: dict) -> dict:
    """선택 삭제 — tb_symbol / tb_symbol_mapp / tb_map_symbol 동시 정리."""
    raw = body.get("ids") or body.get("checkList") or []
    if isinstance(raw, str):
        raw = [x for x in re.split(r"[,\s]+", raw) if x]
    ids = []
    for v in raw:
        try:
            ids.append(int(str(v).strip()))
        except (TypeError, ValueError):
            continue
    if not ids:
        return {"ok": False, "message": "삭제를 위한 상징부호를 선택해주세요."}

    with db_connect() as con:
        marks = ",".join("?" * len(ids))
        maps = [
            r[0]
            for r in con.execute(
                f"SELECT map_symbol_id FROM tb_symbol"
                f" WHERE CAST(symbol_id AS INTEGER) IN ({marks})",
                ids,
            ).fetchall()
            if r[0]
        ]
        con.execute(
            f"DELETE FROM tb_symbol_mapp WHERE CAST(symbol_id AS INTEGER) IN ({marks})",
            ids,
        )
        cur = con.execute(
            f"DELETE FROM tb_symbol WHERE CAST(symbol_id AS INTEGER) IN ({marks})", ids
        )
        deleted = cur.rowcount
        if maps:
            mm = ",".join("?" * len(maps))
            con.execute(
                f"DELETE FROM tb_map_symbol WHERE map_symbol_id IN ({mm})", maps
            )
        con.commit()
    return {"ok": True, "deleted": deleted, "message": f"{deleted}건이 삭제되었습니다."}


def build_symbol_excel(qs: dict):
    """상징부호 자료내려받기 — 검색조건 그대로 전체 건 엑셀."""
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wh, params = _symbol_where(qs)
    with db_connect() as con:
        rows = con.execute(
            f"""{SYMBOL_SELECT}{wh}
            ORDER BY CAST(s.symbol_id AS INTEGER) DESC
            """,
            params,
        ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "상징부호"
    headers = ["번호", "상징부호 ID", "구분", "유형", "상징부호", "설명", "상태", "등록일"]
    ws.append(headers)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2563EB")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows, start=1):
        gb = (r["symbol_gb"] or "").upper()
        ws.append([
            i,
            r["symbol_id"],
            "텍스트" if gb == "T" else "이미지",
            "유형 " + (r["symbol_shape"] or "-"),
            (r["symbol_nm"] or "-") if gb == "T" else (r["file_nm"] or "-"),
            r["comment"] or "",
            "사용" if (r["use_yn"] or "") == "Y" else "미사용",
            fmt_reg_dt(r["reg_dt"], r["upt_dt"], r["file_nm"]) or "",
        ])
    for col, w in zip("ABCDEFGH", (8, 14, 10, 10, 34, 34, 10, 14)):
        ws.column_dimensions[col].width = w
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), "상징부호_목록.xlsx"


# 게시 상태(status) 정의 — kd_headword.use_yn / appro 조합만으로 표현
#   published(게시중): use_yn='Y'
#   pending(승인대기): appro='Y' AND use_yn!='Y'
#   draft(미게시):     use_yn!='Y' AND appro!='Y'
STATUS_SQL = {
    "published": "h.use_yn = 'Y'",
    "pending": "h.appro = 'Y' AND IFNULL(h.use_yn,'') != 'Y'",
    "draft": "IFNULL(h.use_yn,'') != 'Y' AND IFNULL(h.appro,'') != 'Y'",
}


_silhouette_cache: dict = {}


def _tintable_silhouette(icon_b64: str) -> str:
    """icon(검정 잉크+흰 내부 디테일)에서 '어둡고 불투명한' 픽셀만 흰색 실루엣으로 추출.
    흰색 내부 디테일은 구멍(투명)으로 보존 → 모양 구분 유지 + 그룹색으로 틴트 가능."""
    icon = str(icon_b64 or "").strip()
    if not icon:
        return ""
    if icon.startswith("data:"):
        icon = icon.split(",", 1)[-1]
    if icon in _silhouette_cache:
        return _silhouette_cache[icon]
    uri = ""
    try:
        import base64 as _b64, io as _io
        from PIL import Image, ImageChops
        im = Image.open(_io.BytesIO(_b64.b64decode(icon))).convert("RGBA")
        alpha = im.getchannel("A")
        gray = im.convert("L")
        opaque = alpha.point(lambda v: 255 if v > 90 else 0)
        dark = gray.point(lambda v: 255 if v < 140 else 0)
        sil = ImageChops.multiply(opaque, dark)     # 불투명 AND 어두운 곳만
        if sil.getextrema()[1] == 0:                # 밝은 아이콘이면 알파 전체로 폴백
            sil = opaque
        white = Image.new("RGBA", im.size, (255, 255, 255, 255))
        white.putalpha(sil)
        buf = _io.BytesIO()
        white.save(buf, format="PNG")
        uri = "data:image/png;base64," + _b64.b64encode(buf.getvalue()).decode()
    except Exception:
        uri = ""
    _silhouette_cache[icon] = uri
    return uri


def api_symbol_catalog(qs: dict) -> dict:
    """편집기 부호 그리드용 카탈로그 — symbol.do 와 동일 소스(DB tb_map_symbol).
    icon(base64) data URI 를 그대로 제공(마스크 중복 문제 해소 · symbol.do 와 동일 이미지)."""
    with db_connect() as con:
        rows = con.execute(
            """
            SELECT m.map_symbol_id, m.file_nm, m.icon,
                   (SELECT s.symbol_nm FROM tb_symbol s
                     WHERE s.map_symbol_id = m.map_symbol_id
                       AND IFNULL(s.symbol_nm,'') != '' LIMIT 1) AS nm
            FROM tb_map_symbol m
            WHERE IFNULL(m.file_nm,'') != ''
            ORDER BY CAST(m.map_symbol_id AS INTEGER)
            """
        ).fetchall()
    items = [
        {
            "mapSymbolId": r["map_symbol_id"],
            "fileNm": r["file_nm"],
            "iconSrc": _tintable_silhouette(r["icon"]),
            "label": r["nm"] or ("부호 " + str(r["map_symbol_id"])),
        }
        for r in rows
    ]
    return {"ok": True, "total": len(items), "list": items}


def api_headword_list(qs: dict) -> dict:
    """지역어 지도 표제어 목록 (kd_headword)."""
    search_text = (qs.get("searchText") or [""])[0].strip()
    search_type = (qs.get("SearchType") or ["3"])[0].strip() or "3"
    status = (qs.get("status") or [""])[0].strip().lower()  # ''=전체 / pending / published / draft
    try:
        page = max(1, int((qs.get("page") or ["1"])[0]))
    except ValueError:
        page = 1
    try:
        page_size = int((qs.get("pageSize") or ["10"])[0])
    except ValueError:
        page_size = 10
    if page_size not in (10, 50, 100, 200, 300):
        page_size = 10

    where = []
    params: list = []

    # 등록자(계정) 검색 — 이름은 중복이 많아(615명 중 124개 계정이 동명이인) 계정만 받는다.
    #   사용자 계정관리에서 「나의 지도」 건수를 누르면 ?writer=계정 으로 진입한다.
    writer = (qs.get("writer") or qs.get("usid") or [""])[0].strip()
    if writer:
        where.append("h.usid = ?")
        params.append(writer)

    # 탈퇴 회원이 등록한 지도는 기본 숨김 (관리자 화면에서 includeSecsn=1 로만 조회)
    if (qs.get("includeSecsn") or ["0"])[0] not in ("1", "true", "Y", "y"):
        where.append(
            f"NOT EXISTS (SELECT 1 FROM pt_user u WHERE u.usid = h.usid AND IFNULL(u.auth,'') = '{_AUTH_WITHDRAWN}')"
        )

    # SearchType: 1관리자 2사용자 3전체 4게시승인
    if search_type == "1":
        # 관리자 계열(map*, isskor 등) + 빈 usid
        where.append(
            "(h.usid IS NULL OR trim(h.usid) = '' OR h.usid LIKE 'map%' OR h.usid IN ('isskor','iss','admin','diquest'))"
        )
    elif search_type == "2":
        where.append(
            "(h.usid IS NOT NULL AND trim(h.usid) != '' AND h.usid NOT LIKE 'map%' AND h.usid NOT IN ('isskor','iss','admin','diquest'))"
        )
    elif search_type == "4":
        where.append("(h.appro = 'Y' OR h.use_yn = 'Y')")
    # type 3: no filter

    if search_text:
        where.append("(h.headword LIKE ? OR h.headword_no LIKE ? OR IFNULL(h.meaning,'') LIKE ?)")
        like = f"%{search_text}%"
        params.extend([like, like, like])

    # 구분·검색어까지가 상태 요약(counts)의 모수. 상태 필터는 목록에만 추가 적용.
    base_wh = (" WHERE " + " AND ".join(where)) if where else ""
    base_params = list(params)

    list_where = list(where)
    if status in STATUS_SQL:
        list_where.append("(" + STATUS_SQL[status] + ")")
    wh = (" WHERE " + " AND ".join(list_where)) if list_where else ""

    with db_connect() as con:
        # 상태별 요약 카운트 (현재 구분·검색어 기준)
        crow = con.execute(
            f"""
            SELECT COUNT(*) AS all_cnt,
                   SUM(CASE WHEN {STATUS_SQL['published']} THEN 1 ELSE 0 END) AS published_cnt,
                   SUM(CASE WHEN {STATUS_SQL['pending']}   THEN 1 ELSE 0 END) AS pending_cnt,
                   SUM(CASE WHEN {STATUS_SQL['draft']}     THEN 1 ELSE 0 END) AS draft_cnt
            FROM kd_headword h{base_wh}
            """,
            base_params,
        ).fetchone()
        counts = {
            "all": crow["all_cnt"] or 0,
            "published": crow["published_cnt"] or 0,
            "pending": crow["pending_cnt"] or 0,
            "draft": crow["draft_cnt"] or 0,
        }
        total = con.execute(
            f"SELECT COUNT(*) FROM kd_headword h{wh}", params
        ).fetchone()[0]
        offset = (page - 1) * page_size
        rows = con.execute(
            f"""
            SELECT h.headword_id, h.headword_no, h.headword, h.word_class,
                   h.meaning, h.commentary, h.usid, h.use_yn, h.appro, h.topic_id, h.create_dt,
                   (SELECT COUNT(*) FROM tb_headword_dialect d
                     WHERE d.headword_no = h.headword_no) AS word_count,
                   (SELECT COUNT(DISTINCT d.mutation_group) FROM tb_headword_dialect d
                     WHERE d.headword_no = h.headword_no
                       AND d.mutation_group IS NOT NULL AND trim(d.mutation_group) != '') AS group_count
            FROM kd_headword h
            {wh}
            ORDER BY CAST(h.headword_no AS INTEGER) DESC
            LIMIT ? OFFSET ?
            """,
            params + [page_size, offset],
        ).fetchall()

    def clip(s, n=80):
        s = (s or "").replace("\n", " ").strip()
        return s if len(s) <= n else s[: n - 1] + "…"

    def status_of(use_yn, appro):
        if (use_yn or "") == "Y":
            return "published"
        if (appro or "") == "Y":
            return "pending"
        return "draft"

    items = []
    for r in rows:
        items.append(
            {
                "headwordId": r["headword_id"],
                "headwordNo": r["headword_no"],
                "headword": r["headword"] or "",
                "wordClass": r["word_class"] or "",
                "wordCount": r["word_count"] or 0,
                "groupCount": r["group_count"] or 0,
                "meaning": clip(r["meaning"], 100),
                "meaningFull": r["meaning"] or "",
                "commentary": clip(r["commentary"], 100),
                "commentaryFull": r["commentary"] or "",
                "appro": r["appro"] or "",
                "useYn": r["use_yn"] or "",
                "status": status_of(r["use_yn"], r["appro"]),
                "usid": r["usid"] or "",
                "topicId": r["topic_id"] or "",
                "createDt": _fmt_epoch_ms(r["create_dt"]) or "-",
            }
        )

    total_pages = max(1, (total + page_size - 1) // page_size)
    return {
        "ok": True,
        "total": total,
        "page": page,
        "pageSize": page_size,
        "totalPages": total_pages,
        "counts": counts,
        "list": items,
    }


def api_headword_status(body: dict) -> dict:
    """게시 상태 전이 — use_yn·appro 두 필드만 갱신 (단건/일괄 지원).

    action:
      publish   게시(승인)      → use_yn=Y, appro=N
      unpublish 게시 중지        → use_yn=N, appro=N
      pending   승인대기로 되돌림 → use_yn=N, appro=Y
      reject    반려             → use_yn=N, appro=N (별도 반려 컬럼 없음 → 미게시로 수렴)
    """
    action = (body.get("action") or "").strip().lower()
    nos = body.get("headwordNos")
    if not isinstance(nos, list):
        nos = []
    single = str(body.get("headwordNo") or body.get("headword_no") or "").strip()
    if single:
        nos.append(single)
    nos = [str(n).strip() for n in nos if str(n).strip()]
    if not nos:
        return {"ok": False, "message": "대상 표제어(headwordNo)가 없습니다."}

    trans = {
        "publish": ("Y", "N"),
        "unpublish": ("N", "N"),
        "pending": ("N", "Y"),
        "reject": ("N", "N"),
    }
    if action not in trans:
        return {"ok": False, "message": f"알 수 없는 action: {action}"}
    use_yn, appro = trans[action]

    with db_connect() as con:
        ph = ",".join(["?"] * len(nos))
        cur = con.execute(
            f"UPDATE kd_headword SET use_yn=?, appro=? WHERE headword_no IN ({ph})",
            [use_yn, appro] + nos,
        )
        con.commit()
        updated = cur.rowcount

    return {
        "ok": True,
        "message": "상태가 변경되었습니다.",
        "action": action,
        "updated": updated,
        "count": len(nos),
    }


def api_headword_detail(qs: dict) -> dict:
    headword_no = (qs.get("headwordNo") or qs.get("headword_no") or [""])[0].strip()
    headword_id = (qs.get("headwordId") or qs.get("headword_id") or [""])[0].strip()
    if not headword_no and not headword_id:
        return {"ok": False, "message": "headwordNo 또는 headwordId 가 필요합니다."}
    with db_connect() as con:
        if headword_no:
            row = con.execute(
                "SELECT * FROM kd_headword WHERE headword_no = ? LIMIT 1",
                (headword_no,),
            ).fetchone()
        else:
            row = con.execute(
                "SELECT * FROM kd_headword WHERE headword_id = ? LIMIT 1",
                (headword_id,),
            ).fetchone()
    if not row:
        return {"ok": False, "message": "표제어를 찾을 수 없습니다."}
    d = dict(row)
    return {
        "ok": True,
        "data": {
            "headwordId": d.get("headword_id") or "",
            "headwordNo": d.get("headword_no") or "",
            "topicId": d.get("topic_id") or "",
            "headword": d.get("headword") or "",
            "wordClass": d.get("word_class") or "",
            "meaning": d.get("meaning") or "",
            "commentary": d.get("commentary") or "",
            "usid": d.get("usid") or "",
            "useYn": d.get("use_yn") or "N",
            "appro": d.get("appro") or "N",
            "mapMake": d.get("map_make") or "",
            "createDt": _fmt_epoch_ms(d.get("create_dt")) or "-",
        },
    }


def api_headword_search_by_name(qs: dict) -> dict:
    """표제어 blur 시 동일 표제어 품사·의미 자동완성 (원본 headword_search)."""
    headword = (qs.get("headword") or [""])[0].strip()
    if not headword:
        return {"ok": True, "list": []}
    with db_connect() as con:
        rows = con.execute(
            """
            SELECT headword, word_class, meaning, commentary
            FROM kd_headword
            WHERE headword = ?
            ORDER BY CAST(headword_no AS INTEGER) DESC
            LIMIT 5
            """,
            (headword,),
        ).fetchall()
    return {
        "ok": True,
        "list": [
            {
                "headword": r["headword"] or "",
                "wordClass": r["word_class"] or "",
                "meaning": r["meaning"] or "",
                "commentary": r["commentary"] or "",
            }
            for r in rows
        ],
    }


def api_headword_save(body: dict) -> dict:
    """표제어 등록(C) / 수정(M) — 로컬 SQLite."""
    import time

    mode = (body.get("mode") or "C").strip().upper()
    headword = (body.get("headword") or "").strip()
    word_class = (body.get("wordClass") or body.get("word_class") or "").strip()
    meaning = (body.get("meaning") or "").strip()
    commentary = (body.get("commentary") or "").strip()
    use_yn = (body.get("useYn") or body.get("use_yn") or "N").strip() or "N"
    appro = (body.get("appro") or "N").strip() or "N"
    headword_no = str(body.get("headwordNo") or body.get("headword_no") or "").strip()
    headword_id = str(body.get("headwordId") or body.get("headword_id") or "").strip()

    def provided(*keys):
        return any((k in body and body[k] is not None) for k in keys)

    # 삭제(D): 표제어 + 하위 지역어·지역 일괄 삭제
    if mode == "D":
        if not headword_no and not headword_id:
            return {"ok": False, "message": "삭제 대상 headwordNo 가 없습니다."}
        with db_connect() as con:
            if not headword_no and headword_id:
                row = con.execute(
                    "SELECT headword_no FROM kd_headword WHERE headword_id=?", (headword_id,)
                ).fetchone()
                headword_no = row["headword_no"] if row else ""
            if headword_no:
                con.execute("DELETE FROM tb_headword_dialect_region WHERE headword_no=?", (headword_no,))
                con.execute("DELETE FROM tb_headword_dialect WHERE headword_no=?", (headword_no,))
                cur = con.execute("DELETE FROM kd_headword WHERE headword_no=?", (headword_no,))
            else:
                cur = con.execute("DELETE FROM kd_headword WHERE headword_id=?", (headword_id,))
            con.commit()
            if cur.rowcount == 0:
                return {"ok": False, "message": "삭제할 표제어를 찾지 못했습니다."}
        return {"ok": True, "message": "삭제되었습니다.", "mode": "D", "headwordNo": headword_no}

    if not headword:
        return {"ok": False, "message": "표제어를 입력하세요."}
    if len(headword) > 125:
        return {"ok": False, "message": "표제어는 125자 이내로 입력해 주세요."}
    if mode == "C":
        # 등록 시에만 필수 (해설/commentary 는 nullable → 선택)
        if not word_class:
            return {"ok": False, "message": "품사를 선택하세요."}
        if not meaning:
            return {"ok": False, "message": "의미를 입력하세요."}

    # 원본 로직: 서비스 반영(Y)이면 승인요청 N, 승인요청 Y면 서비스 N
    if use_yn == "Y":
        appro = "N"
    if appro == "Y":
        use_yn = "N"

    now_ms = str(int(time.time() * 1000))

    with db_connect() as con:
        if mode == "M":
            if not headword_no and not headword_id:
                return {"ok": False, "message": "수정 대상 headwordNo 가 없습니다."}
            # 부분 업데이트 — 전달된 컬럼만 갱신 (미전송 필드 보존)
            sets, vals = ["headword=?"], [headword]
            if provided("wordClass", "word_class"):
                sets.append("word_class=?"); vals.append(word_class)
            if provided("meaning"):
                sets.append("meaning=?"); vals.append(meaning)
            if provided("commentary"):
                sets.append("commentary=?"); vals.append(commentary)
            if provided("useYn", "use_yn", "appro"):
                sets.append("use_yn=?"); vals.append(use_yn)
                sets.append("appro=?"); vals.append(appro)
            where = "headword_no=?" if headword_no else "headword_id=?"
            vals.append(headword_no or headword_id)
            cur = con.execute(
                "UPDATE kd_headword SET " + ", ".join(sets) + " WHERE " + where, vals
            )
            if cur.rowcount == 0:
                return {"ok": False, "message": "수정할 표제어를 찾지 못했습니다."}
            con.commit()
            return {
                "ok": True,
                "message": "저장되었습니다.",
                "mode": "M",
                "headwordNo": headword_no,
            }

        # Create
        max_id = con.execute(
            "SELECT MAX(CAST(headword_id AS INTEGER)) FROM kd_headword"
        ).fetchone()[0]
        max_no = con.execute(
            "SELECT MAX(CAST(headword_no AS INTEGER)) FROM kd_headword"
        ).fetchone()[0]
        new_id = str(int(max_id or 0) + 1)
        new_no = str(int(max_no or 0) + 1)
        con.execute(
            """
            INSERT INTO kd_headword (
              headword_id, topic_id, headword_no, sub_no, use_no,
              headword, original_word, word_class, meaning, usid,
              use_yn, appro, map_make, commentary, create_dt
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                new_id,
                body.get("topicId") or "",
                new_no,
                body.get("subNo") or "0",
                body.get("useNo") or "",
                headword,
                body.get("originalWord") or "",
                word_class,
                meaning,
                body.get("usid") or "admin",
                use_yn,
                appro,
                body.get("mapMake") or "Y",
                commentary,
                now_ms,
            ),
        )
        con.commit()
        return {
            "ok": True,
            "message": "등록되었습니다.",
            "mode": "C",
            "headwordId": new_id,
            "headwordNo": new_no,
        }


def api_dialect_list(qs: dict) -> dict:
    """표제어별 지역어(어휘) 목록 + 지역 건수."""
    headword_no = (qs.get("headwordNo") or qs.get("headword_no") or [""])[0].strip()
    if not headword_no:
        return {"ok": False, "message": "headwordNo 가 필요합니다."}
    with db_connect() as con:
        rows = con.execute(
            """
            SELECT d.hd_id, d.headword_no, d.word, d.face_color, d.mutation_group,
                   d.mutation_seq, d.map_symbol_id, d.symbol_color, d.create_dt,
                   (SELECT COUNT(*) FROM tb_headword_dialect_region r
                     WHERE r.hd_id = d.hd_id) AS region_count
            FROM tb_headword_dialect d
            WHERE d.headword_no = ?
            ORDER BY
              CASE WHEN d.mutation_group IS NULL OR trim(d.mutation_group)='' THEN 9999
                   ELSE CAST(d.mutation_group AS INTEGER) END,
              CASE WHEN d.mutation_seq IS NULL OR trim(d.mutation_seq)='' THEN 9999
                   ELSE CAST(d.mutation_seq AS INTEGER) END,
              CAST(d.hd_id AS INTEGER)
            """,
            (headword_no,),
        ).fetchall()
        # group colors: first non-empty face_color per group
        items = []
        for r in rows:
            # full=1 이면 전체 지역, 아니면 미리보기 8개
            full = (qs.get("full") or ["0"])[0] in ("1", "true", "True", "yes")
            lim_sql = "" if full else " LIMIT 8"
            # 지도 렌더용 지역 해석 힌트를 kd_region_code 로 보강.
            #   MyMapRegions.search 는 "광주시" 같은 시·군명만 매칭됨(전체명 "경기도 광주시"는 실패).
            #   → mapKey(시·군명) + sido(중복 해소용) 를 함께 내려 클라이언트가 정확히 muni_ 로 해석.
            regs = con.execute(
                f"""
                SELECT r.hdr_id, r.region_nm, r.region_id,
                       (SELECT c.dialect_map_key FROM kd_region_code c WHERE c.region_id = r.region_id LIMIT 1) AS map_key,
                       (SELECT c.sigungu       FROM kd_region_code c WHERE c.region_id = r.region_id LIMIT 1) AS sgg,
                       (SELECT c.sido          FROM kd_region_code c WHERE c.region_id = r.region_id LIMIT 1) AS sido
                FROM tb_headword_dialect_region r
                WHERE r.hd_id = ?
                ORDER BY CAST(r.hdr_id AS INTEGER)
                {lim_sql}
                """,
                (r["hd_id"],),
            ).fetchall()

            def region_name(x):
                nm = (x["region_nm"] or "").strip()
                if nm:
                    return nm
                return (x["map_key"] or "").strip() or (x["sgg"] or "").strip()
            items.append(
                {
                    "hdId": r["hd_id"],
                    "headwordNo": r["headword_no"],
                    "word": r["word"] or "",
                    "faceColor": r["face_color"] or "",
                    "mutationGroup": r["mutation_group"] or "",
                    "mutationSeq": r["mutation_seq"] or "",
                    "mapSymbolId": r["map_symbol_id"] or "",
                    "symbolColor": r["symbol_color"] or "",
                    "regionCount": r["region_count"] or 0,
                    "regions": [
                        {
                            "hdrId": x["hdr_id"] or "",
                            "regionNm": region_name(x),
                            "regionId": x["region_id"] or "",
                            "mapKey": (x["sgg"] or x["map_key"] or "").strip(),
                            "sido": (x["sido"] or "").strip(),
                        }
                        for x in regs
                    ],
                }
            )
    return {"ok": True, "headwordNo": headword_no, "total": len(items), "list": items}


def api_dialect_save(body: dict) -> dict:
    """지역어 등록/수정/삭제."""
    import time

    mode = (body.get("mode") or "C").strip().upper()
    headword_no = str(body.get("headwordNo") or body.get("headword_no") or "").strip()
    hd_id = str(body.get("hdId") or body.get("hd_id") or "").strip()
    word = (body.get("word") or "").strip()
    group = str(body.get("mutationGroup") or body.get("mutation_group") or "").strip()
    seq = str(body.get("mutationSeq") or body.get("mutation_seq") or "").strip()
    map_symbol_id = str(body.get("mapSymbolId") or body.get("map_symbol_id") or "").strip()
    face_color = str(body.get("faceColor") or body.get("face_color") or "").strip()
    symbol_color = str(body.get("symbolColor") or body.get("symbol_color") or "").strip()

    with db_connect() as con:
        if mode == "D":
            if not hd_id:
                return {"ok": False, "message": "삭제 대상 hdId 가 없습니다."}
            con.execute("DELETE FROM tb_headword_dialect_region WHERE hd_id = ?", (hd_id,))
            cur = con.execute("DELETE FROM tb_headword_dialect WHERE hd_id = ?", (hd_id,))
            con.commit()
            if cur.rowcount == 0:
                return {"ok": False, "message": "삭제할 지역어를 찾지 못했습니다."}
            return {"ok": True, "message": "삭제되었습니다.", "mode": "D", "hdId": hd_id}

        if not headword_no:
            return {"ok": False, "message": "headwordNo 가 필요합니다."}
        if not word:
            return {"ok": False, "message": "지역어를 입력하세요."}

        # group sequential rule soft-check on create
        if mode == "C" and group:
            try:
                gnum = int(group)
            except ValueError:
                gnum = 0
            if gnum > 1:
                exists = con.execute(
                    """
                    SELECT 1 FROM tb_headword_dialect
                    WHERE headword_no=? AND mutation_group=?
                    LIMIT 1
                    """,
                    (headword_no, str(gnum - 1)),
                ).fetchone()
                if not exists:
                    # allow if any lower group exists
                    lower = con.execute(
                        """
                        SELECT 1 FROM tb_headword_dialect
                        WHERE headword_no=? AND mutation_group IS NOT NULL
                          AND trim(mutation_group)!=''
                          AND CAST(mutation_group AS INTEGER) < ?
                        LIMIT 1
                        """,
                        (headword_no, gnum),
                    ).fetchone()
                    if not lower and gnum > 1:
                        prev = con.execute(
                            """
                            SELECT 1 FROM tb_headword_dialect
                            WHERE headword_no=? AND mutation_group=?
                            LIMIT 1
                            """,
                            (headword_no, str(gnum - 1)),
                        ).fetchone()
                        if not prev:
                            return {
                                "ok": False,
                                "message": f"{gnum}그룹을 만들려면 먼저 {gnum-1}그룹이 있어야 합니다.",
                            }

        if mode == "M":
            if not hd_id:
                return {"ok": False, "message": "수정 대상 hdId 가 없습니다."}
            cur = con.execute(
                """
                UPDATE tb_headword_dialect
                SET word=?, mutation_group=?, mutation_seq=?, map_symbol_id=?,
                    face_color=?, symbol_color=?
                WHERE hd_id=?
                """,
                (word, group, seq, map_symbol_id, face_color, symbol_color, hd_id),
            )
            if cur.rowcount == 0:
                return {"ok": False, "message": "수정할 지역어를 찾지 못했습니다."}
            con.commit()
            return {"ok": True, "message": "저장되었습니다.", "mode": "M", "hdId": hd_id}

        # Create — auto seq at end of group
        if not seq:
            mx = con.execute(
                """
                SELECT MAX(CAST(mutation_seq AS INTEGER)) FROM tb_headword_dialect
                WHERE headword_no=? AND IFNULL(mutation_group,'')=?
                """,
                (headword_no, group),
            ).fetchone()[0]
            seq = str(int(mx or 0) + 1)
        if not group:
            group = "1"
        max_hd = con.execute(
            "SELECT MAX(CAST(hd_id AS INTEGER)) FROM tb_headword_dialect"
        ).fetchone()[0]
        new_hd = str(int(max_hd or 0) + 1)
        now_ms = str(int(time.time() * 1000))
        con.execute(
            """
            INSERT INTO tb_headword_dialect (
              hd_id, headword_no, word, face_color, mutation_group, mutation_seq,
              map_symbol_id, symbol_color, create_dt
            ) VALUES (?,?,?,?,?,?,?,?,?)
            """,
            (
                new_hd,
                headword_no,
                word,
                face_color,
                group,
                seq,
                map_symbol_id,
                symbol_color,
                now_ms,
            ),
        )
        con.commit()
        return {
            "ok": True,
            "message": "지역어가 추가되었습니다.",
            "mode": "C",
            "hdId": new_hd,
            "headwordNo": headword_no,
        }


def api_dialect_region_save(body: dict) -> dict:
    """지역 추가/삭제 (간단 텍스트 region_nm)."""
    import time

    mode = (body.get("mode") or "C").strip().upper()
    hd_id = str(body.get("hdId") or body.get("hd_id") or "").strip()
    headword_no = str(body.get("headwordNo") or body.get("headword_no") or "").strip()
    word = (body.get("word") or "").strip()
    region_nm = (body.get("regionNm") or body.get("region_nm") or "").strip()
    hdr_id = str(body.get("hdrId") or body.get("hdr_id") or "").strip()

    with db_connect() as con:
        if mode == "D":
            if not hdr_id:
                return {"ok": False, "message": "삭제 대상 hdrId 가 없습니다."}
            cur = con.execute(
                "DELETE FROM tb_headword_dialect_region WHERE hdr_id = ?", (hdr_id,)
            )
            con.commit()
            if cur.rowcount == 0:
                return {"ok": False, "message": "삭제할 지역을 찾지 못했습니다."}
            return {"ok": True, "message": "지역이 삭제되었습니다.", "mode": "D"}

        if not hd_id or not region_nm:
            return {"ok": False, "message": "hdId 와 지역명이 필요합니다."}
        drow = con.execute(
            "SELECT headword_no, word FROM tb_headword_dialect WHERE hd_id=?",
            (hd_id,),
        ).fetchone()
        if not drow:
            return {"ok": False, "message": "지역어를 찾을 수 없습니다."}
        if not headword_no:
            headword_no = drow["headword_no"]
        if not word:
            word = drow["word"] or ""
        max_hdr = con.execute(
            "SELECT MAX(CAST(hdr_id AS INTEGER)) FROM tb_headword_dialect_region"
        ).fetchone()[0]
        new_hdr = str(int(max_hdr or 0) + 1)
        now_ms = str(int(time.time() * 1000))
        region_id = str(body.get("regionId") or body.get("region_id") or "").strip()
        con.execute(
            """
            INSERT INTO tb_headword_dialect_region (
              hdr_id, headword_no, word, region_id, hd_id, serial_nm, basis_year, region_nm, create_dt
            ) VALUES (?,?,?,?,?,?,?,?,?)
            """,
            (new_hdr, headword_no, word, region_id, hd_id, "", "", region_nm, now_ms),
        )
        con.commit()
        return {
            "ok": True,
            "message": "지역이 추가되었습니다.",
            "mode": "C",
            "hdrId": new_hdr,
            "hdId": hd_id,
        }


# ── 사용자(회원) 관리 : pt_user (로컬 SQLite 미러) ──────────────────────────
_USERGROUP_NM = {"1": "시스템관리자", "2": "정보제공관리자", "3": "일반사용자"}
# auth: 9정상 / 8정지(로그인 5회 실패) / 5탈퇴 / 0미인증
# 탈퇴는 물리삭제 대신 상태 전환. 탈퇴 일시는 pt_user.secsn_dt 에 별도 보관한다
# (update_dt 는 비밀번호 재설정 등 다른 수정에도 갱신되어 탈퇴 시점 근거로 쓸 수 없음).
_AUTH_WITHDRAWN = "5"
_AUTH_LABEL = {"9": "정상", "8": "정지", "5": "탈퇴", "0": "미인증", "4": "기타"}


def _fmt_user_dt(v) -> str:
    """write_dt/update_dt(에폭 ms 문자열 또는 datetime 문자열) → YYYY-MM-DD."""
    if v is None or str(v).strip() == "":
        return ""
    s = str(v).strip()
    if s.isdigit():  # 에폭 ms
        try:
            return datetime.fromtimestamp(int(s) / 1000).strftime("%Y-%m-%d")
        except Exception:
            return ""
    return s[:10]  # 'YYYY-MM-DD ...' 형태


def _parse_user_dt(v, keep=None):
    """'YYYY-MM-DD' / 'YYYY.MM.DD' → 에폭 ms 문자열.
    빈 값이면 keep(기존 값)을 그대로 돌려준다. 형식이 어긋나면 ValueError."""
    s = str(v or "").strip()
    if not s:
        return keep
    if s.isdigit():          # 이미 에폭 ms
        return s
    t = s.replace(".", "-").replace("/", "-")[:10]
    try:
        d = datetime.strptime(t, "%Y-%m-%d")
    except ValueError:
        raise ValueError("날짜는 YYYY-MM-DD 형식으로 입력해 주세요. (%s)" % s)
    return str(int(d.timestamp() * 1000))


def api_user_list(qs: dict) -> dict:
    """회원 목록 조회 — pt_user(로컬). 검색/페이징 지원."""
    def _q(k, d=""):
        v = qs.get(k)
        return (v[0] if isinstance(v, list) else v) or d

    page = max(1, int(_q("page", "1") or 1))
    size = int(_q("size", "10") or 10)
    if size not in (10, 20, 50, 100):
        size = 10
    usergroup_id = str(_q("usergroupId")).strip()
    se = str(_q("se", "1")).strip()          # 1계정 2이름 3이메일 4부서명
    search = str(_q("search")).strip()
    start_dt = str(_q("startDt")).strip()    # YYYY.MM.DD / YYYY-MM-DD
    end_dt = str(_q("endDt")).strip()
    status = str(_q("status")).strip()        # 상태: 9정상 8정지 0미인증 ('' 전체)
    only_locked = str(_q("onlyLocked")).strip() in ("1", "true", "Y")

    where, params = ["1=1"], []
    if usergroup_id:
        where.append("usergroup_id = ?"); params.append(usergroup_id)
    if status in ("9", "8", "5", "0", "4"):
        where.append("auth = ?"); params.append(status)
    elif only_locked:
        where.append("auth = '8'")
    if search:
        col = {"1": "usid", "2": "username"}.get(se)   # 이메일·부서명 검색은 화면에서 제거됨
        if col:
            where.append(f"{col} LIKE ?"); params.append(f"%{search}%")
    # 등록일 범위(write_dt = 에폭 ms 문자열) 필터
    def _to_ms(d, end=False):
        d = d.replace(".", "-").strip()
        if not d:
            return None
        try:
            dt = datetime.strptime(d[:10], "%Y-%m-%d")
            if end:
                dt = dt.replace(hour=23, minute=59, second=59)
            return int(dt.timestamp() * 1000)
        except Exception:
            return None
    ms_s, ms_e = _to_ms(start_dt), _to_ms(end_dt, end=True)
    # 에폭 ms 저장분에 한해 범위 적용 (숫자형만)
    if ms_s is not None:
        where.append("(CASE WHEN write_dt GLOB '[0-9]*' THEN CAST(write_dt AS INTEGER) ELSE NULL END) >= ?")
        params.append(ms_s)
    if ms_e is not None:
        where.append("(CASE WHEN write_dt GLOB '[0-9]*' THEN CAST(write_dt AS INTEGER) ELSE NULL END) <= ?")
        params.append(ms_e)

    # 탈퇴 회원은 기본 제외. 관리자가 「탈퇴 회원 포함」을 켜면 함께 조회한다.
    include_secsn = (qs.get("includeSecsn") or ["0"])[0] in ("1", "true", "Y", "y")
    if not include_secsn and status != _AUTH_WITHDRAWN:
        where.append(f"IFNULL(auth,'') <> '{_AUTH_WITHDRAWN}'")

    wh = " AND ".join(where)
    con = db_connect()
    try:
        total = con.execute(f"SELECT COUNT(*) FROM pt_user u WHERE {wh}", params).fetchone()[0]
        offset = (page - 1) * size
        # usid 기준 등록 건수 집계
        #   나의 지도   : kd_headword (지역어 지도 표제어, map/dialect.do 목록과 동일 기준)
        #   의견 제시   : tb_board_post board_id='qna'
        rows = con.execute(
            f"""SELECT u.user_id, u.usergroup_id, u.usid, u.username, u.dept_nm,
                       u.auth, u.fail_count, u.write_dt, u.secsn_dt,
                       (SELECT COUNT(*) FROM tb_board_post p
                         WHERE p.board_id = 'qna'
                           AND IFNULL(p.use_yn,'Y') <> 'N'
                           AND p.create_id = u.usid) AS opinion_cnt,
                       (SELECT COUNT(*) FROM kd_headword h
                         WHERE h.usid = u.usid) AS mymap_cnt
                FROM pt_user u WHERE {wh}
                ORDER BY CAST(u.user_id AS INTEGER) DESC
                LIMIT ? OFFSET ?""",
            params + [size, offset],
        ).fetchall()
    finally:
        con.close()

    out = []
    for r in rows:
        auth = str(r["auth"] or "")
        out.append({
            "userId": str(r["user_id"] or ""),
            "usergroupId": str(r["usergroup_id"] or ""),
            "groupName": _USERGROUP_NM.get(str(r["usergroup_id"] or ""), str(r["usergroup_id"] or "")),
            "usid": r["usid"] or "",
            "username": r["username"] or "",
            "deptNm": r["dept_nm"] or "",
            "opinionCnt": int(r["opinion_cnt"] or 0),
            "myMapCnt": int(r["mymap_cnt"] or 0),
            "secsnDt": _fmt_user_dt(r["secsn_dt"]),
            "auth": auth,
            "authLabel": _AUTH_LABEL.get(auth, auth),
            "failCount": str(r["fail_count"] or "0"),
            "writeDt": _fmt_user_dt(r["write_dt"]),
        })
    return {
        "ok": True, "total": total, "page": page, "size": size,
        "totalPages": max(1, (total + size - 1) // size), "rows": out,
    }



def api_board_post_list(qs: dict) -> dict:
    """게시판 글 목록 — tb_board_post. 의견 제시(qna) 등 board_id 별 조회.
    createId 를 주면 해당 계정이 작성한 글만 (사용자 계정관리 → 「의견 제시」 건수 클릭)."""
    board_id = (qs.get("boardId") or ["qna"])[0].strip() or "qna"
    create_id = (qs.get("createId") or qs.get("writer") or [""])[0].strip()
    search = (qs.get("search") or qs.get("searchKeyword") or [""])[0].strip()
    # 검색 구분 — '' 전체 / 1 제목 / 2 내용 / 0 제목+내용 / 3 작성자 아이디 / 4 작성자 이름
    cond = (qs.get("searchCondition") or [""])[0].strip()
    try:
        page = max(1, int((qs.get("page") or ["1"])[0]))
    except ValueError:
        page = 1
    try:
        size = int((qs.get("pageSize") or ["10"])[0])
    except ValueError:
        size = 10
    size = size if size in (10, 20, 50, 100) else 10

    # use_yn='N' 은 삭제된 글(소프트 삭제). 기본 제외, includeDeleted=1 이면 함께 조회.
    include_deleted = (qs.get("includeDeleted") or ["0"])[0] in ("1", "true", "Y", "y")
    where = ["p.board_id = ?"]
    params: list = [board_id]
    if not include_deleted:
        where.append("IFNULL(p.use_yn,'Y') <> 'N'")
    if create_id:
        where.append("p.create_id = ?")
        params.append(create_id)
    if search:
        like = f"%{search}%"
        if cond == "1":
            where.append("IFNULL(p.post_title,'') LIKE ?"); params.append(like)
        elif cond == "2":
            where.append("IFNULL(p.post_content,'') LIKE ?"); params.append(like)
        elif cond == "3":
            # 작성자 아이디는 계정 정확일치 (관리자 화면의 건수 링크와 결과가 일치해야 함)
            where.append("p.create_id = ?"); params.append(search)
        elif cond == "4":
            # 작성자 이름은 동명이인이 있어 부분일치로 찾는다
            where.append(
                "EXISTS (SELECT 1 FROM pt_user u WHERE u.usid = p.create_id AND IFNULL(u.username,'') LIKE ?)"
            )
            params.append(like)
        else:
            where.append(
                "(IFNULL(p.post_title,'') LIKE ? OR IFNULL(p.post_content,'') LIKE ?"
                " OR p.create_id = ?"
                " OR EXISTS (SELECT 1 FROM pt_user u WHERE u.usid = p.create_id AND IFNULL(u.username,'') LIKE ?))"
            )
            params += [like, like, search, like]
    # 탈퇴 회원이 쓴 글도 동일하게 기본 숨김
    if (qs.get("includeSecsn") or ["0"])[0] not in ("1", "true", "Y", "y"):
        where.append(
            f"NOT EXISTS (SELECT 1 FROM pt_user u WHERE u.usid = p.create_id AND IFNULL(u.auth,'') = '{_AUTH_WITHDRAWN}')"
        )
    wh = " AND ".join(where)

    con = db_connect()
    try:
        total = con.execute(f"SELECT COUNT(*) FROM tb_board_post p WHERE {wh}", params).fetchone()[0]
        rows = con.execute(
            f"""SELECT p.post_id, p.post_title, p.view_count, p.public_yn, p.use_yn,
                       p.fix_yn, p.category_code, p.create_id, p.create_dt,
                       (SELECT u.username FROM pt_user u WHERE u.usid = p.create_id LIMIT 1) AS username,
                       (SELECT COUNT(*) FROM tb_board_answer a
                         WHERE a.post_id = p.post_id AND IFNULL(a.use_yn,'Y') <> 'N') AS answer_cnt,
                       (SELECT COUNT(*) FROM tb_board_file f
                         WHERE f.post_id = p.post_id AND IFNULL(f.use_yn,'Y') <> 'N') AS file_cnt
                FROM tb_board_post p WHERE {wh}
                ORDER BY CAST(p.post_id AS INTEGER) DESC
                LIMIT ? OFFSET ?""",
            params + [size, (page - 1) * size],
        ).fetchall()
    finally:
        con.close()

    cats = board_categories(board_id)
    out = []
    for r in rows:
        cat = (r["category_code"] or "").strip()
        out.append({
            "postId": str(r["post_id"] or ""),
            "title": r["post_title"] or "",
            "categoryCode": cat,
            "categoryName": cats.get(cat, cat),
            "createId": r["create_id"] or "",
            "username": r["username"] or "",
            "createDt": _fmt_user_dt(r["create_dt"]),
            "viewCount": str(r["view_count"] or "0"),
            "publicYn": r["public_yn"] or "",
            "fixYn": (r["fix_yn"] or "N"),
            "fileCount": int(r["file_cnt"] or 0),
            "deleted": str(r["use_yn"] or "Y") == "N",
            "answered": int(r["answer_cnt"] or 0) > 0,
        })
    return {"ok": True, "total": total, "page": page, "size": size,
            "totalPages": max(1, (total + size - 1) // size), "rows": out}



def board_categories(board_id: str = "qna") -> dict:
    """게시판 분류 코드 → 명칭. tb_board.category_group 으로 tb_code_detail 을 조회한다.
    (명칭을 코드에 박아 두면 DB 와 어긋나므로 항상 DB 값을 쓴다)"""
    con = db_connect()
    try:
        grp = con.execute(
            "SELECT category_group FROM tb_board WHERE board_id = ?", (board_id,)
        ).fetchone()
        if not grp or not (grp["category_group"] or "").strip():
            return {}
        rows = con.execute(
            """SELECT code_code, code_name FROM tb_code_detail
                WHERE group_code = ?
                  AND IFNULL(use_yn,'Y') <> 'N' AND IFNULL(delete_yn,'N') <> 'Y'
                ORDER BY CAST(IFNULL(code_sort_seq,'9999') AS INTEGER)""",
            (grp["category_group"],),
        ).fetchall()
    finally:
        con.close()
    return {(r["code_code"] or "").strip(): (r["code_name"] or "").strip() for r in rows}


def api_board_post_detail(qs: dict) -> dict:
    """의견 제시 상세 — 글 + 답변 1건(있으면)."""
    post_id = (qs.get("postId") or [""])[0].strip()
    if not post_id:
        return {"ok": False, "message": "postId가 필요합니다."}
    con = db_connect()
    try:
        p = con.execute(
            """SELECT p.*,
                      (SELECT u.username FROM pt_user u WHERE u.usid = p.create_id LIMIT 1) AS username,
                      (SELECT u.username FROM pt_user u WHERE u.usid = p.update_id LIMIT 1) AS update_username
               FROM tb_board_post p WHERE p.post_id = ?""", (post_id,)).fetchone()
        if not p:
            return {"ok": False, "message": "해당 글을 찾을 수 없습니다."}
        a = con.execute(
            """SELECT a.*,
                      (SELECT u.username FROM pt_user u WHERE u.usid = a.create_id LIMIT 1) AS username,
                      (SELECT u.username FROM pt_user u WHERE u.usid = a.update_id LIMIT 1) AS update_username
               FROM tb_board_answer a
               WHERE a.post_id = ? AND IFNULL(a.use_yn,'Y') <> 'N'
               ORDER BY CAST(a.answer_id AS INTEGER) DESC LIMIT 1""", (post_id,)).fetchone()
    finally:
        con.close()

    cats = board_categories(p["board_id"] or "qna")
    cat = (p["category_code"] or "").strip()
    post = {
        "postId": str(p["post_id"] or ""),
        "boardId": p["board_id"] or "",
        "title": p["post_title"] or "",
        "content": p["post_content"] or "",
        "categoryCode": cat,
        "categoryName": cats.get(cat, cat),
        "viewCount": str(p["view_count"] or "0"),
        "publicYn": p["public_yn"] or "",
        "useYn": p["use_yn"] or "",
        "deleted": str(p["use_yn"] or "Y") == "N",
        "createId": p["create_id"] or "",
        "username": p["username"] or "",
        "createDt": _fmt_user_dt(p["create_dt"]),
        "updateId": p["update_id"] or "",
        "updateName": p["update_username"] or "",
        "updateDt": _fmt_user_dt(p["update_dt"]),
    }
    answer = None
    if a:
        answer = {
            "answerId": str(a["answer_id"] or ""),
            "content": a["answer_content"] or "",
            # 작성자/등록일시 · 수정자/수정일시를 각각 따로 내려 준다
            "createId": a["create_id"] or "",
            "username": a["username"] or "",
            "createDt": _fmt_user_dt(a["create_dt"]),
            "updateId": a["update_id"] or "",
            "updateName": a["update_username"] or "",
            "updateDt": _fmt_user_dt(a["update_dt"]),
        }
    return {"ok": True, "post": post, "answer": answer,
            "answered": answer is not None,
            "categories": cats}


def api_board_post_update(body: dict) -> dict:
    """질문(원글) 수정 — 제목·내용·분류·게시여부."""
    post_id = str(body.get("postId") or "").strip()
    if not post_id:
        return {"ok": False, "message": "postId가 필요합니다."}
    title = str(body.get("title") or "").strip()
    if not title:
        return {"ok": False, "message": "제목을 입력해 주세요."}
    content = str(body.get("content") or "")
    cat = str(body.get("categoryCode") or "").strip()
    public_yn = "Y" if str(body.get("publicYn") or "Y") == "Y" else "N"
    use_yn = "Y" if str(body.get("useYn") or "Y") == "Y" else "N"
    editor = str(body.get("editorId") or "admin").strip() or "admin"

    con = db_connect()
    try:
        row = con.execute(
            "SELECT create_dt, update_dt FROM tb_board_post WHERE post_id = ?", (post_id,)
        ).fetchone()
        if not row:
            return {"ok": False, "message": "해당 글을 찾을 수 없습니다."}
        # 등록일시·수정일시는 관리자가 직접 고칠 수 있다.
        # 값을 보내지 않으면 등록일시는 유지, 수정일시는 지금 시각으로 찍는다.
        try:
            create_dt = _parse_user_dt(body.get("createDt"), keep=row["create_dt"])
            update_dt = _parse_user_dt(body.get("updateDt"),
                                       keep=str(int(time.time() * 1000)))
        except ValueError as e:
            return {"ok": False, "message": str(e)}
        con.execute(
            """UPDATE tb_board_post
                  SET post_title = ?, post_content = ?, category_code = ?,
                      public_yn = ?, use_yn = ?, create_dt = ?,
                      update_id = ?, update_dt = ?
                WHERE post_id = ?""",
            (title, content, cat, public_yn, use_yn, create_dt,
             editor, update_dt, post_id),
        )
        con.commit()
    finally:
        con.close()
    return {"ok": True, "message": "질문을 저장했습니다."}


def api_board_answer_save(body: dict) -> dict:
    """답변 등록·수정. 기존 답변이 있으면 수정, 없으면 새로 등록한다."""
    post_id = str(body.get("postId") or "").strip()
    if not post_id:
        return {"ok": False, "message": "postId가 필요합니다."}
    content = str(body.get("content") or "").strip()
    if not content:
        return {"ok": False, "message": "답변 내용을 입력해 주세요."}
    # 작성자/수정자를 각각 받는다. writerId 는 신규 등록 시 create_id 로 들어간다.
    editor = str(body.get("editorId") or "admin").strip() or "admin"
    writer = str(body.get("writerId") or "").strip()
    now_ms = str(int(time.time() * 1000))

    con = db_connect()
    try:
        if not con.execute("SELECT 1 FROM tb_board_post WHERE post_id = ?", (post_id,)).fetchone():
            return {"ok": False, "message": "해당 글을 찾을 수 없습니다."}
        row = con.execute(
            """SELECT answer_id, create_id, create_dt FROM tb_board_answer
                WHERE post_id = ? AND IFNULL(use_yn,'Y') <> 'N'
                ORDER BY CAST(answer_id AS INTEGER) DESC LIMIT 1""", (post_id,)).fetchone()
        # 등록일시·수정일시는 관리자가 직접 고칠 수 있다.
        try:
            create_dt = _parse_user_dt(body.get("createDt"),
                                       keep=(row["create_dt"] if row else now_ms))
            update_dt = _parse_user_dt(body.get("updateDt"), keep=now_ms)
        except ValueError as e:
            return {"ok": False, "message": str(e)}
        if row:
            con.execute(
                """UPDATE tb_board_answer
                      SET answer_content = ?, create_id = ?, create_dt = ?,
                          update_id = ?, update_dt = ?
                    WHERE answer_id = ?""",
                (content, writer or row["create_id"], create_dt,
                 editor, update_dt, row["answer_id"]),
            )
            msg = "답변을 수정했습니다."
        else:
            nxt = con.execute(
                "SELECT IFNULL(MAX(CAST(answer_id AS INTEGER)), 0) + 1 FROM tb_board_answer"
            ).fetchone()[0]
            con.execute(
                """INSERT INTO tb_board_answer
                       (answer_id, post_id, answer_content, use_yn, create_id, create_dt)
                   VALUES (?, ?, ?, 'Y', ?, ?)""",
                (str(nxt), post_id, content, writer or editor, create_dt),
            )
            msg = "답변을 등록했습니다."
        con.commit()
    finally:
        con.close()
    return {"ok": True, "message": msg}


def api_board_post_delete(body: dict) -> dict:
    """글 삭제·복원 — 물리삭제 없이 use_yn 만 전환한다(게시물 복원 관리와 동일 방식)."""
    post_id = str(body.get("postId") or "").strip()
    restore = str(body.get("restore") or "") in ("1", "true", "Y", "y")
    if not post_id:
        return {"ok": False, "message": "postId가 필요합니다."}
    editor = str(body.get("editorId") or "admin").strip() or "admin"
    con = db_connect()
    try:
        row = con.execute("SELECT use_yn FROM tb_board_post WHERE post_id = ?", (post_id,)).fetchone()
        if not row:
            return {"ok": False, "message": "해당 글을 찾을 수 없습니다."}
        cur = str(row["use_yn"] or "Y")
        if restore and cur != "N":
            return {"ok": False, "message": "삭제된 글이 아닙니다."}
        if not restore and cur == "N":
            return {"ok": False, "message": "이미 삭제된 글입니다."}
        con.execute(
            "UPDATE tb_board_post SET use_yn = ?, update_id = ?, update_dt = ? WHERE post_id = ?",
            ("Y" if restore else "N", editor, str(int(time.time() * 1000)), post_id),
        )
        con.commit()
    finally:
        con.close()
    return {"ok": True, "message": "글을 복원했습니다." if restore else "글을 삭제했습니다."}

def api_user_set_auth(body: dict) -> dict:
    """승인상태 변경 — 관리자가 수정 화면에서 직접 지정한다.
    탈퇴('5')는 물리삭제 대신 상태 전환이며 탈퇴 일시를 기록하고,
    다른 상태로 되돌리면 탈퇴 일시를 지운다(보존기한 무기한 정책)."""
    user_id = str(body.get("userId") or "").strip()
    auth = str(body.get("auth") or "").strip()
    if not user_id:
        return {"ok": False, "message": "userId가 필요합니다."}
    if auth not in _AUTH_LABEL:
        return {"ok": False, "message": f"허용되지 않은 상태값입니다: {auth}"}
    con = db_connect()
    try:
        row = con.execute("SELECT usid, auth FROM pt_user WHERE user_id = ?", (user_id,)).fetchone()
        if not row:
            return {"ok": False, "message": "해당 회원을 찾을 수 없습니다."}
        before = str(row["auth"] or "")
        if before == auth:
            return {"ok": False, "message": f"이미 「{_AUTH_LABEL.get(auth, auth)}」 상태입니다."}
        now_ms = str(int(time.time() * 1000))
        if auth == _AUTH_WITHDRAWN:
            con.execute(
                "UPDATE pt_user SET auth = ?, secsn_dt = ?, fail_count = '0' WHERE user_id = ?",
                (auth, now_ms, user_id),
            )
            msg = f"{row['usid']} 계정을 탈퇴 상태로 변경했습니다. 등록 자료는 사용자단에서 숨겨집니다."
        else:
            # 정지 해제 시 실패 횟수도 함께 초기화
            con.execute(
                "UPDATE pt_user SET auth = ?, secsn_dt = NULL, fail_count = '0' WHERE user_id = ?",
                (auth, user_id),
            )
            msg = f"{row['usid']} 계정의 승인상태를 「{_AUTH_LABEL.get(auth, auth)}」(으)로 변경했습니다."
        con.commit()
    finally:
        con.close()
    return {"ok": True, "message": msg}


def api_user_detail(qs: dict) -> dict:
    """회원 단건 조회 — pt_user(로컬)."""
    def _q(k, d=""):
        v = qs.get(k)
        return (v[0] if isinstance(v, list) else v) or d
    user_id = str(_q("userId")).strip()
    if not user_id:
        return {"ok": False, "message": "userId가 필요합니다."}
    con = db_connect()
    try:
        r = con.execute(
            """SELECT u.user_id, u.usergroup_id, u.usid, u.password, u.username, u.dept_nm,
                      u.mobile, u.auth, u.fail_count, u.writer, u.write_dt, u.updater, u.update_dt,
                      u.secsn_dt,
                      (SELECT COUNT(*) FROM kd_headword h
                        WHERE h.usid = u.usid) AS mymap_cnt,
                      (SELECT COUNT(*) FROM tb_board_post p
                        WHERE p.board_id = 'qna'
                          AND IFNULL(p.use_yn,'Y') <> 'N'
                          AND p.create_id = u.usid) AS opinion_cnt
               FROM pt_user u WHERE u.user_id = ?""", (user_id,)
        ).fetchone()
    finally:
        con.close()
    if not r:
        return {"ok": False, "message": "해당 회원을 찾을 수 없습니다."}
    auth = str(r["auth"] or "")
    return {
        "ok": True,
        "user": {
            "userId": str(r["user_id"] or ""),
            "usergroupId": str(r["usergroup_id"] or ""),
            "groupName": _USERGROUP_NM.get(str(r["usergroup_id"] or ""), str(r["usergroup_id"] or "")),
            "usid": r["usid"] or "",
            "hasPassword": bool((r["password"] or "").strip()),
            "username": r["username"] or "",
            "deptNm": r["dept_nm"] or "",
            "myMapCnt": int(r["mymap_cnt"] or 0),
            "opinionCnt": int(r["opinion_cnt"] or 0),
            "mobile": r["mobile"] or "",
            "auth": auth,
            "authLabel": _AUTH_LABEL.get(auth, auth),
            "failCount": str(r["fail_count"] or "0"),
            "secsnDt": _fmt_user_dt(r["secsn_dt"]),
            "writer": r["writer"] or "",
            "writeDt": _fmt_user_dt(r["write_dt"]),
            "updater": r["updater"] or "",
            "updateDt": _fmt_user_dt(r["update_dt"]),
        },
    }


def api_user_unlock(body: dict) -> dict:
    """정지(잠금) 계정 해제 — auth=9, fail_count=0."""
    user_id = str(body.get("userId") or "").strip()
    if not user_id:
        return {"ok": False, "message": "userId가 필요합니다."}
    con = db_connect()
    try:
        row = con.execute(
            "SELECT usid, auth, fail_count FROM pt_user WHERE user_id = ?", (user_id,)
        ).fetchone()
        if not row:
            return {"ok": False, "message": "해당 회원을 찾을 수 없습니다."}
        con.execute(
            "UPDATE pt_user SET auth = '9', fail_count = '0' WHERE user_id = ?", (user_id,)
        )
        con.commit()
    finally:
        con.close()
    return {
        "ok": True, "message": f"'{row['usid']}' 계정 잠금을 해제했습니다.",
        "userId": user_id, "auth": "9", "authLabel": "정상", "failCount": "0",
    }


def api_user_reset_pw(body: dict) -> dict:
    """관리자 임의 비밀번호 재설정 — bcrypt($2a$10$) 해시로 password UPDATE."""
    user_id = str(body.get("userId") or "").strip()
    new_pw = str(body.get("newPassword") or "")
    if not user_id:
        return {"ok": False, "message": "userId가 필요합니다."}
    if len(new_pw) < 8:
        return {"ok": False, "message": "비밀번호는 8자 이상이어야 합니다."}
    try:
        import bcrypt
    except Exception:
        return {"ok": False, "message": "서버에 bcrypt 모듈이 없습니다. (pip install bcrypt)"}
    con = db_connect()
    try:
        row = con.execute(
            "SELECT usid FROM pt_user WHERE user_id = ?", (user_id,)
        ).fetchone()
        if not row:
            return {"ok": False, "message": "해당 회원을 찾을 수 없습니다."}
        hashed = bcrypt.hashpw(new_pw.encode("utf-8"),
                               bcrypt.gensalt(rounds=10, prefix=b"2a")).decode("ascii")
        con.execute("UPDATE pt_user SET password = ? WHERE user_id = ?", (hashed, user_id))
        con.commit()
    finally:
        con.close()
    return {"ok": True, "message": f"'{row['usid']}' 비밀번호를 재설정했습니다.", "userId": user_id}


# ── Open API 사용현황 : pt_user.api_key / api_url / api_purpose / api_dt ──

def _ensure_api_purpose_col(con) -> None:
    cols = {str(r[1]) for r in con.execute("PRAGMA table_info(pt_user)").fetchall()}
    if "api_purpose" not in cols:
        con.execute("ALTER TABLE pt_user ADD COLUMN api_purpose TEXT")
        con.commit()


def _fmt_api_dt(v) -> str:
    s = str(v or "").strip()
    if not s:
        return ""
    if s.isdigit():
        try:
            return datetime.fromtimestamp(int(s) / 1000).strftime("%Y.%m.%d %H:%M")
        except Exception:
            return ""
    return s[:16]


def _api_dt_range_ms(d: str, end=False):
    d = (d or "").replace(".", "-").strip()
    if not d:
        return None
    try:
        dt = datetime.strptime(d[:10], "%Y-%m-%d")
        if end:
            dt = dt.replace(hour=23, minute=59, second=59)
        return int(dt.timestamp() * 1000)
    except Exception:
        return None


def api_openapi_usage_list(qs: dict) -> dict:
    """Open API 인증키 발급 현황 — 활용목적 포함."""
    def q1(k, d=""):
        return (qs.get(k) or [d])[0].strip()
    try:
        page = max(1, int(q1("page", "1")))
    except ValueError:
        page = 1
    try:
        page_size = int(q1("pageSize", "10") or q1("size", "10"))
    except ValueError:
        page_size = 10
    if page_size not in (10, 20, 50, 100):
        page_size = 10
    cond = q1("searchCondition")
    keyword = q1("searchKeyword") or q1("q")
    start_ms = _api_dt_range_ms(q1("searchStartDt") or q1("startDt"))
    end_ms = _api_dt_range_ms(q1("searchEndDt") or q1("endDt"), end=True)

    where = ["IFNULL(u.api_key,'') <> ''"]
    params = []
    if start_ms is not None:
        where.append("(CASE WHEN u.api_dt GLOB '[0-9]*' THEN CAST(u.api_dt AS INTEGER) ELSE NULL END) >= ?")
        params.append(start_ms)
    if end_ms is not None:
        where.append("(CASE WHEN u.api_dt GLOB '[0-9]*' THEN CAST(u.api_dt AS INTEGER) ELSE NULL END) <= ?")
        params.append(end_ms)
    if keyword:
        colmap = {
            "usid": "u.usid",
            "username": "u.username",
            "api_key": "u.api_key",
            "api_url": "u.api_url",
            "api_purpose": "u.api_purpose",
        }
        if cond in colmap:
            where.append(f"IFNULL({colmap[cond]},'') LIKE ?")
            params.append("%" + keyword + "%")
        else:
            where.append(
                "(IFNULL(u.usid,'') LIKE ? OR IFNULL(u.username,'') LIKE ? "
                "OR IFNULL(u.api_key,'') LIKE ? OR IFNULL(u.api_url,'') LIKE ? "
                "OR IFNULL(u.api_purpose,'') LIKE ?)"
            )
            params.extend(["%" + keyword + "%"] * 5)
    wh = " AND ".join(where)
    with db_connect() as con:
        _ensure_api_purpose_col(con)
        total = con.execute(f"SELECT COUNT(*) FROM pt_user u WHERE {wh}", params).fetchone()[0]
        total_pages = max(1, (total + page_size - 1) // page_size)
        if page > total_pages:
            page = total_pages
        offset = (page - 1) * page_size
        rows = con.execute(
            f"""SELECT u.user_id, u.usid, u.username, u.api_key, u.api_url,
                       u.api_purpose, u.api_dt
                FROM pt_user u WHERE {wh}
                ORDER BY CAST(u.api_dt AS INTEGER) DESC, CAST(u.user_id AS INTEGER) DESC
                LIMIT ? OFFSET ?""",
            params + [page_size, offset],
        ).fetchall()
    items = []
    for r in rows:
        items.append({
            "userId": str(r["user_id"] or ""),
            "usid": r["usid"] or "",
            "username": r["username"] or "",
            "apiKey": r["api_key"] or "",
            "apiUrl": r["api_url"] or "",
            "apiPurpose": r["api_purpose"] or "",
            "apiDt": _fmt_api_dt(r["api_dt"]),
            "callTotal": 0,
            "callToday": 0,
            "callThisMonth": 0,
            "callLastMonth": 0,
        })
    return {"ok": True, "total": total, "page": page, "pageSize": page_size,
            "totalPages": total_pages, "list": items}


def build_openapi_excel(qs: dict):
    """Open API 사용현황 엑셀 — 활용목적 포함."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    def q1(k, d=""):
        return (qs.get(k) or [d])[0].strip()
    items = []
    cond = q1("searchCondition")
    keyword = q1("searchKeyword") or q1("q")
    start_ms = _api_dt_range_ms(q1("searchStartDt") or q1("startDt"))
    end_ms = _api_dt_range_ms(q1("searchEndDt") or q1("endDt"), end=True)
    where = ["IFNULL(u.api_key,'') <> ''"]
    params = []
    if start_ms is not None:
        where.append("(CASE WHEN u.api_dt GLOB '[0-9]*' THEN CAST(u.api_dt AS INTEGER) ELSE NULL END) >= ?")
        params.append(start_ms)
    if end_ms is not None:
        where.append("(CASE WHEN u.api_dt GLOB '[0-9]*' THEN CAST(u.api_dt AS INTEGER) ELSE NULL END) <= ?")
        params.append(end_ms)
    if keyword:
        colmap = {
            "usid": "u.usid", "username": "u.username", "api_key": "u.api_key",
            "api_url": "u.api_url", "api_purpose": "u.api_purpose",
        }
        if cond in colmap:
            where.append(f"IFNULL({colmap[cond]},'') LIKE ?")
            params.append("%" + keyword + "%")
        else:
            where.append(
                "(IFNULL(u.usid,'') LIKE ? OR IFNULL(u.username,'') LIKE ? "
                "OR IFNULL(u.api_key,'') LIKE ? OR IFNULL(u.api_url,'') LIKE ? "
                "OR IFNULL(u.api_purpose,'') LIKE ?)"
            )
            params.extend(["%" + keyword + "%"] * 5)
    wh = " AND ".join(where)
    with db_connect() as con:
        _ensure_api_purpose_col(con)
        rows = con.execute(
            f"""SELECT u.usid, u.username, u.api_key, u.api_url, u.api_purpose, u.api_dt
                FROM pt_user u WHERE {wh}
                ORDER BY CAST(u.api_dt AS INTEGER) DESC""",
            params,
        ).fetchall()
        for r in rows:
            items.append((
                r["usid"] or "", r["username"] or "", r["api_key"] or "",
                r["api_url"] or "", r["api_purpose"] or "", _fmt_api_dt(r["api_dt"]),
            ))

    wb = Workbook()
    ws = wb.active
    ws.title = "Open API 사용현황"
    headers = ["번호", "아이디", "이름", "API Key", "사용 URL", "활용목적", "발급일시"]
    ws.append(headers)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2563EB")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
    for i, row in enumerate(items, start=1):
        ws.append([i, *row])
    for col, w in zip("ABCDEFG", (8, 16, 14, 34, 40, 40, 18)):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), "OpenAPI_사용현황.xlsx"


# ── 문학 속 지역어 (tb_literature + tb_literature_example) ──

def _lit_q1(qs: dict, key: str, default: str = "") -> str:
    v = qs.get(key, [default])
    if isinstance(v, list):
        return (v[0] if v else default) or default
    return str(v or default)


def _lit_word_classes(qs: dict) -> list[str]:
    """품사 검색값 — 단일/콤마구분/반복 파라미터를 모두 허용.
    복합 품사는 가운뎃점(·)을 쓰므로 콤마는 선택값 구분자로만 쓴다."""
    raw: list[str] = []
    for key in ("wordClass", "word_class"):
        v = qs.get(key, [])
        if isinstance(v, list):
            raw.extend(str(x) for x in v)
        elif v:
            raw.append(str(v))
    out: list[str] = []
    seen: set[str] = set()
    for item in raw:
        for part in item.split(","):
            p = part.strip()
            if p and p not in seen:
                seen.add(p)
                out.append(p)
    return out


def _lit_pos_variants(label: str) -> tuple[list[str], bool]:
    """검색 품사 라벨 → DB word_class 표기 변형. (variants, include_empty)."""
    s = (label or "").strip()
    if not s:
        return [], False
    include_empty = s == "품사 없음"
    parts = [p.strip() for p in re.split(r"[·,/]", s) if p.strip()]
    out: set[str] = {s, s.replace(" ", "")}
    if include_empty:
        out.update(["품사 없음", "없음"])
    if s == "접사":
        out.update(["접사", "접미사", "접두사"])
    if s == "어미":
        out.update(["어미", "종결어미"])
    if parts:
        n = len(parts)
        perms: list[tuple[str, ...]] = [tuple(parts)]
        if n == 2:
            perms.append((parts[1], parts[0]))
        elif n == 3:
            a, b, c = parts
            perms.extend([(a, c, b), (b, a, c), (b, c, a), (c, a, b), (c, b, a)])
        for perm in perms:
            compact = tuple(p.replace(" ", "") for p in perm)
            for seq in (perm, compact):
                out.add("·".join(seq))
                out.add(",".join(seq))
                out.add(", ".join(seq))
                out.add(" ".join(seq))
    variants = [v for v in out if v]
    return variants, include_empty


def _lit_match_sql(col: str, mode: str) -> str:
    m = (mode or "contains").lower()
    if m in ("match", "exact", "eq", "일치"):
        return f"{col} = ?"
    if m in ("startswith", "prefix", "start", "시작문자"):
        return f"{col} LIKE ?"
    if m in ("endswith", "suffix", "end", "끝문자"):
        return f"{col} LIKE ?"
    return f"{col} LIKE ?"


def _lit_match_val(term: str, mode: str) -> str:
    m = (mode or "contains").lower()
    if m in ("match", "exact", "eq", "일치"):
        return term
    if m in ("startswith", "prefix", "start", "시작문자"):
        return f"{term}%"
    if m in ("endswith", "suffix", "end", "끝문자"):
        return f"%{term}"
    return f"%{term}%"


def _lit_field_col(field: str) -> str | None:
    f = (field or "").strip().lower()
    return {
        "dlt": "l.dlt_tp",
        "dlt_tp": "l.dlt_tp",
        "std": "l.std_tp",
        "std_tp": "l.std_tp",
        "mean": "l.mean",
        "writer": "e.writer",
        "book": "e.book_name",
        "book_name": "e.book_name",
        "example": "e.word_example",
        "word_example": "e.word_example",
        "rel_dlt": "l.rel_dlt",
        "region": "l.region_nm",
    }.get(f)


def api_literature_list(qs: dict) -> dict:
    """관리자 문학 속 지역어 목록 (dialect_local.db)."""
    try:
        page = max(1, int(_lit_q1(qs, "page", "1") or 1))
    except ValueError:
        page = 1
    try:
        size = min(100, max(1, int(_lit_q1(qs, "size", "10") or 10)))
    except ValueError:
        size = 10
    region = _lit_q1(qs, "region").strip()
    word_classes = _lit_word_classes(qs)
    use_yn = _lit_q1(qs, "useYn").strip().upper()
    main_fix = _lit_q1(qs, "mainFixYn").strip().upper()
    q = _lit_q1(qs, "q").strip()

    # targets: JSON array [{field, mode, term, conn}]  or t1_field/t1_mode/t1_term ...
    targets = []
    raw_targets = _lit_q1(qs, "targets").strip()
    if raw_targets:
        try:
            parsed = json.loads(raw_targets)
            if isinstance(parsed, list):
                targets = parsed
        except Exception:
            targets = []
    if not targets:
        for i in range(1, 4):
            term = _lit_q1(qs, f"t{i}_term").strip()
            if not term:
                continue
            targets.append({
                "field": _lit_q1(qs, f"t{i}_field", "dlt"),
                "mode": _lit_q1(qs, f"t{i}_mode", "contains"),
                "term": term,
                "conn": _lit_q1(qs, f"t{i}_conn", "in"),
            })
    if q and not targets:
        targets = [{"field": "dlt", "mode": "contains", "term": q, "conn": "in"}]

    where = ["1=1"]
    params: list = []

    if region:
        where.append("l.region_nm LIKE ?")
        params.append(f"%{region}%")
    if word_classes:
        ors = []
        for wc in word_classes:
            variants, include_empty = _lit_pos_variants(wc)
            if variants:
                ph = ",".join("?" * len(variants))
                if include_empty:
                    ors.append(f"(IFNULL(l.word_class,'') = '' OR IFNULL(l.word_class,'') IN ({ph}))")
                else:
                    ors.append(f"IFNULL(l.word_class,'') IN ({ph})")
                params.extend(variants)
            elif include_empty:
                ors.append("IFNULL(l.word_class,'') = ''")
        if ors:
            where.append("(" + " OR ".join(ors) + ")")
    if use_yn in ("Y", "N"):
        where.append("UPPER(COALESCE(l.use_yn,'')) = ?")
        params.append(use_yn)
    # 메인고정 컬럼 없음 → 목록 표시는 항상 N. Y 필터 시 0건
    if main_fix == "Y":
        where.append("1=0")

    need_example_join = False
    for t in targets:
        field = str(t.get("field") or "dlt")
        col = _lit_field_col(field)
        if not col:
            continue
        if col.startswith("e."):
            need_example_join = True
        term = str(t.get("term") or "").strip()
        if not term:
            continue
        mode = str(t.get("mode") or "contains")
        conn = str(t.get("conn") or "in").lower()
        clause = _lit_match_sql(col, mode)
        val = _lit_match_val(term, mode)
        if conn in ("notin", "not", "제외", "and_not"):
            where.append(f"NOT ({clause})")
        else:
            where.append(clause)
        params.append(val)

    where_sql = " AND ".join(where)
    join_sql = (
        "LEFT JOIN tb_literature_example e ON l.liter_id = e.liter_id"
        if need_example_join or True
        else ""
    )
    # always left join so writer/book can be selected for list display
    join_sql = "LEFT JOIN tb_literature_example e ON l.liter_id = e.liter_id"

    offset = (page - 1) * size
    with db_connect() as con:
        total = con.execute(
            f"""
            SELECT COUNT(DISTINCT l.liter_id) AS c
            FROM tb_literature l
            {join_sql}
            WHERE {where_sql}
            """,
            params,
        ).fetchone()["c"]

        rows = con.execute(
            f"""
            SELECT l.liter_id, l.dlt_tp, l.std_tp, l.word_class, l.mean,
                   l.region_nm, l.use_yn, l.rel_dlt,
                   (SELECT e2.writer FROM tb_literature_example e2
                     WHERE e2.liter_id = l.liter_id
                     ORDER BY CAST(e2.liter_exam_id AS INTEGER) ASC LIMIT 1) AS writer,
                   (SELECT e2.book_name FROM tb_literature_example e2
                     WHERE e2.liter_id = l.liter_id
                     ORDER BY CAST(e2.liter_exam_id AS INTEGER) ASC LIMIT 1) AS book_name
            FROM tb_literature l
            {join_sql}
            WHERE {where_sql}
            GROUP BY l.liter_id
            ORDER BY CAST(l.liter_id AS INTEGER) ASC
            LIMIT ? OFFSET ?
            """,
            params + [size, offset],
        ).fetchall()

    items = []
    for r in rows:
        items.append({
            "literId": str(r["liter_id"] or ""),
            "dltTp": r["dlt_tp"] or "",
            "stdTp": r["std_tp"] or "",
            "wordClass": r["word_class"] or "",
            "mean": r["mean"] or "",
            "regionNm": r["region_nm"] or "",
            "useYn": (r["use_yn"] or "N").upper() if (r["use_yn"] or "").strip() else "N",
            "mainFixYn": "N",
            "writer": (r["writer"] or "").strip(),
            "bookName": (r["book_name"] or "").strip(),
            "relDlt": r["rel_dlt"] or "",
        })

    total_pages = max(1, (int(total) + size - 1) // size) if total else 1
    return {
        "ok": True,
        "total": int(total),
        "page": page,
        "size": size,
        "totalPages": total_pages,
        "rows": items,
        "db": str(DB_PATH),
    }


def api_literature_detail(qs: dict) -> dict:
    liter_id = _lit_q1(qs, "id") or _lit_q1(qs, "literId") or _lit_q1(qs, "liter_id")
    liter_id = liter_id.strip()
    if not liter_id:
        return {"ok": False, "message": "id(liter_id)가 필요합니다."}

    with db_connect() as con:
        row = con.execute(
            "SELECT * FROM tb_literature WHERE liter_id = ? LIMIT 1",
            (liter_id,),
        ).fetchone()
        if not row:
            return {"ok": False, "message": f"자료를 찾을 수 없습니다. (id={liter_id})"}

        ex_rows = con.execute(
            """
            SELECT * FROM tb_literature_example
            WHERE liter_id = ?
            ORDER BY CAST(liter_exam_id AS INTEGER) ASC
            """,
            (liter_id,),
        ).fetchall()

    examples = []
    for ex in ex_rows:
        examples.append({
            "literExamId": str(ex["liter_exam_id"] or ""),
            "wordExample": ex["word_example"] or "",
            "stdExample": ex["std_example"] or "",
            "writer": ex["writer"] or "",
            "bookName": ex["book_name"] or "",
            "publishCompany": ex["publish_company"] or "",
            "publishYear": ex["publish_year"] or "",
            "volumeNo": ex["volume_no"] or "",
            "pageNo": ex["page_no"] or "",
            "sidoCd": ex["sido_cd"] or "",
            "sigunguCd": ex["sigungu_cd"] or "",
            "sidoNm": ex["sido_nm"] or "",
            "sigunguNm": ex["sigungu_nm"] or "",
            "useYn": ex["use_yn"] or "Y",
        })

    data = {
        "literId": str(row["liter_id"] or ""),
        "dltTp": row["dlt_tp"] or "",
        "stdTp": row["std_tp"] or "",
        "addMean": row["add_mean"] or "",
        "wordClass": row["word_class"] or "",
        "mean": row["mean"] or "",
        "regionNm": row["region_nm"] or "",
        "relDlt": row["rel_dlt"] or "",
        "wordDesc": row["word_desc"] or "",
        "useYn": (row["use_yn"] or "N").upper() if (row["use_yn"] or "").strip() else "N",
        "exhBookNm": row["exh_book_nm"] or "",
        "exhAuthor": row["exh_author"] or "",
        "exhPublishCom": row["exh_publish_com"] or "",
        "exhPublishYear": row["exh_publish_year"] or "",
        "examples": examples,
    }
    return {"ok": True, "data": data}


def api_literature_save(body: dict) -> dict:
    """문학 지역어 등록/수정 + 용례 교체 저장."""
    liter_id = str(body.get("literId") or body.get("liter_id") or "").strip()
    dlt_tp = str(body.get("dltTp") or body.get("dlt_tp") or "").strip()
    std_tp = str(body.get("stdTp") or body.get("std_tp") or "").strip()
    if not dlt_tp:
        return {"ok": False, "message": "표제어를 입력해주세요."}
    if not std_tp:
        return {"ok": False, "message": "대응표준어를 입력해주세요."}

    add_mean = str(body.get("addMean") or body.get("add_mean") or "").strip()
    word_class = str(body.get("wordClass") or body.get("word_class") or "").strip()
    mean = str(body.get("mean") or "").strip()
    region_nm = body.get("regionNm") or body.get("region_nm") or ""
    if isinstance(region_nm, list):
        region_nm = ", ".join([str(x).strip() for x in region_nm if str(x).strip()])
    else:
        region_nm = str(region_nm or "").strip()
    rel_dlt = str(body.get("relDlt") or body.get("rel_dlt") or "").strip()
    word_desc = str(body.get("wordDesc") or body.get("word_desc") or "").strip()
    use_yn = str(body.get("useYn") or body.get("use_yn") or "Y").strip().upper() or "Y"
    if use_yn not in ("Y", "N"):
        use_yn = "Y"
    exh_book = str(body.get("exhBookNm") or body.get("exh_book_nm") or "").strip()
    exh_author = str(body.get("exhAuthor") or body.get("exh_author") or "").strip()
    exh_com = str(body.get("exhPublishCom") or body.get("exh_publish_com") or "").strip()
    exh_year = str(body.get("exhPublishYear") or body.get("exh_publish_year") or "").strip()
    examples = body.get("examples") or []
    if not isinstance(examples, list):
        examples = []

    now_ms = str(int(time.time() * 1000))

    with db_connect() as con:
        if liter_id:
            exists = con.execute(
                "SELECT liter_id FROM tb_literature WHERE liter_id = ?", (liter_id,)
            ).fetchone()
            if not exists:
                return {"ok": False, "message": f"자료를 찾을 수 없습니다. (id={liter_id})"}
            con.execute(
                """
                UPDATE tb_literature SET
                  dlt_tp=?, std_tp=?, add_mean=?, word_class=?, mean=?,
                  region_nm=?, rel_dlt=?, word_desc=?, use_yn=?,
                  exh_book_nm=?, exh_author=?, exh_publish_com=?, exh_publish_year=?,
                  upt_id=?, upt_dt=?
                WHERE liter_id=?
                """,
                (
                    dlt_tp, std_tp, add_mean, word_class, mean,
                    region_nm, rel_dlt, word_desc, use_yn,
                    exh_book, exh_author, exh_com, exh_year,
                    "0", now_ms, liter_id,
                ),
            )
            con.execute("DELETE FROM tb_literature_example WHERE liter_id = ?", (liter_id,))
        else:
            row = con.execute(
                "SELECT COALESCE(MAX(CAST(liter_id AS INTEGER)), 0) + 1 AS n FROM tb_literature"
            ).fetchone()
            liter_id = str(row["n"])
            con.execute(
                """
                INSERT INTO tb_literature (
                  liter_id, dlt_tp, std_tp, add_mean, word_class, mean,
                  region_nm, rel_dlt, word_desc, use_yn, cause,
                  reg_id, reg_dt, upt_id, upt_dt,
                  exh_book_nm, exh_author, exh_publish_com, exh_publish_year
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    liter_id, dlt_tp, std_tp, add_mean, word_class, mean,
                    region_nm, rel_dlt, word_desc, use_yn, "",
                    "0", now_ms, "0", now_ms,
                    exh_book, exh_author, exh_com, exh_year,
                ),
            )

        # next exam id
        max_ex = con.execute(
            "SELECT COALESCE(MAX(CAST(liter_exam_id AS INTEGER)), 0) AS n FROM tb_literature_example"
        ).fetchone()["n"]
        next_ex = int(max_ex or 0)

        for ex in examples:
            if not isinstance(ex, dict):
                continue
            word_ex = str(ex.get("wordExample") or ex.get("word_example") or "").strip()
            if not word_ex:
                continue
            next_ex += 1
            con.execute(
                """
                INSERT INTO tb_literature_example (
                  liter_exam_id, liter_id, word_example, std_example,
                  writer, book_name, publish_company, publish_year,
                  sido_cd, sigungu_cd, sido_nm, sigungu_nm,
                  use_yn, reg_id, reg_dt, upt_id, upt_dt, page_no, volume_no
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    str(next_ex), liter_id, word_ex,
                    str(ex.get("stdExample") or ex.get("std_example") or "").strip(),
                    str(ex.get("writer") or "").strip(),
                    str(ex.get("bookName") or ex.get("book_name") or "").strip(),
                    str(ex.get("publishCompany") or ex.get("publish_company") or "").strip(),
                    str(ex.get("publishYear") or ex.get("publish_year") or "").strip(),
                    str(ex.get("sidoCd") or ex.get("sido_cd") or "").strip(),
                    str(ex.get("sigunguCd") or ex.get("sigungu_cd") or "").strip(),
                    str(ex.get("sidoNm") or ex.get("sido_nm") or "").strip(),
                    str(ex.get("sigunguNm") or ex.get("sigungu_nm") or "").strip(),
                    "Y", "0", now_ms, "0", now_ms,
                    str(ex.get("pageNo") or ex.get("page_no") or "").strip(),
                    str(ex.get("volumeNo") or ex.get("volume_no") or "").strip(),
                ),
            )
        con.commit()

    return {"ok": True, "message": "저장되었습니다.", "literId": liter_id}


def api_literature_delete(body: dict) -> dict:
    """문학 지역어 단건/다건 삭제 (용례 포함)."""
    ids = body.get("ids") or body.get("literIds") or []
    if not ids:
        one = body.get("literId") or body.get("id") or body.get("liter_id")
        if one:
            ids = [one]
    ids = [str(x).strip() for x in ids if str(x).strip()]
    if not ids:
        return {"ok": False, "message": "삭제할 항목을 선택해주세요."}

    with db_connect() as con:
        placeholders = ",".join("?" * len(ids))
        con.execute(
            f"DELETE FROM tb_literature_example WHERE liter_id IN ({placeholders})",
            ids,
        )
        cur = con.execute(
            f"DELETE FROM tb_literature WHERE liter_id IN ({placeholders})",
            ids,
        )
        con.commit()
        deleted = cur.rowcount if cur.rowcount is not None else len(ids)

    return {"ok": True, "message": f"{deleted}건 삭제되었습니다.", "deleted": deleted}


# ────────────────────────────────────────────────────────────────────────────
# 세대별 지역어 변화(단어 카드) 관리 — data/processed/word_stories.json 직접 편집
#   · 프론트(dialect_wordcard.html)가 읽는 그 파일이 곧 원본이다. DB를 쓰지 않는다.
#   · 저장할 때 알 수 없는 필드(callTable·related·lineage 등)는 그대로 보존한다.
#   · 쓰기는 임시파일 → os.replace 로 원자적으로, 직전 내용은 .bak 으로 남긴다.
# ────────────────────────────────────────────────────────────────────────────
WORDCARD_JSON = USER_MAP_ROOT / "data" / "processed" / "word_stories.json"

WC_GROUPS = ["20M", "20F", "50M", "50F", "70M", "70F"]


def _wc_load() -> dict:
    if not WORDCARD_JSON.is_file():
        raise FileNotFoundError(f"단어 카드 자료를 찾을 수 없습니다: {WORDCARD_JSON}")
    with WORDCARD_JSON.open("r", encoding="utf-8") as f:
        db = json.load(f)
    db.setdefault("words", [])
    db.setdefault("types", {})
    db.setdefault("coding", [])
    db.setdefault("meta", {})
    return db


def _wc_write(db: dict) -> None:
    text = json.dumps(db, ensure_ascii=False, indent=1) + "\n"
    if WORDCARD_JSON.is_file():
        try:
            bak = WORDCARD_JSON.with_suffix(WORDCARD_JSON.suffix + ".bak")
            bak.write_bytes(WORDCARD_JSON.read_bytes())
        except Exception:
            pass  # 백업 실패로 저장 자체를 막지는 않는다
    tmp = WORDCARD_JSON.with_name(WORDCARD_JSON.name + ".tmp")
    tmp.write_text(text, encoding="utf-8")
    os.replace(str(tmp), str(WORDCARD_JSON))


def _wc_ct_total(word: dict) -> int:
    ct = word.get("ct") or {}
    n = 0
    for g in WC_GROUPS:
        for v in (ct.get(g) or []):
            try:
                n += int(v)
            except Exception:
                pass
    return n


def _wc_row(word: dict, types: dict) -> dict:
    """목록 한 줄 — 무거운 교차표는 합계만 보낸다."""
    t = types.get(word.get("type")) or {}
    return {
        "id": str(word.get("id") or ""),
        "word": word.get("word") or "",
        "cat": word.get("cat") or "",
        "type": word.get("type") or "",
        "typeLabel": t.get("label") or word.get("type") or "",
        "typeColor": t.get("color") or "#64748b",
        "typeBg": t.get("bg") or "#f1f5f9",
        "hook": word.get("hook") or "",
        "page": word.get("page"),
        "section": word.get("section") or "",
        "hasCT": bool(word.get("hasCT")),
        "ctTotal": _wc_ct_total(word),
        "factCnt": len(word.get("facts") or []),
        "variantCnt": len(word.get("variants") or []),
    }


def api_wordcard_meta(qs: dict) -> dict:
    """유형·코딩 범주·기존 분류(cat) 목록 — 등록/수정 폼의 선택지."""
    db = _wc_load()
    cats = sorted({(w.get("cat") or "").strip() for w in db["words"] if (w.get("cat") or "").strip()})
    return {
        "ok": True,
        "types": db["types"],
        "coding": db["coding"],
        "cats": cats,
        "groups": WC_GROUPS,
        "meta": db["meta"],
        "path": str(WORDCARD_JSON),
        "total": len(db["words"]),
    }


def api_wordcard_list(qs: dict) -> dict:
    def q1(k, d=""):
        v = qs.get(k)
        return (v[0] if isinstance(v, list) else v) or d

    kw = str(q1("searchValue")).strip()
    wtype = str(q1("searchType")).strip()
    expose = str(q1("searchExpose")).strip()  # Y: 교차표 있음(프론트 노출), N: 없음
    try:
        page = max(1, int(q1("page", "1")))
    except Exception:
        page = 1
    try:
        page_size = int(q1("pageSize", "10"))
    except Exception:
        page_size = 10
    page_size = min(max(page_size, 1), 500)

    db = _wc_load()
    rows = [_wc_row(w, db["types"]) for w in db["words"]]

    if kw:
        def hit(r):
            hay = " ".join([r["id"], r["word"], r["cat"], r["hook"]])
            return kw in hay
        rows = [r for r in rows if hit(r)]
    if wtype:
        rows = [r for r in rows if r["type"] == wtype]
    if expose == "Y":
        rows = [r for r in rows if r["hasCT"]]
    elif expose == "N":
        rows = [r for r in rows if not r["hasCT"]]

    total = len(rows)
    # 앞단에 실제로 뜨는 건수 — 목록 총건수(파일 전체)와 헷갈리지 않게 함께 보낸다
    exposed = sum(1 for r in rows if r["hasCT"])
    total_pages = max(1, (total + page_size - 1) // page_size)
    page = min(page, total_pages)
    start = (page - 1) * page_size
    return {
        "ok": True,
        "total": total,
        "exposed": exposed,
        "hidden": total - exposed,
        "page": page,
        "pageSize": page_size,
        "totalPages": total_pages,
        "list": rows[start:start + page_size],
    }


def api_wordcard_detail(qs: dict) -> dict:
    def q1(k, d=""):
        v = qs.get(k)
        return (v[0] if isinstance(v, list) else v) or d

    wid = str(q1("id")).strip()
    if not wid:
        return {"ok": False, "message": "항목 ID가 없습니다."}
    db = _wc_load()
    for w in db["words"]:
        if str(w.get("id")) == wid:
            return {"ok": True, "item": w, "types": db["types"], "coding": db["coding"]}
    return {"ok": False, "message": f"항목을 찾을 수 없습니다: {wid}"}


def _wc_clean_ct(raw) -> dict:
    """교차표 정리 — 6집단 × 4범주 정수. 전부 0인 집단은 버린다."""
    out = {}
    if not isinstance(raw, dict):
        return out
    for g in WC_GROUPS:
        cells = raw.get(g)
        if not isinstance(cells, list):
            continue
        vals = []
        for i in range(4):
            try:
                vals.append(max(0, int(cells[i])))
            except Exception:
                vals.append(0)
        if sum(vals) > 0:
            out[g] = vals
    return out


def api_wordcard_save(body: dict) -> dict:
    mode = str(body.get("mode") or "").upper()
    wid = str(body.get("id") or "").strip()
    old_id = str(body.get("oldId") or "").strip()
    word = str(body.get("word") or "").strip()

    if not wid:
        return {"ok": False, "message": "항목번호(ID)를 입력해주세요."}
    if not word:
        return {"ok": False, "message": "단어를 입력해주세요."}

    db = _wc_load()
    words = db["words"]
    if body.get("type") and body["type"] not in db["types"]:
        return {"ok": False, "message": f"없는 변화 유형입니다: {body['type']}"}

    idx = {str(w.get("id")): i for i, w in enumerate(words)}
    if mode == "M":
        target = old_id or wid
        if target not in idx:
            return {"ok": False, "message": f"수정할 항목을 찾을 수 없습니다: {target}"}
        if wid != target and wid in idx:
            return {"ok": False, "message": f"이미 쓰고 있는 항목번호입니다: {wid}"}
        base = dict(words[idx[target]])
        pos = idx[target]
    else:
        if wid in idx:
            return {"ok": False, "message": f"이미 쓰고 있는 항목번호입니다: {wid}"}
        base = {}
        pos = None

    facts = [str(s).strip() for s in (body.get("facts") or []) if str(s).strip()]
    variants = []
    for v in (body.get("variants") or []):
        form = str((v or {}).get("form") or "").strip()
        if not form:
            continue
        item = {"form": form}
        if str(v.get("tag") or "").strip():
            item["tag"] = str(v["tag"]).strip()
        if str(v.get("note") or "").strip():
            item["note"] = str(v["note"]).strip()
        regions = [str(r).strip() for r in (v.get("regions") or []) if str(r).strip()]
        if regions:
            item["regions"] = regions
        variants.append(item)

    ct = _wc_clean_ct(body.get("ct"))

    base["id"] = wid
    base["word"] = word
    base["cat"] = str(body.get("cat") or "").strip()
    base["type"] = str(body.get("type") or "qualitative").strip()
    base["hook"] = str(body.get("hook") or "").strip()
    base["story"] = str(body.get("story") or "").strip()
    base["facts"] = facts
    base["section"] = str(body.get("section") or "4.1").strip()
    try:
        base["page"] = int(body.get("page"))
    except Exception:
        base["page"] = base.get("page") or 0

    if variants:
        base["variants"] = variants
    else:
        base.pop("variants", None)

    stats_in = body.get("stats") or {}
    chi = str(stats_in.get("chiSq") or "").strip()
    if chi:
        try:
            stats = {"chiSq": float(chi)}
        except Exception:
            return {"ok": False, "message": "카이제곱 값은 숫자로 입력해주세요."}
        try:
            stats["df"] = int(str(stats_in.get("df") or "").strip())
        except Exception:
            pass
        if str(stats_in.get("p") or "").strip():
            stats["p"] = str(stats_in["p"]).strip()
        base["stats"] = stats
    else:
        base.pop("stats", None)

    # 교차표가 있는 항목만 프론트 차트에 노출된다 → hasCT 는 입력값이 아니라 결과다
    if ct:
        base["ct"] = ct
        base["hasCT"] = True
    else:
        base.pop("ct", None)
        base["hasCT"] = False

    link = str(body.get("link") or "").strip()
    if link:
        base["link"] = link
    else:
        base.pop("link", None)

    if pos is None:
        words.append(base)
    else:
        words[pos] = base

    _wc_write(db)
    return {
        "ok": True,
        "id": wid,
        "hasCT": base["hasCT"],
        "message": ("등록되었습니다." if pos is None else "수정되었습니다.")
        + ("" if base["hasCT"] else " 교차표가 없어 앞단 화면에는 노출되지 않습니다."),
    }


def api_wordcard_delete(body: dict) -> dict:
    ids = body.get("ids") or []
    if not ids:
        one = body.get("id")
        if one:
            ids = [one]
    ids = [str(x).strip() for x in ids if str(x).strip()]
    if not ids:
        return {"ok": False, "message": "삭제할 항목을 선택해주세요."}

    db = _wc_load()
    before = len(db["words"])
    db["words"] = [w for w in db["words"] if str(w.get("id")) not in set(ids)]
    deleted = before - len(db["words"])
    if not deleted:
        return {"ok": False, "message": "삭제할 항목을 찾지 못했습니다."}
    _wc_write(db)
    return {"ok": True, "deleted": deleted, "message": f"{deleted}건 삭제되었습니다."}


# ────────────────────────────────────────────────────────────────────────────
# 지역어 기상도 — 전용 테이블 (wb_weather_*)
#   · 원자료 파일 1개 = 제보자 1명. 지역·연차·세대·성별은 파일명에서 읽는다.
#       {지역2}{연차2}{세대2}{성별1}VE.xlsx  예) CB2420FVE = 충북·2024·20대·여
#   · 첫 시트가 정본이다. 12개 파일이 항목범위별 분할 시트를 함께 갖고 있고
#     그중 GB2450MVE 는 저장 당시 마지막 시트가 활성이라 '활성 시트'를 쓰면 517행이 유실된다.
#   · 열은 위치가 아니라 헤더 이름으로 찾는다. 원본 서식이 15가지로 흔들려서다.
#     정리 양식(5열)과 옛 원본(14열·열순서 뒤바뀜)을 모두 받는다.
#   · 판정 임계값은 scripts/etl_awareness_region.py 와 같아야 한다(w1 ≥.6 · w2 ≥.3 · w3 ≥.1).
#
# 운영에서는 이 표들이 CUBRID 같은 스키마에 들어간다. 로컬 프로토타입에서는
# 운영 미러(dialect_local.db)를 건드리지 않으려고 별도 파일에 둔다.
# ────────────────────────────────────────────────────────────────────────────
WEATHER_DB = Path(
    os.environ.get("WEATHER_DB", str(USER_MAP_ROOT / "data" / "gisangdo.db"))
)

WB_REGION_NAMES = {"GG": "경기", "GW": "강원", "CB": "충북", "CN": "충남", "JB": "전북",
                   "JN": "전남", "GB": "경북", "GN": "경남", "JJ": "제주"}
WB_REGION_ORDER = ["GG", "GW", "CB", "CN", "JB", "JN", "GB", "GN", "JJ"]
WB_VALID_GRADE = {"1", "2", "3", "4"}
# 전각 숫자('１')만 반각으로 맞춘다. 판정 규칙은 etl_awareness_region.grade_of 와 같아야 한다.
WB_FULLWIDTH = str.maketrans("１２３４", "1234")

WEATHER_DDL = """
CREATE TABLE IF NOT EXISTS wb_weather_file (
  weather_file_id INTEGER PRIMARY KEY AUTOINCREMENT, file_nm TEXT NOT NULL UNIQUE,
  region_cd TEXT NOT NULL, region_nm TEXT, research_year INTEGER,
  research_degree TEXT, generation INTEGER, sex TEXT,
  row_cnt INTEGER DEFAULT 0, item_cnt INTEGER DEFAULT 0, src_layout TEXT,
  use_yn TEXT DEFAULT 'Y', reg_id TEXT, reg_dt TEXT, upt_id TEXT, upt_dt TEXT);
CREATE INDEX IF NOT EXISTS ix_wwf_region ON wb_weather_file (region_cd, generation, sex);

CREATE TABLE IF NOT EXISTS wb_weather_response (
  response_id INTEGER PRIMARY KEY AUTOINCREMENT, weather_file_id INTEGER NOT NULL,
  line_no INTEGER NOT NULL, serial_no TEXT, item_cd TEXT NOT NULL, item_base TEXT,
  headword TEXT, dialect_form TEXT, grade TEXT, grade_valid_yn TEXT DEFAULT 'N',
  upt_dt TEXT,          -- 관리자가 고친 행. 재업로드 경고와 캐시 무효화에 쓴다
  use_yn TEXT DEFAULT 'Y', reg_dt TEXT,
  FOREIGN KEY (weather_file_id) REFERENCES wb_weather_file (weather_file_id));
CREATE INDEX IF NOT EXISTS ix_wwr_file ON wb_weather_response (weather_file_id, line_no);
CREATE INDEX IF NOT EXISTS ix_wwr_item ON wb_weather_response (item_base, grade_valid_yn);
CREATE INDEX IF NOT EXISTS ix_wwr_head ON wb_weather_response (headword);

"""


def weather_db():
    WEATHER_DB.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(str(WEATHER_DB))
    con.row_factory = sqlite3.Row
    con.executescript(WEATHER_DDL)
    # 이미 만들어진 DB 에 칸이 없으면 더한다 (이력표를 없애며 생긴 칸)
    cols = [r[1] for r in con.execute("PRAGMA table_info(wb_weather_response)")]
    if "upt_dt" not in cols:
        con.execute("ALTER TABLE wb_weather_response ADD COLUMN upt_dt TEXT")
        con.commit()
    return con

def _wb_norm_header(h) -> str:
    """표기 흔들림을 표준 이름으로 모은다."""
    s = re.sub(r"\s+", "", str(h or ""))
    if s.startswith("시작시간"):
        return "시작시간"
    if s.startswith("종료시간"):
        return "종료시간"
    if s.startswith("지속시간"):
        return "지속시간"
    alias = {"표제어": "표제어형", "표제어형": "표제어형", "표준어형": "표제어형",
             "인지도/사용도": "사용도/인지도", "사용도/인지도": "사용도/인지도"}
    return alias.get(s, s)


def _wb_parse_filename(fname: str):
    m = re.match(r"^([A-Z]{2})(\d{2})(\d{2})([MF])VE", os.path.basename(fname))
    if not m:
        return None
    rg, yy, gen, sx = m.groups()
    return {"region_cd": rg, "region_nm": WB_REGION_NAMES.get(rg, rg),
            "research_year": 2000 + int(yy), "research_degree": yy,
            "generation": int(gen), "sex": sx}


def _wb_item_base(code) -> str | None:
    m = re.match(r"^(\d{5})", str(code or "").strip())
    return m.group(1) if m else None


def _wb_read_sheet(data: bytes):
    """업로드된 xlsx → (레이아웃, [행dict]). 첫 시트만 읽는다."""
    sheets = _read_xlsx(data)
    if not sheets:
        return "UNKNOWN", []
    _name, grid = sheets[0]
    if not grid:
        return "UNKNOWN", []
    hdr = [_wb_norm_header(x) for x in grid[0]]
    idx = {}
    for i, name in enumerate(hdr):
        if name and name not in idx:
            idx[name] = i
    layout = "V5" if len(hdr) <= 6 else "RAW"

    def g(row, key):
        i = idx.get(key)
        if i is None or i >= len(row):
            return ""
        v = row[i]
        return "" if v is None else str(v).strip()

    out, n = [], 0
    for row in grid[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        n += 1
        out.append({"line_no": n, "serial_no": g(row, "일련번호") or None,
                    "item_cd": g(row, "항목번호"), "headword": g(row, "표제어형"),
                    "dialect_form": g(row, "방언형(기저형)"), "grade": g(row, "사용도/인지도")})
    return layout, out


def api_weather_upload(raw: bytes, ctype: str) -> dict:
    """기상도 원자료 업로드 — 파일 1개 = 제보자 1명. 같은 파일명은 덮어쓴다."""
    mp = _parse_multipart(raw, ctype)
    if not mp["files"]:
        return {"ok": False, "message": "업로드된 파일이 없습니다."}
    results, issues = [], []
    con = weather_db()
    for f in mp["files"]:
        fname = f["filename"]
        meta = _wb_parse_filename(fname)
        if not meta:
            results.append({"ok": False, "fileName": fname,
                            "message": "파일명 규약 불일치 ({지역2}{연차2}{세대2}{성별1}VE.xlsx)"})
            continue
        try:
            layout, rows = _wb_read_sheet(f["data"])
        except Exception as e:
            results.append({"ok": False, "fileName": fname, "message": f"엑셀 읽기 실패: {e}"})
            continue
        kept = [r for r in rows if r["item_cd"]]
        dropped = len(rows) - len(kept)
        if not kept:
            results.append({"ok": False, "fileName": fname,
                            "message": "항목번호가 있는 행이 없습니다. 첫 시트를 확인해 주세요."})
            continue

        # 같은 파일명은 재업로드로 보고 기존 행을 지운다
        old = con.execute("SELECT weather_file_id FROM wb_weather_file WHERE file_nm=?",
                          (fname,)).fetchone()
        replaced = bool(old)
        if old:
            con.execute("DELETE FROM wb_weather_response WHERE weather_file_id=?",
                        (old["weather_file_id"],))
            con.execute("DELETE FROM wb_weather_file WHERE weather_file_id=?",
                        (old["weather_file_id"],))
        cur = con.execute(
            """INSERT INTO wb_weather_file
               (file_nm,region_cd,region_nm,research_year,research_degree,generation,sex,
                row_cnt,item_cnt,src_layout,use_yn,reg_id,reg_dt)
               VALUES (?,?,?,?,?,?,?,?,?,?,'Y','admin',datetime('now'))""",
            (fname, meta["region_cd"], meta["region_nm"], meta["research_year"],
             meta["research_degree"], meta["generation"], meta["sex"], len(kept),
             len({_wb_item_base(r["item_cd"]) for r in kept} - {None}), layout))
        fid = cur.lastrowid
        bad = 0
        batch = []
        for r in kept:
            if r["grade"]:
                r["grade"] = str(r["grade"]).strip().translate(WB_FULLWIDTH)
            gv = "Y" if r["grade"] in WB_VALID_GRADE else "N"
            if r["grade"] and gv == "N" and r["grade"] != "*":
                bad += 1
                issues.append({"fileName": fname, "serialNo": r["serial_no"],
                               "itemCd": r["item_cd"], "grade": r["grade"]})
            batch.append((fid, r["line_no"], r["serial_no"], r["item_cd"],
                          _wb_item_base(r["item_cd"]), r["headword"], r["dialect_form"],
                          r["grade"] or None, gv))
        con.executemany(
            """INSERT INTO wb_weather_response
               (weather_file_id,line_no,serial_no,item_cd,item_base,headword,
                dialect_form,grade,grade_valid_yn,use_yn,reg_dt)
               VALUES (?,?,?,?,?,?,?,?,?,'Y',datetime('now'))""", batch)
        results.append({
            "ok": True, "fileName": fname, "region": meta["region_nm"],
            "year": meta["research_year"], "generation": f'{meta["generation"]}대',
            "sex": "여" if meta["sex"] == "F" else "남", "layout": layout,
            "rows": len(kept), "dropped": dropped, "gradeBad": bad, "replaced": replaced,
        })
    con.commit()
    con.close()
    okn = sum(1 for r in results if r.get("ok"))
    msg = f"{okn}개 파일 적재" + (f" · 실패 {len(results)-okn}개" if okn < len(results) else "")
    if issues:
        msg += f" · 등급 이상값 {len(issues)}건"
    return {"ok": True, "message": msg, "results": results, "issues": issues[:100]}


def _weather_etl():
    """ETL 모듈 — 구조 조립·판정·DB 로더가 모두 거기 한 곳에 있다."""
    import importlib.util
    path = USER_MAP_ROOT / "scripts" / "etl_awareness_region.py"
    if not path.is_file():
        raise FileNotFoundError(f"ETL 스크립트를 찾을 수 없습니다: {path}")
    spec = importlib.util.spec_from_file_location("etl_awareness_region", str(path))
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    return m


_WEATHER_AWARE_CACHE = {}


def _weather_sig(con):
    n = con.execute("SELECT COUNT(*) FROM wb_weather_response").fetchone()[0]
    d = con.execute("SELECT MAX(reg_dt) FROM wb_weather_file").fetchone()[0]
    e = con.execute("SELECT MAX(upt_dt) FROM wb_weather_response").fetchone()[0]
    return (n, d, e)


def api_weather_awareness(qs: dict) -> dict:
    """관리자 목록·응답관리가 읽는 자료 — 전용 테이블 기준.

    프론트(server.py)와 같은 구조를 돌려준다. 목록이 정적 JSON 을 직접 읽던 것을
    이 API 로 바꾸면, 업로드 직후 내보내기를 하지 않아도 화면이 최신 상태가 된다.

    year: 조사 연차. 연차가 섞이면 '제보자 38명' 같은 수치가 서로 다른 조사의 합이 되므로
          화면 필터가 아니라 원천을 가르는 조건으로 다룬다.
    """
    def q1(k, d=""):
        v = qs.get(k)
        return (v[0] if isinstance(v, list) else v) or d

    year = re.sub(r"\D", "", str(q1("year")))[-2:]
    con = weather_db()
    try:
        sig = _weather_sig(con) + (year,)
    finally:
        con.close()
    if _WEATHER_AWARE_CACHE.get(sig) is not None:
        return _WEATHER_AWARE_CACHE[sig]
    etl = _weather_etl()
    recs, nfiles = etl.load_records_from_db(str(WEATHER_DB), year)
    out = etl.build_output(recs, nfiles)
    etl.fill_db_qc(out, str(WEATHER_DB), year)
    _WEATHER_AWARE_CACHE.clear()
    _WEATHER_AWARE_CACHE[sig] = out
    return out


def _weather_file_by_panel_id(con, panel_id: str):
    """명부 id (GW2420M = 지역2+연차2+세대2+성별1) → wb_weather_file 한 건."""
    m = re.match(r"^([A-Z]{2})(\d{2})(\d{2})([MF])$", str(panel_id or "").strip())
    if not m:
        return None
    rg, yy, gen, sx = m.groups()
    return con.execute(
        """SELECT * FROM wb_weather_file
           WHERE region_cd=? AND research_degree=? AND generation=? AND sex=? AND use_yn='Y'""",
        (rg, yy, int(gen), sx)).fetchone()


def api_weather_export(body: dict) -> dict:
    """전용 테이블 → 정적 JSON(data/processed/awareness_by_region.json) 다시 뽑기.

    정적 호스팅(GitHub Pages)에는 API 가 없어 프론트가 이 파일로 되돌아간다.
    관리자에서 원자료를 올린 뒤 이걸 눌러야 정적 배포본에도 반영된다.
    구조 조립은 scripts/etl_awareness_region.py 의 build_output() 한 곳에서만 한다.
    """
    import importlib.util

    etl_path = USER_MAP_ROOT / "scripts" / "etl_awareness_region.py"
    if not etl_path.is_file():
        return {"ok": False, "message": f"ETL 스크립트를 찾을 수 없습니다: {etl_path}"}
    spec = importlib.util.spec_from_file_location("etl_awareness_region", str(etl_path))
    etl = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(etl)

    out_path = USER_MAP_ROOT / "data" / "processed" / "awareness_by_region.json"
    # 직전 내용을 .bak 으로 남긴다 — 되돌릴 수 있게
    if out_path.is_file():
        try:
            out_path.with_suffix(out_path.suffix + ".bak").write_bytes(out_path.read_bytes())
        except Exception:
            pass

    recs, nfiles = etl.load_records_from_db(str(WEATHER_DB))
    out = etl.build_output(recs, nfiles)
    etl.fill_db_qc(out, str(WEATHER_DB))
    text = json.dumps(out, ensure_ascii=False, indent=1)
    tmp = out_path.with_name(out_path.name + ".tmp")
    tmp.write_text(text, encoding="utf-8")
    os.replace(str(tmp), str(out_path))

    qc = out["meta"]["qc"]
    return {"ok": True, "path": str(out_path), "bytes": len(text.encode("utf-8")),
            "informants": nfiles, "items": len(out["items"]),
            "cells": qc.get("cells"), "states": qc.get("states"),
            "message": (f"정적 JSON 내보냄 — 제보자 {nfiles}명 · 항목 {len(out['items'])}개 · "
                        f"{len(text.encode('utf-8'))/1024:.0f} KB")}


def api_weather_responses_save(body: dict) -> dict:
    """응답 목록 저장 — 화면 한 줄 = wb_weather_response 한 행. 고친 줄만 한 번에 받는다.

    화면이 자료를 읽은 뒤 누군가 재업로드를 하면 response_id 가 다른 행을 가리키게 된다.
    그래서 rid 만 믿지 않고 (파일명, 행번호)를 함께 받아 서로 맞는지 확인한 뒤에만 고친다.
    어긋난 줄은 고치지 않고 이유를 돌려준다 — 엉뚱한 응답의 등급을 덮어쓰는 사고가 실제로 있었다.
    """
    rows = body.get("rows") or []
    editor = str(body.get("editor") or "관리자").strip() or "관리자"
    if not isinstance(rows, list) or not rows:
        return {"ok": False, "message": "저장할 내용이 없습니다."}

    con = weather_db()
    saved, skipped = 0, []
    now = time.strftime("%Y-%m-%d %H:%M:%S")
    try:
        for r in rows:
            r = r or {}
            rid = r.get("rid")
            try:
                rid = int(rid)
            except (TypeError, ValueError):
                skipped.append({"rid": r.get("rid"), "why": "행 번호가 없습니다."})
                continue

            cur = con.execute(
                """SELECT r.response_id, r.line_no, r.dialect_form, r.headword, r.grade,
                          r.item_base, r.weather_file_id, f.file_nm
                   FROM wb_weather_response r JOIN wb_weather_file f USING(weather_file_id)
                   WHERE r.response_id=?""", (rid,)).fetchone()
            if not cur:
                skipped.append({"rid": rid, "why": "그 응답이 더는 없습니다. 새로 고침 뒤 다시 시도해 주세요."})
                continue
            # 화면이 읽은 자리와 지금 자리가 같은지 — 재업로드로 어긋났으면 여기서 걸린다
            if str(r.get("file") or "") and (r.get("file") != cur["file_nm"]
                                             or int(r.get("lineNo") or -1) != cur["line_no"]):
                skipped.append({"rid": rid, "why": "자료가 바뀌었습니다(재업로드). 새로 고침 뒤 다시 시도해 주세요."})
                continue

            grade = r.get("grade")
            grade = "" if grade in (None, "") else str(grade).strip().translate(WB_FULLWIDTH)
            if grade and grade not in WB_VALID_GRADE:
                skipped.append({"rid": rid, "why": "사용도/인지도는 1~4 여야 합니다."})
                continue
            form = str(r.get("form") or "").strip()

            before_form = (cur["dialect_form"] or "").strip() or (cur["headword"] or "").strip()
            before_grade = (cur["grade"] or "").strip()
            if form == before_form and grade == before_grade:
                continue                                    # 바뀐 게 없다

            # 원자료가 '방언형 칸은 비고 표제어형 칸에 어형'인 행(조사자 제시형)이면
            # 그 칸을 그대로 고쳐야 화면과 엑셀의 대응이 유지된다.
            col = "dialect_form" if (cur["dialect_form"] or "").strip() else "headword"
            con.execute(
                "UPDATE wb_weather_response SET %s=?, grade=?, grade_valid_yn=?, upt_dt=? "
                "WHERE response_id=?" % col,
                (form or None, grade or None, "Y" if grade in WB_VALID_GRADE else "N", now, rid))
            saved += 1
        con.commit()
    finally:
        con.close()

    _WEATHER_AWARE_CACHE.clear()
    msg = "%d건 저장" % saved
    if skipped:
        msg += " · %d건 건너뜀" % len(skipped)
    return {"ok": True, "saved": saved, "skipped": skipped, "message": msg}


def api_weather_responses(qs: dict) -> dict:
    """항목 하나의 응답 행을 원자료 그대로 내려준다 — 관리자 응답 목록 화면용.

    집계(build_output)는 제보자 단위로 '최선 등급' 하나만 남기므로 어형이 여럿인 응답이
    가려진다. 편집은 엑셀 한 행 = 화면 한 행이어야 하므로 여기서는 접지 않는다.
    정렬은 일련번호가 아니라 지역·세대·성별·행번호 순 — 일련번호는 15%가 비어 있고
    중복도 있어 정렬 키로 쓸 수 없다."""
    import sqlite3

    item = re.sub(r"[^0-9]", "", str(qs.get("item", [""])[0] if isinstance(qs.get("item"), list) else qs.get("item") or ""))
    if not item:
        return {"ok": False, "message": "항목코드가 필요합니다."}
    if not WEATHER_DB.exists():
        return {"ok": False, "message": "적재된 자료가 없습니다."}

    order = {c: i for i, c in enumerate(WB_REGION_ORDER)}
    con = sqlite3.connect(str(WEATHER_DB))
    con.row_factory = sqlite3.Row
    try:
        rows = con.execute(
            """SELECT r.response_id, r.line_no, r.serial_no, r.item_cd, r.headword,
                      r.dialect_form, r.grade, r.grade_valid_yn, r.upt_dt,
                      f.file_nm, f.region_cd, f.region_nm, f.research_degree,
                      f.generation, f.sex
               FROM wb_weather_response r JOIN wb_weather_file f USING(weather_file_id)
               WHERE r.item_base=? AND r.use_yn='Y' AND f.use_yn='Y'""", (item,)).fetchall()
    finally:
        con.close()

    out = []
    for r in rows:
        form = (r["dialect_form"] or "").strip()
        pres = (r["headword"] or "").strip()
        out.append({
            "rid": r["response_id"],
            "file": r["file_nm"],
            "lineNo": r["line_no"],           # (file, lineNo) 가 편집 열쇠
            "serialNo": r["serial_no"] or "",
            "itemCd": r["item_cd"] or "",
            "region": r["region_cd"],
            "regionNm": r["region_nm"] or WB_REGION_NAMES.get(r["region_cd"], r["region_cd"]),
            "year": r["research_degree"] or "",
            "age": r["generation"],
            "sex": r["sex"],
            "headword": pres,
            "form": form,
            # 방언형 칸이 비면 표제어형 칸의 값을 쓴다 (ETL 과 같은 규칙)
            "shown": form if form and form != "*" else pres,
            "grade": (r["grade"] or "") if r["grade_valid_yn"] == "Y" else "",
            "gradeRaw": r["grade"] or "",
            # 관리자가 고친 행. 화면의 '관리자가 고침' 검색이 저장된 것까지 찾으려면 필요하다
            "edited": bool(r["upt_dt"]),
        })
    out.sort(key=lambda x: (order.get(x["region"], 99), x["age"] or 0,
                            x["sex"] or "", x["file"], x["lineNo"]))
    return {"ok": True, "item": item, "total": len(out), "rows": out}


def api_weather_files(qs: dict) -> dict:
    """적재된 원자료 파일 목록."""
    con = weather_db()
    rows = [dict(r) for r in con.execute(
        """SELECT weather_file_id,file_nm,region_cd,region_nm,research_year,generation,
                  sex,row_cnt,item_cnt,src_layout,use_yn,reg_dt
           FROM wb_weather_file ORDER BY region_cd, generation, sex""")]
    # 관리자가 고친 행 — 재업로드하면 엑셀 값으로 되돌아가므로 화면이 미리 경고해야 한다
    edited = dict(con.execute(
        """SELECT weather_file_id, COUNT(*) FROM wb_weather_response
           WHERE upt_dt IS NOT NULL GROUP BY weather_file_id"""))
    resp = con.execute("SELECT COUNT(*) c FROM wb_weather_response").fetchone()
    bad = con.execute(
        """SELECT COUNT(*) c FROM wb_weather_response
           WHERE grade IS NOT NULL AND grade<>'*' AND grade_valid_yn='N'""").fetchone()
    con.close()
    for r in rows:
        r["sexNm"] = "여" if r["sex"] == "F" else "남"
        r["genNm"] = f'{r["generation"]}대'
        r["editedCnt"] = edited.get(r["weather_file_id"], 0)
    return {"ok": True, "total": len(rows), "list": rows,
            "responseCnt": resp["c"], "gradeBadCnt": bad["c"]}


def api_weather_delete(body: dict) -> dict:
    """적재 파일 삭제 (응답 행까지)."""
    ids = body.get("ids") or ([body.get("id")] if body.get("id") else [])
    ids = [str(x).strip() for x in ids if str(x).strip()]
    if not ids:
        return {"ok": False, "message": "삭제할 파일을 선택해주세요."}
    con = weather_db()
    ph = ",".join("?" * len(ids))
    con.execute(f"DELETE FROM wb_weather_response WHERE weather_file_id IN ({ph})", ids)
    cur = con.execute(f"DELETE FROM wb_weather_file WHERE weather_file_id IN ({ph})", ids)
    n = cur.rowcount
    con.commit()
    con.close()
    return {"ok": True, "deleted": n,
            "message": f"{n}건 삭제되었습니다. 집계 재계산이 필요합니다."}


class Handler(http.server.SimpleHTTPRequestHandler):
    def translate_path(self, path):
        # strip query for file mapping
        path_only = path.split("?", 1)[0]
        fs = super().translate_path(path_only)
        stripped = fs.rstrip()
        if stripped.endswith(".do"):
            html = stripped[:-3] + ".html"
            if os.path.exists(html):
                return html
        return fs

    def _send_json(self, obj, status=200):
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)


    def _serve_file_range(self, fs: Path, ctype: str):
        """정적 파일을 HTTP Range 지원으로 스트리밍 (오디오 탐색용)."""
        try:
            size = fs.stat().st_size
        except OSError:
            self.send_error(404, "Not Found")
            return
        rng = self.headers.get("Range")
        start, end = 0, size - 1
        status = 200
        if rng:
            m = re.match(r"bytes=(\d*)-(\d*)", rng.strip())
            if m:
                g1, g2 = m.group(1), m.group(2)
                if g1:
                    start = int(g1)
                    end = int(g2) if g2 else size - 1
                elif g2:  # suffix: 마지막 N바이트
                    start = max(0, size - int(g2))
                if start > end or start >= size:
                    self.send_response(416)
                    self.send_header("Content-Range", f"bytes */{size}")
                    self.end_headers()
                    return
                status = 206
        length = end - start + 1
        self.send_response(status)
        self.send_header("Content-Type", ctype)
        self.send_header("Accept-Ranges", "bytes")
        self.send_header("Content-Length", str(length))
        if status == 206:
            self.send_header("Content-Range", f"bytes {start}-{end}/{size}")
        self.send_header("Cache-Control", "no-cache")
        self.end_headers()
        if self.command == "HEAD":
            return
        with fs.open("rb") as f:
            f.seek(start)
            remaining = length
            chunk = 64 * 1024
            while remaining > 0:
                buf = f.read(min(chunk, remaining))
                if not buf:
                    break
                try:
                    self.wfile.write(buf)
                except (BrokenPipeError, ConnectionResetError):
                    break
                remaining -= len(buf)

    def do_POST(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        length = int(self.headers.get("Content-Length") or 0)
        raw = self.rfile.read(length) if length else b""
        ctype = self.headers.get("Content-Type", "")

        # 멀티파트 업로드(등록 ①)는 JSON 파싱 전에 처리
        if path in (
            "/mariadb/neibis-api/survey/std-vocab/bulk",
            "/mariadb/neibis-api/std-vocab/bulk",
        ):
            try:
                self._send_json(api_std_vocab_bulk(raw, ctype))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/weather/upload",
            "/mariadb/neibis-api/v1/weather/upload",
        ):
            try:
                self._send_json(api_weather_upload(raw, ctype))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/oral/upload",
            "/mariadb/neibis-api/v1/oral/upload",
            "/mariadb/neibis-api/survey/oral/upload",
        ):
            try:
                self._send_json(api_oral_upload(raw, ctype))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        try:
            body = json.loads(raw.decode("utf-8") or "{}")
        except Exception:
            body = {}

        if path in (
            "/mariadb/neibis-api/archive/literature/save",
            "/mariadb/neibis-api/v1/archive/literature/save",
            "/mariadb/neibis-api/literature/save",
        ):
            try:
                self._send_json(api_literature_save(body))
            except FileNotFoundError as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/archive/literature/delete",
            "/mariadb/neibis-api/v1/archive/literature/delete",
            "/mariadb/neibis-api/literature/delete",
        ):
            try:
                self._send_json(api_literature_delete(body))
            except FileNotFoundError as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/board/post/update",
            "/mariadb/neibis-api/v1/board/post/update",
        ):
            try:
                self._send_json(api_board_post_update(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/board/post/delete",
            "/mariadb/neibis-api/v1/board/post/delete",
        ):
            try:
                self._send_json(api_board_post_delete(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/board/answer/save",
            "/mariadb/neibis-api/v1/board/answer/save",
        ):
            try:
                self._send_json(api_board_answer_save(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/system/user/auth",
            "/mariadb/neibis-api/v1/system/user/auth",
        ):
            try:
                self._send_json(api_user_set_auth(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/system/user/unlock",
            "/mariadb/neibis-api/v1/system/user/unlock",
        ):
            try:
                self._send_json(api_user_unlock(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/system/user/reset-pw",
            "/mariadb/neibis-api/v1/system/user/reset-pw",
        ):
            try:
                self._send_json(api_user_reset_pw(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/weather/responses/save",
            "/mariadb/neibis-api/v1/weather/responses/save",
        ):
            try:
                self._send_json(api_weather_responses_save(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/weather/export",
            "/mariadb/neibis-api/v1/weather/export",
        ):
            try:
                self._send_json(api_weather_export(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/weather/delete",
            "/mariadb/neibis-api/v1/weather/delete",
        ):
            try:
                self._send_json(api_weather_delete(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/wordcard/save",
            "/mariadb/neibis-api/v1/wordcard/save",
        ):
            try:
                self._send_json(api_wordcard_save(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/wordcard/delete",
            "/mariadb/neibis-api/v1/wordcard/delete",
        ):
            try:
                self._send_json(api_wordcard_delete(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/symbol/save",
            "/mariadb/neibis-api/v1/symbol/save",
        ):
            try:
                self._send_json(api_symbol_save(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/symbol/delete",
            "/mariadb/neibis-api/v1/symbol/delete",
        ):
            try:
                self._send_json(api_symbol_delete(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/headword/save",
            "/mariadb/neibis-api/v1/headword/save",
        ):
            try:
                self._send_json(api_headword_save(body))
            except FileNotFoundError as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/headword/status",
            "/mariadb/neibis-api/v1/headword/status",
        ):
            try:
                self._send_json(api_headword_status(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/dialect/save",
            "/mariadb/neibis-api/v1/dialect/save",
            "/mariadb/neibis-api/headword/dialect/save",
        ):
            try:
                self._send_json(api_dialect_save(body))
            except FileNotFoundError as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/dialect/region/save",
            "/mariadb/neibis-api/v1/dialect/region/save",
        ):
            try:
                self._send_json(api_dialect_region_save(body))
            except FileNotFoundError as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/oral/save-meta",
            "/mariadb/neibis-api/v1/oral/save-meta",
            "/mariadb/neibis-api/survey/oral/save-meta",
        ):
            try:
                self._send_json(api_oral_save_meta(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/oral/save-lines",
            "/mariadb/neibis-api/v1/oral/save-lines",
            "/mariadb/neibis-api/survey/oral/save-lines",
        ):
            try:
                self._send_json(api_oral_save_lines(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/oral/save-raw",
            "/mariadb/neibis-api/v1/oral/save-raw",
            "/mariadb/neibis-api/survey/oral/save-raw",
        ):
            try:
                self._send_json(api_oral_save_raw(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/oral/create",
            "/mariadb/neibis-api/v1/oral/create",
            "/mariadb/neibis-api/survey/oral/create",
        ):
            try:
                self._send_json(api_oral_create(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/source/save",
            "/mariadb/neibis-api/v1/source/save",
        ):
            try:
                self._send_json(api_source_save(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/source/delete",
            "/mariadb/neibis-api/v1/source/delete",
        ):
            try:
                self._send_json(api_source_delete(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in ("/mariadb/neibis-api/survey/save", "/mariadb/neibis-api/v1/survey/save"):
            try:
                self._send_json(api_survey_save(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in ("/mariadb/neibis-api/survey/delete", "/mariadb/neibis-api/v1/survey/delete"):
            try:
                self._send_json(api_survey_delete(body))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        self.send_error(404, "Not Found")

    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        qs = urllib.parse.parse_qs(parsed.query)

        if path in (
            "/mariadb/neibis-api/board/post/detail",
            "/mariadb/neibis-api/v1/board/post/detail",
        ):
            try:
                self._send_json(api_board_post_detail(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/board/post/list",
            "/mariadb/neibis-api/v1/board/post/list",
        ):
            try:
                self._send_json(api_board_post_list(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/system/user/list",
            "/mariadb/neibis-api/v1/system/user/list",
        ):
            try:
                self._send_json(api_user_list(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/stats/openapi",
            "/mariadb/neibis-api/v1/stats/openapi",
        ):
            try:
                self._send_json(api_openapi_usage_list(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/stats/openapi/excel",
            "/mariadb/neibis-api/v1/stats/openapi/excel",
        ):
            try:
                data, fname = build_openapi_excel(qs)
                enc = urllib.parse.quote(fname)
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", "attachment; filename*=UTF-8''" + enc)
                self.send_header("Content-Length", str(len(data)))
                self.send_header("Cache-Control", "no-store")
                self.end_headers()
                self.wfile.write(data)
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/system/user/detail",
            "/mariadb/neibis-api/v1/system/user/detail",
        ):
            try:
                self._send_json(api_user_detail(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/weather/awareness",
            "/mariadb/neibis-api/v1/weather/awareness",
        ):
            try:
                self._send_json(api_weather_awareness(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/weather/responses",
            "/mariadb/neibis-api/v1/weather/responses",
        ):
            try:
                self._send_json(api_weather_responses(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/weather/files",
            "/mariadb/neibis-api/v1/weather/files",
        ):
            try:
                self._send_json(api_weather_files(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/wordcard/list",
            "/mariadb/neibis-api/v1/wordcard/list",
        ):
            try:
                self._send_json(api_wordcard_list(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/wordcard/detail",
            "/mariadb/neibis-api/v1/wordcard/detail",
        ):
            try:
                self._send_json(api_wordcard_detail(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/wordcard/meta",
            "/mariadb/neibis-api/v1/wordcard/meta",
        ):
            try:
                self._send_json(api_wordcard_meta(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/symbol/detail",
            "/mariadb/neibis-api/v1/symbol/detail",
        ):
            try:
                self._send_json(api_symbol_detail(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/symbol/excel",
            "/mariadb/neibis-api/v1/symbol/excel",
        ):
            try:
                data, fname = build_symbol_excel(qs)
                enc = urllib.parse.quote(fname)
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", "attachment; filename*=UTF-8''" + enc)
                self.send_header("Content-Length", str(len(data)))
                self.send_header("Cache-Control", "no-store")
                self.end_headers()
                self.wfile.write(data)
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/symbol/list",
            "/mariadb/neibis-api/v1/symbol/list",
        ):
            try:
                self._send_json(api_symbol_list(qs))
            except FileNotFoundError as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/vocab/list",
            "/mariadb/neibis-api/v1/vocab/list",
            "/mariadb/neibis-api/survey/vocab/list",
        ):
            try:
                self._send_json(api_vocab_list(qs))
            except FileNotFoundError as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/archive/literature/list",
            "/mariadb/neibis-api/v1/archive/literature/list",
            "/mariadb/neibis-api/literature/list",
        ):
            try:
                self._send_json(api_literature_list(qs))
            except FileNotFoundError as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/archive/literature/detail",
            "/mariadb/neibis-api/v1/archive/literature/detail",
            "/mariadb/neibis-api/literature/detail",
        ):
            try:
                self._send_json(api_literature_detail(qs))
            except FileNotFoundError as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/vocab/detail",
            "/mariadb/neibis-api/v1/vocab/detail",
            "/mariadb/neibis-api/survey/vocab/detail",
        ):
            try:
                self._send_json(api_vocab_detail(qs))
            except FileNotFoundError as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/symbol/catalog",
            "/mariadb/neibis-api/v1/symbol/catalog",
        ):
            try:
                self._send_json(api_symbol_catalog(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/headword/list",
            "/mariadb/neibis-api/v1/headword/list",
            "/mariadb/neibis-api/map/dialect/list",
        ):
            try:
                self._send_json(api_headword_list(qs))
            except FileNotFoundError as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/headword/detail",
            "/mariadb/neibis-api/v1/headword/detail",
        ):
            try:
                self._send_json(api_headword_detail(qs))
            except FileNotFoundError as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/headword/search",
            "/mariadb/neibis-api/v1/headword/search",
        ):
            try:
                self._send_json(api_headword_search_by_name(qs))
            except FileNotFoundError as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/dialect/list",
            "/mariadb/neibis-api/v1/dialect/list",
            "/mariadb/neibis-api/headword/dialect/list",
        ):
            try:
                self._send_json(api_dialect_list(qs))
            except FileNotFoundError as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/oral/list",
            "/mariadb/neibis-api/v1/oral/list",
            "/mariadb/neibis-api/survey/oral/list",
        ):
            try:
                self._send_json(api_oral_list(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/oral/topics",
            "/mariadb/neibis-api/v1/oral/topics",
            "/mariadb/neibis-api/survey/oral/topics",
        ):
            try:
                self._send_json(api_oral_topics(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/oral/detail",
            "/mariadb/neibis-api/v1/oral/detail",
            "/mariadb/neibis-api/survey/oral/detail",
        ):
            try:
                self._send_json(api_oral_detail(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/oral/meta",
            "/mariadb/neibis-api/v1/oral/meta",
            "/mariadb/neibis-api/survey/oral/meta",
        ):
            try:
                self._send_json(api_oral_meta(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/oral/raw",
            "/mariadb/neibis-api/v1/oral/raw",
            "/mariadb/neibis-api/survey/oral/raw",
        ):
            try:
                self._send_json(api_oral_raw(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in (
            "/mariadb/neibis-api/oral/source/search",
            "/mariadb/neibis-api/v1/oral/source/search",
            "/mariadb/neibis-api/survey/oral/source/search",
        ):
            try:
                self._send_json(api_oral_source_search(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in ("/mariadb/neibis-api/source/list", "/mariadb/neibis-api/v1/source/list"):
            try:
                self._send_json(api_source_list(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in ("/mariadb/neibis-api/source/regions", "/mariadb/neibis-api/v1/source/regions"):
            try:
                self._send_json(api_source_regions(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in ("/mariadb/neibis-api/source/detail", "/mariadb/neibis-api/v1/source/detail"):
            try:
                self._send_json(api_source_detail(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in ("/mariadb/neibis-api/survey/list", "/mariadb/neibis-api/v1/survey/list"):
            try:
                self._send_json(api_survey_list(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in ("/mariadb/neibis-api/survey/legacy-list", "/mariadb/neibis-api/v1/survey/legacy-list"):
            try:
                self._send_json(api_survey_legacy_list(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in ("/mariadb/neibis-api/survey/legacy-detail", "/mariadb/neibis-api/v1/survey/legacy-detail"):
            try:
                self._send_json(api_survey_legacy_detail(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in ("/mariadb/neibis-api/survey/legacy-respondents", "/mariadb/neibis-api/v1/survey/legacy-respondents"):
            try:
                self._send_json(api_survey_legacy_respondents(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in ("/mariadb/neibis-api/survey/legacy-answer-detail", "/mariadb/neibis-api/v1/survey/legacy-answer-detail"):
            try:
                self._send_json(api_survey_legacy_answer_detail(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in ("/mariadb/neibis-api/survey/detail", "/mariadb/neibis-api/v1/survey/detail"):
            try:
                self._send_json(api_survey_detail(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in ("/mariadb/neibis-api/survey/respondents", "/mariadb/neibis-api/v1/survey/respondents"):
            try:
                self._send_json(api_survey_respondents(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in ("/mariadb/neibis-api/survey/answer-detail", "/mariadb/neibis-api/v1/survey/answer-detail"):
            try:
                self._send_json(api_survey_answer_detail(qs))
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        if path in ("/mariadb/neibis-api/survey/excel", "/mariadb/neibis-api/v1/survey/excel"):
            try:
                sid = (qs.get("id") or qs.get("surveyNo") or [""])[0].strip()
                data, fname = build_survey_excel(sid)
                if not data:
                    self._send_json({"ok": False, "message": "설문을 찾을 수 없습니다."}, 404)
                    return
                enc = urllib.parse.quote(fname)
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", "attachment; filename*=UTF-8''" + enc)
                self.send_header("Content-Length", str(len(data)))
                self.send_header("Cache-Control", "no-store")
                self.end_headers()
                self.wfile.write(data)
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)}, 500)
            return

        # 구술발화 원천 미디어/전사파일 서빙 (Range 지원)
        if path.startswith("/oral-media/"):
            name = urllib.parse.unquote(path[len("/oral-media/"):])
            m = re.fullmatch(r"([A-Za-z0-9_\-]+)\.(wav|eaf)", name)
            if not m:
                self.send_error(404, "Not Found")
                return
            fs = _find_oral_media(m.group(1), m.group(2))
            if fs is None:
                self.send_error(404, "Not Found")
                return
            ctype = "audio/wav" if m.group(2) == "wav" else "application/xml; charset=utf-8"
            self._serve_file_range(fs, ctype)
            return

        if path == "/mariadb/neibis-api/health":
            self._send_json(
                {
                    "ok": True,
                    "db": str(DB_PATH),
                    "dbExists": DB_PATH.is_file(),
                }
            )
            return

        # 사용자 지도 자산 (OpenLayers 데이터·스크립트·symbol_mask)
        if path.startswith("/user-map/"):
            rel = path[len("/user-map/"):]
            # path traversal guard
            rel = rel.lstrip("/").replace("\\", "/")
            if ".." in rel.split("/"):
                self.send_error(403, "Forbidden")
                return
            fs = (USER_MAP_ROOT / rel).resolve()
            try:
                fs.relative_to(USER_MAP_ROOT.resolve())
            except ValueError:
                self.send_error(403, "Forbidden")
                return
            if not fs.is_file():
                self.send_error(404, "Not Found")
                return
            ctype = "application/octet-stream"
            if fs.suffix == ".js":
                ctype = "application/javascript; charset=utf-8"
            elif fs.suffix == ".css":
                ctype = "text/css; charset=utf-8"
            elif fs.suffix == ".png":
                ctype = "image/png"
            elif fs.suffix == ".json":
                ctype = "application/json; charset=utf-8"
            data = fs.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", ctype)
            self.send_header("Content-Length", str(len(data)))
            self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
            self.end_headers()
            self.wfile.write(data)
            return

        # 1차: /Users/aaa/inseq/korean 루트 디렉토리 정적 파일 우선 확인
        if path != "/" and not path.startswith("/mariadb/"):
            rel_path = path.lstrip("/")
            alt_file = (USER_MAP_ROOT / rel_path).resolve()
            try:
                alt_file.relative_to(USER_MAP_ROOT.resolve())
                if alt_file.is_file():
                    ctype = self.guess_type(str(alt_file))
                    data = alt_file.read_bytes()
                    self.send_response(200)
                    self.send_header("Content-Type", ctype)
                    self.send_header("Content-Length", str(len(data)))
                    self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
                    self.end_headers()
                    self.wfile.write(data)
                    return
            except Exception:
                pass

        # 2차: neibis-cms 정적 파일 및 default handler
        return super().do_GET()

    def end_headers(self):
        # 개발 편의: 정적 파일(HTML/JS/CSS)에도 no-cache 부여 (이미 Cache-Control 지정된 응답은 유지)
        try:
            buf = getattr(self, "_headers_buffer", None) or []
            if not any(b"cache-control" in line.lower() for line in buf):
                self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
        except Exception:
            pass
        super().end_headers()

    def log_message(self, fmt, *args):
        # quieter default
        sys_stderr = getattr(self, "", None)
        print("[%s] %s" % (self.log_date_time_string(), fmt % args))


class Server(socketserver.ThreadingMixIn, http.server.HTTPServer):
    daemon_threads = True
    allow_reuse_address = True


if __name__ == "__main__":
    handler = functools.partial(Handler, directory=str(ROOT))
    with Server(("", PORT), handler) as httpd:
        print(f"NEIBIS prototype serving {ROOT} at http://localhost:{PORT}")
        print(f"DB: {DB_PATH} (exists={DB_PATH.is_file()})")
        print("API: GET /mariadb/neibis-api/symbol/list?searchValue=&page=1&pageSize=10")
        print("API: GET /mariadb/neibis-api/vocab/list?page=1&pageSize=10")
        httpd.serve_forever()
